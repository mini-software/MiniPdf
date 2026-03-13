import openpyxl
wb = openpyxl.load_workbook('xlsx/Wedding_timeline_planner1_copy.xlsx')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'Sheet: {name}')
    print(f'  Dimensions: {ws.dimensions}')
    print(f'  Min/Max row: {ws.min_row}-{ws.max_row}, col: {ws.min_column}-{ws.max_column}')
    
    # Check column widths
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width:
            print(f'  Col {col_letter}: width={dim.width}, hidden={dim.hidden}')
    
    # Check merged cells
    for mc in ws.merged_cells.ranges:
        print(f'  Merged: {mc}')
    
    # Check page setup
    ps = ws.page_setup
    print(f'  Page setup: orientation={ps.orientation}, paperSize={ps.paperSize}, fitToWidth={ps.fitToWidth}, fitToHeight={ps.fitToHeight}')
    if ws.sheet_properties and ws.sheet_properties.pageSetUpPr:
        print(f'  fitToPage: {ws.sheet_properties.pageSetUpPr.fitToPage}')
    
    # Check row heights
    for row_num, dim in ws.row_dimensions.items():
        if dim.height:
            print(f'  Row {row_num}: height={dim.height}')
    
    # Print first 50 rows of content
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=False):
        vals = []
        for cell in row:
            if cell.value is not None:
                font_info = ''
                if cell.font:
                    font_info = f'(font={cell.font.name},sz={cell.font.size},b={cell.font.bold})'
                wrap = cell.alignment.wrapText if cell.alignment else None
                v = repr(cell.value)[:80]
                vals.append(f'{cell.coordinate}={v} wrap={wrap} {font_info}')
        if vals:
            sep = '  |  '
            print(f'  Row {row[0].row}: ' + sep.join(vals))
