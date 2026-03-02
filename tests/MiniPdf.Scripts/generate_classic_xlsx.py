"""
Generate 90 classic .xlsx files for testing Excel-to-PDF conversion.
Each file corresponds to a test case in ClassicExcelToPdfTests.cs.

Cases 61-90 include embedded images to exercise MiniPdf image rendering.

Usage:
    pip install openpyxl pillow
    python generate_classic_xlsx.py

Output directory: ./output/
"""

import io
import os
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False


def _make_jpeg_bytes(width: int, height: int, color: tuple) -> bytes:
    """Return raw JPEG bytes for a solid-color image (requires Pillow)."""
    img = PILImage.new("RGB", (width, height), color)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


def _add_image(ws, jpeg_bytes: bytes, anchor: str, width_px: int = 120, height_px: int = 90):
    """Embed a JPEG image into the worksheet at the given cell anchor."""
    if not HAS_PIL:
        return
    from openpyxl.drawing.image import Image as XLImage
    buf = io.BytesIO(jpeg_bytes)
    xl_img = XLImage(buf)
    xl_img.width = width_px
    xl_img.height = height_px
    xl_img.anchor = anchor
    ws.add_image(xl_img)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def save(wb: Workbook, filename: str):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  [OK] {filename}")


# ── 01. Basic table with headers ────────────────────────────────────────
def classic01_basic_table_with_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "London"])
    ws.append(["Charlie", 35, "Tokyo"])
    ws.append(["Diana", 28, "Paris"])
    save(wb, "classic01_basic_table_with_headers.xlsx")


# ── 02. Multiple worksheets ─────────────────────────────────────────────
def classic02_multiple_worksheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Quarter", "Revenue"])
    ws1.append(["Q1", 100])
    ws1.append(["Q2", 200])
    ws1.append(["Q3", 350])
    ws1.append(["Q4", 480])

    ws2 = wb.create_sheet("Costs")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 500])
    ws2.append(["Salary", 3000])
    ws2.append(["Utilities", 200])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Revenue", 1130])
    ws3.append(["Total Costs", 3700])
    ws3.append(["Net", -2570])
    save(wb, "classic02_multiple_worksheets.xlsx")


# ── 03. Empty workbook (no data rows) ───────────────────────────────────
def classic03_empty_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # No data at all
    save(wb, "classic03_empty_workbook.xlsx")


# ── 04. Single cell ─────────────────────────────────────────────────────
def classic04_single_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    save(wb, "classic04_single_cell.xlsx")


# ── 05. Wide table (26 columns A–Z) ─────────────────────────────────────
def classic05_wide_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = list(string.ascii_uppercase)  # A-Z
    ws.append(headers)
    for row_idx in range(1, 6):
        ws.append([f"{ch}{row_idx}" for ch in headers])
    save(wb, "classic05_wide_table.xlsx")


# ── 06. Tall table (200 rows → multi-page) ──────────────────────────────
def classic06_tall_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Row#", "Value", "Description"])
    for i in range(1, 201):
        ws.append([f"Row{i}", f"Val{i}", f"This is the description for row number {i}"])
    save(wb, "classic06_tall_table.xlsx")


# ── 07. Numbers only ────────────────────────────────────────────────────
def classic07_numbers_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([1.0, 2.0, 3.0])
    ws.append([4.0, 5.0, 6.0])
    ws.append([7.0, 8.0, 9.0])
    ws.append([10.0, 100.0, 1000.0])
    save(wb, "classic07_numbers_only.xlsx")


# ── 08. Mixed text and numbers ──────────────────────────────────────────
def classic08_mixed_text_and_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Amount"])
    ws.append(["Item", 10.5])
    ws.append(["Tax", 0.08])
    ws.append(["Total", 10.58])
    ws.append(["Discount", -1.5])
    ws.append(["Final", 9.08])
    save(wb, "classic08_mixed_text_and_numbers.xlsx")


# ── 09. Long text content ───────────────────────────────────────────────
def classic09_long_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Long Text Column"])
    ws.append(["X" * 500])
    ws.append(["A" * 300 + " " + "B" * 200])
    ws.append(["Short"])
    ws.append(["Y" * 1000])
    save(wb, "classic09_long_text.xlsx")


# ── 10. Special XML characters ──────────────────────────────────────────
def classic10_special_xml_characters():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Special Characters"])
    ws.append(["A&B"])
    ws.append(["<tag>"])
    ws.append(['"quoted"'])
    ws.append(["it's"])
    ws.append(["Tom & Jerry < Batman > Superman"])
    ws.append(['He said "hello" & she replied \'hi\''])
    save(wb, "classic10_special_xml_characters.xlsx")


# ── 11. Sparse rows (gaps between data rows) ────────────────────────────
def classic11_sparse_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="First")
    ws.cell(row=5, column=1, value="Fifth")
    ws.cell(row=10, column=1, value="Tenth")
    ws.cell(row=20, column=1, value="Twentieth")
    ws.cell(row=50, column=1, value="Fiftieth")
    save(wb, "classic11_sparse_rows.xlsx")


# ── 12. Sparse columns (A, D filled; B, C empty) ───────────────────────
def classic12_sparse_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Left"
    ws["D1"] = "Right"
    ws["A2"] = "Data1"
    ws["F2"] = "FarRight"
    ws["A3"] = "Row3"
    ws["J3"] = "VeryFar"
    save(wb, "classic12_sparse_columns.xlsx")


# ── 13. Date-like strings ───────────────────────────────────────────────
def classic13_date_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date", "Event"])
    ws.append(["2025-01-15", "Launch"])
    ws.append(["2025-06-30", "Release"])
    ws.append(["2025-12-25", "Holiday"])
    ws.append(["2026-01-01", "New Year"])
    ws.append(["2026-02-23", "Today"])
    save(wb, "classic13_date_strings.xlsx")


# ── 14. Decimal numbers ─────────────────────────────────────────────────
def classic14_decimal_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Constant", "Value"])
    ws.append(["Pi", 3.14159])
    ws.append(["e", 2.71828])
    ws.append(["Sqrt(2)", 1.41421])
    ws.append(["Phi", 1.61803])
    ws.append(["Ln(2)", 0.69315])
    save(wb, "classic14_decimal_numbers.xlsx")


# ── 15. Negative numbers ────────────────────────────────────────────────
def classic15_negative_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value"])
    ws.append(["Loss", -100.0])
    ws.append(["Small Loss", -0.5])
    ws.append(["Zero", 0.0])
    ws.append(["Gain", 50.0])
    ws.append(["Big Loss", -99999.99])
    ws.append(["Tiny", -0.001])
    save(wb, "classic15_negative_numbers.xlsx")


# ── 16. Percentage-like strings ──────────────────────────────────────────
def classic16_percentage_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Metric", "Rate"])
    ws.append(["Conversion", "12.5%"])
    ws.append(["Bounce", "45.3%"])
    ws.append(["Retention", "88.7%"])
    ws.append(["Churn", "3.2%"])
    ws.append(["Growth", "156.0%"])
    save(wb, "classic16_percentage_strings.xlsx")


# ── 17. Currency-like strings ───────────────────────────────────────────
def classic17_currency_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Price"])
    ws.append(["Widget", "$19.99"])
    ws.append(["Gadget", "$149.00"])
    ws.append(["Premium", "$1,299.99"])
    ws.append(["Budget", "$4.50"])
    ws.append(["Euro Item", "€49.99"])
    ws.append(["Yen Item", "¥5000"])
    save(wb, "classic17_currency_strings.xlsx")


# ── 18. Large dataset (1000 rows × 10 cols) ─────────────────────────────
def classic18_large_dataset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [f"Col{c}" for c in range(10)]
    ws.append(headers)
    for r in range(1000):
        ws.append([f"R{r}C{c}" for c in range(10)])
    save(wb, "classic18_large_dataset.xlsx")


# ── 19. Single column list ──────────────────────────────────────────────
def classic19_single_column_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Items"])
    for i in range(1, 21):
        ws.append([f"Item {i}"])
    save(wb, "classic19_single_column_list.xlsx")


# ── 20. All empty cells ─────────────────────────────────────────────────
def classic20_all_empty_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    save(wb, "classic20_all_empty_cells.xlsx")


# ── 21. Header only (no data rows) ──────────────────────────────────────
def classic21_header_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Col1", "Col2", "Col3", "Col4", "Col5"])
    save(wb, "classic21_header_only.xlsx")


# ── 22. Very long sheet name ────────────────────────────────────────────
def classic22_long_sheet_name():
    wb = Workbook()
    # Excel sheet name max is 31 characters
    ws = wb.active
    ws.title = "VeryLongSheetNameThatIsMaxLen"
    ws.append(["Data", "Value"])
    ws.append(["Row1", 100])
    ws.append(["Row2", 200])
    save(wb, "classic22_long_sheet_name.xlsx")


# ── 23. Unicode / CJK text ──────────────────────────────────────────────
def classic23_unicode_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Language", "Greeting", "Extra"])
    ws.append(["English", "Hello", "World"])
    ws.append(["Chinese", "你好", "世界"])
    ws.append(["Japanese", "こんにちは", "世界"])
    ws.append(["Korean", "안녕하세요", "세계"])
    ws.append(["Arabic", "مرحبا", "العالم"])
    ws.append(["Emoji", "😀🎉", "✅❌"])
    save(wb, "classic23_unicode_text.xlsx")


# ── 24. Red text (colored) ──────────────────────────────────────────────
def classic24_red_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    red_font = Font(color="FF0000", size=11)
    normal_font = Font(size=11)

    ws.append(["Status", "Message"])
    cell_a = ws.cell(row=2, column=1, value="Error")
    cell_a.font = red_font
    cell_b = ws.cell(row=2, column=2, value="Something went wrong")
    cell_b.font = red_font

    cell_a2 = ws.cell(row=3, column=1, value="OK")
    cell_a2.font = normal_font
    cell_b2 = ws.cell(row=3, column=2, value="All systems operational")
    cell_b2.font = normal_font

    cell_a3 = ws.cell(row=4, column=1, value="Warning")
    cell_a3.font = Font(color="FFA500", size=11)
    cell_b3 = ws.cell(row=4, column=2, value="Check disk space")
    cell_b3.font = Font(color="FFA500", size=11)
    save(wb, "classic24_red_text.xlsx")


# ── 25. Multiple colors ─────────────────────────────────────────────────
def classic25_multiple_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    colors = [
        ("Red", "FF0000"),
        ("Green", "00FF00"),
        ("Blue", "0000FF"),
        ("Yellow", "FFFF00"),
        ("Magenta", "FF00FF"),
        ("Cyan", "00FFFF"),
        ("Orange", "FFA500"),
        ("Purple", "800080"),
    ]
    ws.append(["Color Name", "Sample Text"])
    for name, color_hex in colors:
        row = ws.max_row + 1
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.font = Font(color=color_hex, size=11)
        cell_b = ws.cell(row=row, column=2, value=f"This is {name.lower()} text")
        cell_b.font = Font(color=color_hex, size=11)
    save(wb, "classic25_multiple_colors.xlsx")


# ── 26. Inline strings ──────────────────────────────────────────────────
def classic26_inline_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Inline1", "Inline2", "Inline3"])
    ws.append(["ValueA", "ValueB", "ValueC"])
    ws.append(["Test1", "Test2", "Test3"])
    save(wb, "classic26_inline_strings.xlsx")


# ── 27. Single row (horizontal data) ────────────────────────────────────
def classic27_single_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
    save(wb, "classic27_single_row.xlsx")


# ── 28. Duplicate values ────────────────────────────────────────────────
def classic28_duplicate_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Yes", "No", "Yes", "No"])
    ws.append(["No", "Yes", "No", "Yes"])
    ws.append(["Yes", "Yes", "Yes", "Yes"])
    ws.append(["No", "No", "No", "No"])
    ws.append(["Yes", "No", "Yes", "No"])
    save(wb, "classic28_duplicate_values.xlsx")


# ── 29. Formula-result values (pre-computed, no formula strings) ─────────
def classic29_formula_results():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "Sum", "Product"])
    ws.append([10, 20, 30, 200])
    ws.append([5, 15, 20, 75])
    ws.append([100, 200, 300, 20000])
    ws.append(["", "", 350, 20275])
    save(wb, "classic29_formula_results.xlsx")


# ── 30. Mixed empty and filled sheets ───────────────────────────────────
def classic30_mixed_empty_and_filled_sheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Empty"
    # No data in first sheet

    ws2 = wb.create_sheet("Data")
    ws2.append(["Hello", "World"])
    ws2.append(["Foo", "Bar"])
    ws2.append(["Baz", "Qux"])

    ws3 = wb.create_sheet("AlsoEmpty")
    # No data in third sheet

    ws4 = wb.create_sheet("MoreData")
    ws4.append(["Column1", "Column2", "Column3"])
    ws4.append([1, 2, 3])
    save(wb, "classic30_mixed_empty_and_filled_sheets.xlsx")


# ── 31. Bold header row ──────────────────────────────────────────────────
def classic31_bold_header_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bold = Font(bold=True, size=11)
    headers = ["Product", "Category", "Price", "Stock"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
    ws.append(["Laptop", "Electronics", 999.99, 50])
    ws.append(["Desk", "Furniture", 349.00, 20])
    ws.append(["Pen", "Stationery", 1.99, 500])
    ws.append(["Chair", "Furniture", 199.00, 30])
    save(wb, "classic31_bold_header_row.xlsx")


# ── 32. Right-aligned numbers ─────────────────────────────────────────────
def classic32_right_aligned_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    right = Alignment(horizontal="right")
    ws.append(["Label", "Amount"])
    data = [("Revenue", 125000), ("Expenses", 87500), ("Profit", 37500)]
    for row_idx, (label, amount) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        c = ws.cell(row=row_idx, column=2, value=amount)
        c.alignment = right
    save(wb, "classic32_right_aligned_numbers.xlsx")


# ── 33. Centered text ────────────────────────────────────────────────────
def classic33_centered_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    center = Alignment(horizontal="center")
    ws.append(["Mon", "Tue", "Wed", "Thu", "Fri"])
    data_rows = [
        [9, 10, 11, 9, 8],
        [12, 11, 10, 13, 12],
    ]
    for dr in data_rows:
        row_num = ws.max_row + 1
        for col, val in enumerate(dr, start=1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.alignment = center
    save(wb, "classic33_centered_text.xlsx")


# ── 34. Column width set explicitly ──────────────────────────────────────
def classic34_explicit_column_widths():
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.append(["ID", "Description", "Value"])
    ws.append([1, "Short", 10])
    ws.append([2, "A much longer description text here", 200])
    ws.append([3, "Medium length description", 55])
    save(wb, "classic34_explicit_column_widths.xlsx")


# ── 35. Row height set explicitly ────────────────────────────────────────
def classic35_explicit_row_heights():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 50
    ws.append(["Tall Header", "Value"])
    ws.append(["Extra Tall Row", 42])
    ws.append(["Normal Row", 10])
    save(wb, "classic35_explicit_row_heights.xlsx")


# ── 36. Merged cells ──────────────────────────────────────────────────────
def classic36_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:C1")
    ws["A1"] = "Merged Header Spanning Three Columns"
    ws.append(["Col1", "Col2", "Col3"])
    ws.append(["Row2A", "Row2B", "Row2C"])
    ws.append(["Row3A", "Row3B", "Row3C"])
    save(wb, "classic36_merged_cells.xlsx")


# ── 37. Freeze panes ──────────────────────────────────────────────────────
def classic37_freeze_panes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.freeze_panes = "A2"  # freeze first row
    ws.append(["ID", "Name", "Score", "Grade"])
    for i in range(1, 21):
        grade = "A" if i > 17 else "B" if i > 13 else "C" if i > 9 else "D"
        ws.append([i, f"Student{i:02d}", i * 5, grade])
    save(wb, "classic37_freeze_panes.xlsx")


# ── 38. Hyperlink cell ────────────────────────────────────────────────────
def classic38_hyperlink_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Resource", "URL"])
    cell = ws.cell(row=2, column=1, value="GitHub")
    cell.hyperlink = "https://github.com"
    cell.style = "Hyperlink"
    ws.cell(row=2, column=2, value="https://github.com")
    ws.cell(row=3, column=1, value="Docs")
    ws.cell(row=3, column=2, value="https://docs.microsoft.com")
    save(wb, "classic38_hyperlink_cell.xlsx")


# ── 39. Conditional-style financial table ────────────────────────────────
def classic39_financial_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    red = Font(color="FF0000")
    green = Font(color="008000")
    ws.append(["Month", "Budget", "Actual", "Variance"])
    data = [
        ("Jan", 10000, 9500, -500),
        ("Feb", 10000, 10800, 800),
        ("Mar", 10000, 9900, -100),
        ("Apr", 10000, 11200, 1200),
        ("May", 10000, 9700, -300),
        ("Jun", 10000, 10050, 50),
    ]
    for month, budget, actual, variance in data:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=month)
        ws.cell(row=row_idx, column=2, value=budget)
        ws.cell(row=row_idx, column=3, value=actual)
        var_cell = ws.cell(row=row_idx, column=4, value=variance)
        var_cell.font = green if variance >= 0 else red
    save(wb, "classic39_financial_table.xlsx")


# ── 40. Scientific notation numbers ──────────────────────────────────────
def classic40_scientific_notation():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value"])
    ws.append(["Avogadro", 6.022e23])
    ws.append(["Planck", 6.626e-34])
    ws.append(["Speed of Light", 2.998e8])
    ws.append(["Electron mass", 9.109e-31])
    ws.append(["Pi approx", 3.14159265358979])
    save(wb, "classic40_scientific_notation.xlsx")


# ── 41. Integer vs float mixed ────────────────────────────────────────────
def classic41_integer_vs_float():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Type", "Value"])
    ws.append(["Integer", 42])
    ws.append(["Float", 42.0])
    ws.append(["NegInt", -7])
    ws.append(["NegFloat", -7.5])
    ws.append(["Zero", 0])
    ws.append(["ZeroFloat", 0.0])
    ws.append(["Large", 1000000])
    ws.append(["Small", 0.000001])
    save(wb, "classic41_integer_vs_float.xlsx")


# ── 42. Boolean values ────────────────────────────────────────────────────
def classic42_boolean_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Feature", "Enabled"])
    ws.append(["Dark Mode", True])
    ws.append(["Notifications", False])
    ws.append(["Auto-save", True])
    ws.append(["Analytics", False])
    ws.append(["Beta Features", True])
    save(wb, "classic42_boolean_values.xlsx")


# ── 43. Inventory report ─────────────────────────────────────────────────
def classic43_inventory_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    bold = Font(bold=True)
    headers = ["SKU", "Name", "Category", "Qty", "Unit Price", "Total Value"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
    inventory = [
        ("SKU001", "Widget A", "Widgets", 100, 5.99, 599.00),
        ("SKU002", "Widget B", "Widgets", 250, 3.49, 872.50),
        ("SKU003", "Gadget X", "Gadgets", 50, 29.99, 1499.50),
        ("SKU004", "Gadget Y", "Gadgets", 75, 19.99, 1499.25),
        ("SKU005", "Tool Z", "Tools", 30, 49.99, 1499.70),
        ("SKU006", "Part P", "Parts", 500, 0.99, 495.00),
        ("SKU007", "Part Q", "Parts", 1000, 0.49, 490.00),
    ]
    for row in inventory:
        ws.append(list(row))
    save(wb, "classic43_inventory_report.xlsx")


# ── 44. Employee roster ───────────────────────────────────────────────────
def classic44_employee_roster():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(["EmpID", "First", "Last", "Dept", "Title", "Email"])
    employees = [
        (1001, "Alice", "Smith", "Engineering", "Senior Engineer", "alice@example.com"),
        (1002, "Bob", "Jones", "Marketing", "Marketing Manager", "bob@example.com"),
        (1003, "Carol", "Williams", "HR", "HR Specialist", "carol@example.com"),
        (1004, "David", "Brown", "Engineering", "Junior Engineer", "david@example.com"),
        (1005, "Eve", "Davis", "Finance", "Financial Analyst", "eve@example.com"),
        (1006, "Frank", "Miller", "Sales", "Sales Representative", "frank@example.com"),
        (1007, "Grace", "Wilson", "Engineering", "Tech Lead", "grace@example.com"),
        (1008, "Henry", "Moore", "Support", "Support Specialist", "henry@example.com"),
    ]
    for emp in employees:
        ws.append(list(emp))
    save(wb, "classic44_employee_roster.xlsx")


# ── 45. Sales by region (multi-sheet) ────────────────────────────────────
def classic45_sales_by_region():
    wb = Workbook()
    regions = [
        ("North", [("Q1", 45000), ("Q2", 52000), ("Q3", 61000), ("Q4", 71000)]),
        ("South", [("Q1", 38000), ("Q2", 41000), ("Q3", 39000), ("Q4", 44000)]),
        ("East",  [("Q1", 55000), ("Q2", 59000), ("Q3", 63000), ("Q4", 68000)]),
        ("West",  [("Q1", 47000), ("Q2", 51000), ("Q3", 55000), ("Q4", 60000)]),
    ]
    first = True
    for region, quarters in regions:
        ws = wb.active if first else wb.create_sheet(region)
        if first:
            ws.title = region
            first = False
        ws.append(["Quarter", "Sales"])
        for q, s in quarters:
            ws.append([q, s])
    save(wb, "classic45_sales_by_region.xlsx")


# ── 46. Grade book ───────────────────────────────────────────────────────
def classic46_grade_book():
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"
    ws.append(["Student", "HW1", "HW2", "Midterm", "Final", "Total", "Grade"])
    students = [
        ("Alice",   95, 88, 92, 90, 365, "A"),
        ("Bob",     70, 75, 68, 72, 285, "C"),
        ("Carol",   85, 90, 88, 91, 354, "A-"),
        ("David",   60, 55, 62, 58, 235, "D"),
        ("Eve",     92, 95, 97, 98, 382, "A+"),
        ("Frank",   78, 80, 75, 82, 315, "B"),
        ("Grace",   88, 85, 89, 87, 349, "B+"),
    ]
    for s in students:
        ws.append(list(s))
    save(wb, "classic46_grade_book.xlsx")


# ── 47. Time series data ──────────────────────────────────────────────────
def classic47_time_series():
    wb = Workbook()
    ws = wb.active
    ws.title = "Temperatures"
    ws.append(["Day", "High", "Low", "Avg"])
    import random
    random.seed(42)
    for day in range(1, 32):
        # Use "Day-NN" label to avoid LibreOffice auto-converting ISO dates
        day_label = f"Day-{day:02d}"
        high = round(random.uniform(15, 30), 1)
        low = round(random.uniform(5, 15), 1)
        avg = round((high + low) / 2, 1)
        ws.append([day_label, high, low, avg])
    save(wb, "classic47_time_series.xlsx")


# ── 48. Survey results ───────────────────────────────────────────────────
def classic48_survey_results():
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"
    # Deliberately keep question text short (≤16 chars) so all 6 columns fit
    # on one page: Q-col natural width ≈ (16+2)*5 = 90pt, total ≈ 90+80+35+45+50+95
    # = 395pt + 5*20 padding = 495pt ≤ 512pt usable → single column group.
    ws.append(["Question", "StrongAgree", "Agree", "Neutral", "Disagree", "StrongDisagree"])
    questions = [
        ("Easy to use",   30, 45, 15, 7, 3),
        ("Recommend",     25, 40, 20, 10, 5),
        ("Fair price",    20, 35, 25, 15, 5),
        ("Good support",  35, 40, 15, 7, 3),
        ("Satisfied",     28, 42, 18, 8, 4),
    ]
    for q in questions:
        ws.append(list(q))
    save(wb, "classic48_survey_results.xlsx")


# ── 49. Contact list ──────────────────────────────────────────────────────
def classic49_contact_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Name", "Phone", "Email", "City", "Country"])
    contacts = [
        ("Alice Smith",   "+1-555-0101", "alice@example.com",   "New York",  "USA"),
        ("Bob Jones",     "+44-20-7946-0958", "bob@example.co.uk", "London", "UK"),
        ("Carol Wang",    "+86-10-1234-5678", "carol@example.cn", "Beijing",  "China"),
        ("David Muller",  "+49-30-1234567",   "david@example.de", "Berlin",   "Germany"),
        ("Eve Martin",    "+33-1-23-45-67-89","eve@example.fr",   "Paris",    "France"),
        ("Frank Tanaka",  "+81-3-1234-5678",  "frank@example.jp","Tokyo",    "Japan"),
        ("Grace Kim",     "+82-2-1234-5678",  "grace@example.kr","Seoul",    "Korea"),
    ]
    for c in contacts:
        ws.append(list(c))
    save(wb, "classic49_contact_list.xlsx")


# ── 50. Budget vs actuals (three-sheet) ──────────────────────────────────
def classic50_budget_vs_actuals():
    wb = Workbook()

    # Sheet 1: Budget
    ws_budget = wb.active
    ws_budget.title = "Budget"
    ws_budget.append(["Department", "Q1", "Q2", "Q3", "Q4", "Annual"])
    budget_data = [
        ("Engineering", 200000, 200000, 210000, 220000, 830000),
        ("Marketing",   80000,  90000,  85000,  95000, 350000),
        ("Sales",       120000, 130000, 140000, 150000, 540000),
        ("HR",          40000,  40000,  42000,  43000, 165000),
        ("Finance",     35000,  35000,  37000,  38000, 145000),
    ]
    for row in budget_data:
        ws_budget.append(list(row))

    # Sheet 2: Actuals
    ws_actual = wb.create_sheet("Actuals")
    ws_actual.append(["Department", "Q1", "Q2", "Q3", "Q4", "Annual"])
    actual_data = [
        ("Engineering", 195000, 205000, 215000, 225000, 840000),
        ("Marketing",   82000,  88000,  91000,  97000, 358000),
        ("Sales",       118000, 135000, 142000, 148000, 543000),
        ("HR",          39000,  41000,  41500,  44000, 165500),
        ("Finance",     34000,  36000,  37500,  39000, 146500),
    ]
    for row in actual_data:
        ws_actual.append(list(row))

    # Sheet 3: Variance
    ws_var = wb.create_sheet("Variance")
    ws_var.append(["Department", "Q1", "Q2", "Q3", "Q4", "Annual"])
    for b, a in zip(budget_data, actual_data):
        ws_var.append([b[0]] + [a[i] - b[i] for i in range(1, 6)])

    save(wb, "classic50_budget_vs_actuals.xlsx")


# ── 51. Product catalog ───────────────────────────────────────────────────
def classic51_product_catalog():
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws.append(["Part#", "Name", "Description", "Weight(g)", "Price"])
    catalog = [
        ("P-001", "Basic Widget", "Standard widget for everyday use", 150, 4.99),
        ("P-002", "Pro Widget", "Enhanced widget with premium features", 180, 12.99),
        ("P-003", "Mini Gadget", "Compact gadget for mobile use", 90, 19.99),
        ("P-004", "Max Gadget", "Full-size gadget, industrial grade", 450, 89.99),
        ("P-005", "Connector A", "Type-A connector cable, 1m", 80, 7.49),
        ("P-006", "Connector B", "Type-B connector cable, 2m", 110, 9.99),
        ("P-007", "Adapter X", "Universal power adapter", 200, 15.99),
        ("P-008", "Adapter Y", "Travel power adapter", 120, 11.99),
        ("P-009", "Mount Bracket", "Wall mount bracket, steel", 600, 24.99),
        ("P-010", "Carry Case", "Padded carry case, waterproof", 350, 34.99),
    ]
    for row in catalog:
        ws.append(list(row))
    save(wb, "classic51_product_catalog.xlsx")


# ── 52. Pivot-style summary ───────────────────────────────────────────────
def classic52_pivot_summary():
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Region", "Electronics", "Furniture", "Clothing", "Food", "Total"])
    data = [
        ("North", 45000, 12000, 8000, 22000, 87000),
        ("South", 38000, 15000, 11000, 25000, 89000),
        ("East",  52000, 9000,  14000, 18000, 93000),
        ("West",  41000, 18000, 10000, 21000, 90000),
        ("Total", 176000, 54000, 43000, 86000, 359000),
    ]
    bold = Font(bold=True)
    for row_idx, row in enumerate(data, start=2):
        ws.append(list(row))
        if row_idx == len(data) + 1:  # Total row
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).font = bold
    save(wb, "classic52_pivot_summary.xlsx")


# ── 53. Invoice layout ────────────────────────────────────────────────────
def classic53_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    bold = Font(bold=True)
    # Header section
    ws["A1"] = "INVOICE"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Invoice #:"
    ws["B3"] = "INV-2025-0042"
    ws["A4"] = "Date:"
    ws["B4"] = "2025-03-01"
    ws["A5"] = "Due Date:"
    ws["B5"] = "2025-03-31"
    ws["A7"] = "Bill To:"
    ws["A7"].font = bold
    ws["A8"] = "ACME Corporation"
    ws["A9"] = "123 Business Rd, Suite 400"
    ws["A10"] = "New York, NY 10001"
    # Line items
    ws["A12"] = "Item"
    ws["B12"] = "Qty"
    ws["C12"] = "Unit Price"
    ws["D12"] = "Total"
    for col in range(1, 5):
        ws.cell(row=12, column=col).font = bold
    items = [
        ("Consulting Services", 10, 150.00, 1500.00),
        ("Software License", 5, 99.00, 495.00),
        ("Hardware", 2, 249.99, 499.98),
        ("Support Plan (annual)", 1, 1200.00, 1200.00),
    ]
    for r_idx, item in enumerate(items, start=13):
        ws.cell(row=r_idx, column=1, value=item[0])
        ws.cell(row=r_idx, column=2, value=item[1])
        ws.cell(row=r_idx, column=3, value=item[2])
        ws.cell(row=r_idx, column=4, value=item[3])
    ws.cell(row=18, column=3, value="Subtotal")
    ws.cell(row=18, column=4, value=3694.98)
    ws.cell(row=19, column=3, value="Tax (8%)")
    ws.cell(row=19, column=4, value=295.60)
    ws.cell(row=20, column=3, value="Total Due")
    ws.cell(row=20, column=3).font = bold
    ws.cell(row=20, column=4, value=3990.58)
    ws.cell(row=20, column=4).font = bold
    save(wb, "classic53_invoice.xlsx")


# ── 54. Multi-level header simulation ────────────────────────────────────
def classic54_multi_level_header():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bold = Font(bold=True)
    # Row 1: group headers
    ws["A1"] = ""
    ws["B1"] = "Q1"
    ws["D1"] = "Q2"
    ws["F1"] = "Q3"
    for cell_ref in ["B1", "D1", "F1"]:
        ws[cell_ref].font = bold
    # Row 2: sub-headers
    sub = ["ID", "Revenue", "Cost", "Revenue", "Cost", "Revenue", "Cost"]
    for col, h in enumerate(sub, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
    # Data rows
    data = [
        (1, 50000, 30000, 55000, 32000, 60000, 35000),
        (2, 45000, 28000, 48000, 29000, 52000, 31000),
        (3, 60000, 35000, 65000, 37000, 70000, 40000),
    ]
    for row in data:
        ws.append(list(row))
    save(wb, "classic54_multi_level_header.xlsx")


# ── 55. Error / N/A values (as strings) ──────────────────────────────────
def classic55_error_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Metric", "Value", "Status"])
    ws.append(["Sales", 12345, "OK"])
    ws.append(["Revenue", "#N/A", "Missing"])
    ws.append(["Cost", "#REF!", "Broken ref"])
    ws.append(["Profit", "#DIV/0!", "Div by zero"])
    ws.append(["Units", "#VALUE!", "Wrong type"])
    ws.append(["Target", 15000, "OK"])
    save(wb, "classic55_error_values.xlsx")


# ── 56. Alternating row fill colors ──────────────────────────────────────
def classic56_alternating_row_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    fill_light = PatternFill(fill_type="solid", fgColor="DCE6F1")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    ws.append(["#", "Product", "Price"])
    for i in range(1, 11):
        row_num = ws.max_row + 1
        ws.append([i, f"Product {i}", i * 10.0])
        fill = fill_light if i % 2 == 0 else fill_white
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).fill = fill
    save(wb, "classic56_alternating_row_colors.xlsx")


# ── 57. UTF-8 Asian characters (CJK only sheet) ──────────────────────────
def classic57_cjk_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "中文"
    ws.append(["序号", "产品名称", "价格", "库存"])
    ws.append([1, "笔记本电脑", 5999, 100])
    ws.append([2, "智能手机", 2999, 250])
    ws.append([3, "平板电脑", 1999, 150])
    ws.append([4, "蓝牙耳机", 299, 500])
    ws.append([5, "充电器", 99, 1000])
    save(wb, "classic57_cjk_only.xlsx")


# ── 58. Mixed numeric formats ─────────────────────────────────────────────
def classic58_mixed_numeric_formats():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Type", "Value"])
    ws.append(["Integer",           1000000])
    ws.append(["Float 2dp",         3.14])
    ws.append(["Float 5dp",         3.14159])
    ws.append(["Negative int",      -42])
    ws.append(["Negative float",    -3.14])
    ws.append(["Very small",        0.0001])
    ws.append(["Very large",        9999999.99])
    ws.append(["Zero",              0])
    ws.append(["Scientific approx", 1.23e10])
    save(wb, "classic58_mixed_numeric_formats.xlsx")


# ── 59. Multi-sheet with data on each plus a summary ─────────────────────
def classic59_multi_sheet_summary():
    wb = Workbook()

    months = ["Jan", "Feb", "Mar"]
    totals = []
    first = True
    for month in months:
        ws = wb.active if first else wb.create_sheet(month)
        if first:
            ws.title = month
            first = False
        ws.append(["Product", "Units", "Revenue"])
        revenue_total = 0
        for p in range(1, 6):
            units = (p * 7 + len(month)) % 50 + 10
            price = p * 5.0 + 9.99
            revenue = round(units * price, 2)
            revenue_total += revenue
            ws.append([f"Prod{p}", units, revenue])
        totals.append((month, revenue_total))

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Month", "Total Revenue"])
    for month, total in totals:
        ws_summary.append([month, round(total, 2)])

    save(wb, "classic59_multi_sheet_summary.xlsx")


# ── 60. Large wide table (20 columns × 50 rows) ───────────────────────────
def classic60_large_wide_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    num_cols = 20
    num_rows = 50
    headers = [f"Col{c+1:02d}" for c in range(num_cols)]
    ws.append(headers)
    for r in range(1, num_rows + 1):
        ws.append([f"R{r:02d}C{c+1:02d}" for c in range(num_cols)])
    save(wb, "classic60_large_wide_table.xlsx")


# ── 61. Product card with image ──────────────────────────────────────────
def classic61_product_card_with_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Product"
    bold = Font(bold=True)
    img = _make_jpeg_bytes(120, 120, (220, 80, 80))
    _add_image(ws, img, "A1", width_px=120, height_px=90)
    ws.row_dimensions[1].height = 68
    ws["D1"] = "Product Name"
    ws["D1"].font = bold
    ws["D2"] = "Widget Pro 3000"
    ws["D3"] = "Price"
    ws["D3"].font = bold
    ws["D4"] = "$29.99"
    ws["D5"] = "In Stock"
    ws["D6"] = 150
    save(wb, "classic61_product_card_with_image.xlsx")


# ── 62. Company logo header ───────────────────────────────────────────────
def classic62_company_logo_header():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True, size=14)
    logo = _make_jpeg_bytes(160, 60, (0, 80, 160))
    _add_image(ws, logo, "A1", width_px=160, height_px=60)
    ws.row_dimensions[1].height = 45
    ws["C1"] = "ACME Corporation"
    ws["C1"].font = bold
    ws["C2"] = "Annual Report 2025"
    ws.append([])
    ws.append(["Department", "Q1", "Q2", "Q3", "Q4"])
    for dept, q1, q2, q3, q4 in [
        ("Sales", 120, 135, 142, 160),
        ("Engineering", 85, 90, 95, 100),
        ("Marketing", 60, 65, 70, 75),
    ]:
        ws.append([dept, q1, q2, q3, q4])
    save(wb, "classic62_company_logo_header.xlsx")


# ── 63. Two products side by side ─────────────────────────────────────────
def classic63_two_products_side_by_side():
    wb = Workbook()
    ws = wb.active
    ws.title = "Compare"
    img_a = _make_jpeg_bytes(100, 100, (255, 140, 0))
    img_b = _make_jpeg_bytes(100, 100, (0, 160, 100))
    _add_image(ws, img_a, "A1", width_px=90, height_px=90)
    _add_image(ws, img_b, "D1", width_px=90, height_px=90)
    height_rows = 68
    for r in range(1, 7):
        ws.row_dimensions[r].height = height_rows // 6
    ws["A7"] = "Product A"
    ws["D7"] = "Product B"
    ws["A8"] = "Price: $19.99"
    ws["D8"] = "Price: $24.99"
    ws["A9"] = "Rating: 4.2"
    ws["D9"] = "Rating: 4.7"
    save(wb, "classic63_two_products_side_by_side.xlsx")


# ── 64. Employee directory with photo ────────────────────────────────────
def classic64_employee_directory_with_photo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Directory"
    bold = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="003366")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(["Photo", "Name", "Title", "Department", "Email"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    employees = [
        ("Alice Chen", "Engineer", "R&D", "alice@example.com", (100, 149, 237)),
        ("Bob Smith", "Manager", "Sales", "bob@example.com", (60, 179, 113)),
        ("Carol Wang", "Designer", "UX", "carol@example.com", (218, 112, 214)),
    ]
    for i, (name, title, dept, email, color) in enumerate(employees, start=2):
        photo = _make_jpeg_bytes(60, 60, color)
        _add_image(ws, photo, f"A{i}", width_px=50, height_px=45)
        ws.row_dimensions[i].height = 36
        ws[f"B{i}"] = name
        ws[f"C{i}"] = title
        ws[f"D{i}"] = dept
        ws[f"E{i}"] = email
    save(wb, "classic64_employee_directory_with_photo.xlsx")


# ── 65. Inventory with product photos ─────────────────────────────────────
def classic65_inventory_with_product_photos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    bold = Font(bold=True)
    ws.append(["Image", "SKU", "Name", "Qty", "Price"])
    ws[1][0].font = bold
    items = [
        ("SKU-001", "Red Widget", 50, 9.99, (200, 50, 50)),
        ("SKU-002", "Blue Gadget", 30, 14.99, (50, 80, 200)),
        ("SKU-003", "Green Tool", 100, 4.49, (50, 160, 50)),
        ("SKU-004", "Yellow Device", 25, 29.99, (220, 190, 0)),
        ("SKU-005", "Purple Gear", 75, 7.99, (140, 50, 200)),
    ]
    for i, (sku, name, qty, price, color) in enumerate(items, start=2):
        img = _make_jpeg_bytes(60, 60, color)
        _add_image(ws, img, f"A{i}", width_px=50, height_px=45)
        ws.row_dimensions[i].height = 36
        ws[f"B{i}"] = sku
        ws[f"C{i}"] = name
        ws[f"D{i}"] = qty
        ws[f"E{i}"] = price
    save(wb, "classic65_inventory_with_product_photos.xlsx")


# ── 66. Invoice with company logo ─────────────────────────────────────────
def classic66_invoice_with_logo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    logo = _make_jpeg_bytes(140, 55, (20, 60, 140))
    _add_image(ws, logo, "A1", width_px=140, height_px=55)
    ws.row_dimensions[1].height = 42
    bold = Font(bold=True)
    ws["D1"] = "INVOICE"
    ws["D1"].font = Font(bold=True, size=18)
    ws["D2"] = "Invoice #: INV-20250301"
    ws["D3"] = "Date: 2025-03-01"
    ws.append([])
    ws.append(["Description", "Qty", "Unit Price", "Total"])
    for cell in ws[5]:
        cell.font = bold
    ws.append(["Consulting Services", 8, 150.0, 1200.0])
    ws.append(["Software License", 1, 299.0, 299.0])
    ws.append(["Support Package", 1, 99.0, 99.0])
    ws.append([])
    ws.append(["", "", "Total", 1598.0])
    save(wb, "classic66_invoice_with_logo.xlsx")


# ── 67. Real-estate listing with photo ────────────────────────────────────
def classic67_real_estate_listing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Listing"
    house_img = _make_jpeg_bytes(200, 130, (180, 160, 130))
    _add_image(ws, house_img, "A1", width_px=200, height_px=130)
    for r in range(1, 11):
        ws.row_dimensions[r].height = 13
    bold = Font(bold=True)
    ws["D1"] = "123 Maple Street"
    ws["D1"].font = Font(bold=True, size=14)
    ws["D2"] = "Springfield, ST 12345"
    ws["D3"] = "List Price: $485,000"
    ws["D3"].font = bold
    ws.append([])
    ws.append(["Feature", "Detail"])
    for feat, detail in [
        ("Bedrooms", 4), ("Bathrooms", 2.5), ("Sq Ft", 2100),
        ("Lot Size", "0.25 acres"), ("Year Built", 1998),
    ]:
        ws.append([feat, detail])
    save(wb, "classic67_real_estate_listing.xlsx")


# ── 68. Restaurant menu with food photos ──────────────────────────────────
def classic68_restaurant_menu():
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    ws["A1"] = "Today's Menu"
    ws["A1"].font = Font(bold=True, size=16)
    ws.row_dimensions[1].height = 20
    menu_items = [
        ("Grilled Salmon", "$18.99", "Fresh Atlantic salmon with herbs", (240, 180, 100)),
        ("Caesar Salad", "$12.99", "Romaine lettuce, croutons, parmesan", (180, 220, 130)),
        ("Beef Burger", "$14.99", "8oz Angus beef, brioche bun", (190, 130, 80)),
        ("Pasta Primavera", "$13.99", "Seasonal vegetables, olive oil", (255, 220, 150)),
    ]
    row = 2
    for name, price, description, color in menu_items:
        # Text in columns A-C (left); image in column E (right)
        ws[f"A{row}"] = name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = price
        ws[f"A{row + 1}"] = description
        ws.row_dimensions[row].height = 57
        img = _make_jpeg_bytes(80, 80, color)
        _add_image(ws, img, f"E{row}", width_px=80, height_px=75)
        row += 3
    save(wb, "classic68_restaurant_menu.xlsx")


# ── 69. Image-only sheet ──────────────────────────────────────────────────
def classic69_image_only_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    cover_img = _make_jpeg_bytes(300, 200, (40, 100, 180))
    _add_image(ws, cover_img, "A1", width_px=300, height_px=200)
    for r in range(1, 16):
        ws.row_dimensions[r].height = 14
    save(wb, "classic69_image_only_sheet.xlsx")


# ── 70. Product catalog with 3 images ─────────────────────────────────────
def classic70_product_catalog_with_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    bold = Font(bold=True)
    ws["A1"] = "Product Catalog - Spring 2025"
    ws["A1"].font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 20
    products = [
        ("Classic Pen", "$3.99", "A reliable ballpoint pen", (60, 100, 180)),
        ("Leather Notebook", "$12.99", "Premium A5 notebook", (140, 100, 60)),
        ("Desk Organizer", "$24.99", "Bamboo desk tidy set", (100, 160, 100)),
    ]
    row = 3
    for name, price, desc, color in products:
        img = _make_jpeg_bytes(100, 100, color)
        _add_image(ws, img, f"A{row}", width_px=90, height_px=90)
        ws.row_dimensions[row].height = 68
        ws[f"C{row}"] = name
        ws[f"C{row}"].font = bold
        ws[f"D{row}"] = price
        ws[f"C{row + 1}"] = desc
        row += 4
    save(wb, "classic70_product_catalog_with_images.xlsx")


# ── 71. Multi-sheet workbook, each sheet has an image ─────────────────────
def classic71_multi_sheet_with_images():
    wb = Workbook()
    sheets_data = [
        ("Overview", (30, 100, 180), [("Metric", "Value"), ("Revenue", 1_200_000), ("Cost", 800_000)]),
        ("Marketing", (180, 60, 60), [("Channel", "Budget"), ("Digital", 50_000), ("Print", 20_000)]),
        ("HR", (60, 160, 80), [("Department", "Headcount"), ("Engineering", 45), ("Sales", 30)]),
    ]
    for i, (title, color, data) in enumerate(sheets_data):
        ws = wb.active if i == 0 else wb.create_sheet(title)
        ws.title = title
        img = _make_jpeg_bytes(100, 70, color)
        _add_image(ws, img, "A1", width_px=100, height_px=70)
        ws.row_dimensions[1].height = 53
        for j, row_data in enumerate(data, start=2):
            ws.append(list(row_data))
    save(wb, "classic71_multi_sheet_with_images.xlsx")


# ── 72. Bar chart screenshot + data ───────────────────────────────────────
def classic72_bar_chart_image_with_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    bold = Font(bold=True)
    ws["A1"] = "Monthly Sales Data"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Month", "Revenue", "Target"])
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    revenues = [42000, 48000, 51000, 45000, 56000, 62000]
    targets = [45000, 47000, 50000, 50000, 54000, 60000]
    for m, r, t in zip(months, revenues, targets):
        ws.append([m, r, t])
    # Embed a simulated chart image
    chart_img = _make_jpeg_bytes(220, 140, (245, 245, 245))
    _add_image(ws, chart_img, "E3", width_px=220, height_px=140)
    save(wb, "classic72_bar_chart_image_with_data.xlsx")


# ── 73. Event flyer with banner ───────────────────────────────────────────
def classic73_event_flyer_with_banner():
    wb = Workbook()
    ws = wb.active
    ws.title = "Event"
    banner = _make_jpeg_bytes(300, 100, (80, 30, 140))
    _add_image(ws, banner, "A1", width_px=300, height_px=100)
    ws.row_dimensions[1].height = 76
    bold = Font(bold=True)
    ws["A8"] = "Tech Summit 2025"
    ws["A8"].font = Font(bold=True, size=16)
    ws["A9"] = "Date: April 15, 2025"
    ws["A10"] = "Venue: Convention Center Hall A"
    ws["A11"] = "Speakers: 20+ Industry Leaders"
    ws.append([])
    ws.append(["Time", "Session", "Speaker"])
    for row in [
        ("09:00", "Opening Keynote", "Dr. Jane Kim"),
        ("10:30", "AI in Practice", "Prof. Mark Liu"),
        ("13:00", "Cloud Architecture", "Eng. Sara Patel"),
        ("15:00", "Panel Discussion", "All Speakers"),
    ]:
        ws.append(list(row))
    save(wb, "classic73_event_flyer_with_banner.xlsx")


# ── 74. Dashboard with KPIs and screenshot ────────────────────────────────
def classic74_dashboard_with_kpi_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    bold = Font(bold=True)
    ws["A1"] = "Executive Dashboard Q1 2025"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["KPI", "Target", "Actual", "Status"])
    for cell in ws[3]:
        cell.font = bold
    kpis = [
        ("Revenue", 500000, 523000, "✓ Above"),
        ("New Customers", 200, 187, "✗ Below"),
        ("NPS Score", 70, 74, "✓ Above"),
        ("Churn Rate", "< 3%", "2.8%", "✓ Above"),
    ]
    for kpi in kpis:
        ws.append(list(kpi))
    dash_img = _make_jpeg_bytes(180, 120, (240, 248, 255))
    _add_image(ws, dash_img, "F2", width_px=180, height_px=120)
    save(wb, "classic74_dashboard_with_kpi_image.xlsx")


# ── 75. Certificate with seal image ───────────────────────────────────────
def classic75_certificate_with_seal():
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate"
    bg_fill = PatternFill(fill_type="solid", fgColor="FEFEFE")
    bold_large = Font(bold=True, size=20)
    ws["B2"] = "Certificate of Achievement"
    ws["B2"].font = bold_large
    ws["B4"] = "This certifies that"
    ws["B5"] = "Alice Johnson"
    ws["B5"].font = Font(bold=True, size=16)
    ws["B6"] = "has successfully completed the Advanced Python Training"
    ws["B7"] = "Issued: March 1, 2025"
    seal = _make_jpeg_bytes(90, 90, (200, 160, 0))
    _add_image(ws, seal, "F3", width_px=90, height_px=90)
    save(wb, "classic75_certificate_with_seal.xlsx")


# ── 76. Four product images in a grid ─────────────────────────────────────
def classic76_product_image_grid():
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid"
    ws["A1"] = "Best Sellers"
    ws["A1"].font = Font(bold=True, size=14)
    # 2x2 grid compact layout: products in cols A-B and D-E to stay within page width
    products = [
        ("A3", "B3", (220, 80, 80),  "Red Phone Case",  "$9.99"),
        ("D3", "E3", (80, 120, 220), "Blue Speakers",   "$49.99"),
        ("A9", "B9", (80, 180, 80),  "Green Backpack",  "$34.99"),
        ("D9", "E9", (220, 160, 0),  "Yellow Headset",  "$29.99"),
    ]
    for img_anchor, text_anchor, color, name, price in products:
        img = _make_jpeg_bytes(80, 80, color)
        _add_image(ws, img, img_anchor, width_px=80, height_px=80)
        row_num = int("".join(filter(str.isdigit, text_anchor)))
        ws[text_anchor] = name
        ws[text_anchor].font = Font(bold=True)
        next_cell = text_anchor[0] + str(row_num + 1)
        ws[next_cell] = price
    save(wb, "classic76_product_image_grid.xlsx")


# ── 77. News article layout with hero image ───────────────────────────────
def classic77_news_article_with_hero_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Article"
    hero = _make_jpeg_bytes(280, 140, (100, 130, 160))
    _add_image(ws, hero, "A1", width_px=280, height_px=140)
    ws.row_dimensions[1].height = 107
    bold = Font(bold=True)
    ws["A11"] = "AI Transforms Modern Workplaces"
    ws["A11"].font = Font(bold=True, size=14)
    ws["A12"] = "By Jane Reporter | March 1, 2025"
    ws["A12"].font = Font(italic=True, color="808080")
    paragraphs = [
        "Artificial intelligence is reshaping how businesses operate globally.",
        "From automated workflows to intelligent decision support, the impact is clear.",
        "Companies adopting AI early report 30% productivity gains on average.",
        "Experts predict continued acceleration through 2030 and beyond.",
    ]
    for i, para in enumerate(paragraphs, start=14):
        ws[f"A{i}"] = para
    save(wb, "classic77_news_article_with_hero_image.xlsx")


# ── 78. Small icon in every row ───────────────────────────────────────────
def classic78_small_icon_per_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    bold = Font(bold=True)
    ws.append(["Icon", "Task", "Assignee", "Status"])
    for cell in ws[1]:
        cell.font = bold
    tasks = [
        ("Fix login bug", "Alice", "Done", (80, 200, 80)),
        ("Write unit tests", "Bob", "In Progress", (255, 165, 0)),
        ("Deploy to staging", "Carol", "Pending", (200, 80, 80)),
        ("Code review PR #42", "Alice", "Done", (80, 200, 80)),
        ("Update docs", "Dave", "In Progress", (255, 165, 0)),
    ]
    for i, (task, assignee, status, color) in enumerate(tasks, start=2):
        icon = _make_jpeg_bytes(30, 30, color)
        _add_image(ws, icon, f"A{i}", width_px=25, height_px=25)
        ws.row_dimensions[i].height = 20
        ws[f"B{i}"] = task
        ws[f"C{i}"] = assignee
        ws[f"D{i}"] = status
    save(wb, "classic78_small_icon_per_row.xlsx")


# ── 79. Wide panoramic banner ─────────────────────────────────────────────
def classic79_wide_panoramic_banner():
    wb = Workbook()
    ws = wb.active
    ws.title = "Banner"
    banner = _make_jpeg_bytes(480, 80, (10, 60, 120))
    _add_image(ws, banner, "A1", width_px=480, height_px=80)
    ws.row_dimensions[1].height = 61
    ws["A9"] = "Product Launch 2025"
    ws["A9"].font = Font(bold=True, size=18)
    ws["A10"] = "Introducing the next generation of innovation."
    ws["A11"] = "Available starting April 1, 2025"
    ws.append([])
    ws.append(["Model", "Storage", "RAM", "Price"])
    for cfg in [("Pro", "256GB", "16GB", "$999"), ("Max", "512GB", "32GB", "$1499")]:
        ws.append(list(cfg))
    save(wb, "classic79_wide_panoramic_banner.xlsx")


# ── 80. Portrait tall image ───────────────────────────────────────────────
def classic80_portrait_tall_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Portrait"
    portrait = _make_jpeg_bytes(100, 200, (160, 200, 240))
    _add_image(ws, portrait, "A1", width_px=90, height_px=180)
    ws["D1"] = "Profile"
    ws["D1"].font = Font(bold=True, size=14)
    ws["D2"] = "Name: Dr. Emily Zhao"
    ws["D3"] = "Title: Chief Scientist"
    ws["D4"] = "Dept: Research & Innovation"
    ws["D5"] = "Location: Singapore Office"
    ws["D6"] = "Email: emily.zhao@example.com"
    ws["D7"] = "LinkedIn: linkedin.com/in/ezhao"
    save(wb, "classic80_portrait_tall_image.xlsx")


# ── 81. Step-by-step guide with images ────────────────────────────────────
def classic81_step_by_step_with_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "Guide"
    ws["A1"] = "Quick Start Guide"
    ws["A1"].font = Font(bold=True, size=14)
    steps = [
        ("Step 1: Unbox", "Remove the product from packaging.", (200, 230, 255)),
        ("Step 2: Charge", "Connect USB-C cable, charge 2 hours.", (200, 255, 210)),
        ("Step 3: Power On", "Hold power button 3 seconds.", (255, 235, 200)),
        ("Step 4: Configure", "Follow on-screen setup wizard.", (240, 200, 255)),
    ]
    row = 3
    for title, desc, color in steps:
        step_img = _make_jpeg_bytes(80, 80, color)
        _add_image(ws, step_img, f"A{row}", width_px=75, height_px=75)
        ws.row_dimensions[row].height = 57
        ws[f"C{row}"] = title
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"C{row + 1}"] = desc
        row += 4
    save(wb, "classic81_step_by_step_with_images.xlsx")


# ── 82. Before/After comparison with images ───────────────────────────────
def classic82_before_after_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    bold = Font(bold=True)
    ws["A1"] = "Before"
    ws["A1"].font = bold
    ws["D1"] = "After"
    ws["D1"].font = bold
    before_img = _make_jpeg_bytes(130, 130, (180, 140, 130))
    after_img = _make_jpeg_bytes(130, 130, (100, 180, 160))
    _add_image(ws, before_img, "A2", width_px=130, height_px=130)
    _add_image(ws, after_img, "D2", width_px=130, height_px=130)
    for r in range(2, 12):
        ws.row_dimensions[r].height = 13
    ws["A12"] = "Old design – legacy UI"
    ws["D12"] = "New design – modern UI"
    ws.append([])
    ws.append(["Metric", "Before", "After", "Delta"])
    for m, b, a, d in [("Load time", "4.2s", "1.1s", "-74%"), ("Conversion", "2.1%", "4.8%", "+129%")]:
        ws.append([m, b, a, d])
    save(wb, "classic82_before_after_images.xlsx")


# ── 83. Color swatch palette ──────────────────────────────────────────────
def classic83_color_swatch_palette():
    wb = Workbook()
    ws = wb.active
    ws.title = "Palette"
    ws["A1"] = "Brand Color Palette"
    ws["A1"].font = Font(bold=True, size=13)
    colors = [
        ("Primary Blue", (0, 82, 165)),
        ("Primary Red", (197, 27, 50)),
        ("Accent Green", (0, 163, 108)),
        ("Neutral Grey", (128, 128, 128)),
        ("Warm Yellow", (255, 193, 7)),
        ("Dark Navy", (10, 30, 70)),
    ]
    row = 3
    for name, rgb in colors:
        swatch = _make_jpeg_bytes(60, 40, rgb)
        _add_image(ws, swatch, f"A{row}", width_px=55, height_px=38)
        ws.row_dimensions[row].height = 29
        ws[f"C{row}"] = name
        ws[f"D{row}"] = f"RGB({rgb[0]}, {rgb[1]}, {rgb[2]})"
        row += 2
    save(wb, "classic83_color_swatch_palette.xlsx")


# ── 84. Travel destination cards ──────────────────────────────────────────
def classic84_travel_destination_cards():
    wb = Workbook()
    ws = wb.active
    ws.title = "Destinations"
    ws["A1"] = "Top Travel Destinations 2025"
    ws["A1"].font = Font(bold=True, size=14)
    destinations = [
        ("Kyoto, Japan", "Cherry blossoms and ancient temples", (180, 120, 150)),
        ("Prague, Czech Republic", "Medieval architecture and vibrant culture", (130, 160, 200)),
        ("Cape Town, South Africa", "Mountains, ocean, and wildlife", (160, 200, 140)),
    ]
    row = 3
    for place, desc, color in destinations:
        dest_img = _make_jpeg_bytes(140, 90, color)
        _add_image(ws, dest_img, f"A{row}", width_px=130, height_px=90)
        ws.row_dimensions[row].height = 68
        ws[f"D{row}"] = place
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"D{row + 1}"] = desc
        row += 5
    save(wb, "classic84_travel_destination_cards.xlsx")


# ── 85. Science lab results with specimen image ───────────────────────────
def classic85_lab_results_with_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Results"
    bold = Font(bold=True)
    ws["A1"] = "Sample Analysis Report"
    ws["A1"].font = Font(bold=True, size=13)
    specimen_img = _make_jpeg_bytes(120, 100, (210, 220, 230))
    _add_image(ws, specimen_img, "E2", width_px=120, height_px=100)
    ws.append([])
    ws.append(["Parameter", "Value", "Unit", "Reference Range", "Flag"])
    for cell in ws[3]:
        cell.font = bold
    results = [
        ("pH", 7.35, "", "7.35 – 7.45", "Normal"),
        ("Glucose", 5.2, "mmol/L", "3.9 – 5.5", "Normal"),
        ("Sodium", 142, "mEq/L", "136 – 145", "Normal"),
        ("Potassium", 5.0, "mEq/L", "3.5 – 5.0", "Normal"),
        ("Creatinine", 1.4, "mg/dL", "0.6 – 1.2", "High"),
    ]
    for r in results:
        ws.append(list(r))
    save(wb, "classic85_lab_results_with_image.xlsx")


# ── 86. Software screenshot + feature list ────────────────────────────────
def classic86_software_screenshot_features():
    wb = Workbook()
    ws = wb.active
    ws.title = "Features"
    screenshot = _make_jpeg_bytes(200, 150, (230, 240, 250))
    _add_image(ws, screenshot, "A1", width_px=200, height_px=150)
    ws["D1"] = "MiniApp v2.0"
    ws["D1"].font = Font(bold=True, size=15)
    ws["D2"] = "The fastest lightweight app"
    for r in range(1, 13):
        ws.row_dimensions[r].height = 12
    ws.append([])
    ws.append(["Feature", "Available"])
    bold = Font(bold=True)
    ws[ws.max_row][0].font = bold
    features = [
        ("Dark Mode", "Yes"), ("Auto Save", "Yes"), ("Cloud Sync", "Yes"),
        ("Offline Mode", "Yes"), ("API Access", "Pro only"), ("Export to PDF", "Yes"),
    ]
    for f in features:
        ws.append(list(f))
    save(wb, "classic86_software_screenshot_features.xlsx")


# ── 87. Sports results with team logos ────────────────────────────────────
def classic87_sports_results_with_logos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    bold = Font(bold=True)
    ws["A1"] = "League Standings - Season 2025"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Logo", "Team", "W", "L", "D", "Pts"])
    for cell in ws[3]:
        cell.font = bold
    teams = [
        ("Eagles", 18, 4, 2, 56, (0, 80, 180)),
        ("Tigers", 15, 7, 2, 47, (220, 120, 0)),
        ("Sharks", 13, 8, 3, 42, (0, 160, 200)),
        ("Wolves", 10, 10, 4, 34, (120, 30, 30)),
    ]
    row = 4
    for team, w, l, d, pts, color in teams:
        logo = _make_jpeg_bytes(40, 40, color)
        _add_image(ws, logo, f"A{row}", width_px=35, height_px=32)
        ws.row_dimensions[row].height = 26
        ws[f"B{row}"] = team
        ws[f"C{row}"] = w
        ws[f"D{row}"] = l
        ws[f"E{row}"] = d
        ws[f"F{row}"] = pts
        row += 1
    save(wb, "classic87_sports_results_with_logos.xlsx")


# ── 88. Image after data rows ─────────────────────────────────────────────
def classic88_image_after_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True)
    ws.append(["Quarter", "Revenue", "Expenses", "Profit"])
    for cell in ws[1]:
        cell.font = bold
    for q, r, e in [("Q1", 120000, 80000), ("Q2", 135000, 88000),
                    ("Q3", 142000, 91000), ("Q4", 160000, 95000)]:
        ws.append([q, r, e, r - e])
    ws.append([])
    footer_img = _make_jpeg_bytes(200, 80, (220, 235, 255))
    _add_image(ws, footer_img, "A8", width_px=200, height_px=80)
    ws["D8"] = "Prepared by Finance Team"
    ws["D9"] = "Confidential - Q4 2025"
    save(wb, "classic88_image_after_data.xlsx")


# ── 89. Nutrition label with product image ────────────────────────────────
def classic89_nutrition_label_with_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Nutrition"
    bold = Font(bold=True)
    product_img = _make_jpeg_bytes(100, 130, (230, 200, 160))
    _add_image(ws, product_img, "A1", width_px=95, height_px=125)
    ws["C1"] = "Nutrition Facts"
    ws["C1"].font = Font(bold=True, size=14)
    ws["C2"] = "Serving Size: 30g (approx. 1 cup)"
    ws.append([])
    ws.append(["Nutrient", "Amount per serving", "% Daily Value"])
    for cell in ws[4]:
        cell.font = bold
    nutrients = [
        ("Calories", "120 kcal", ""),
        ("Total Fat", "3g", "4%"),
        ("Saturated Fat", "0.5g", "3%"),
        ("Sodium", "160mg", "7%"),
        ("Total Carbohydrate", "22g", "8%"),
        ("Dietary Fiber", "3g", "11%"),
        ("Sugars", "4g", ""),
        ("Protein", "3g", ""),
    ]
    for n in nutrients:
        ws.append(list(n))
    save(wb, "classic89_nutrition_label_with_image.xlsx")


# ── 90. Project status with milestone images ──────────────────────────────
def classic90_project_status_with_milestones():
    wb = Workbook()
    ws = wb.active
    ws.title = "Project"
    bold = Font(bold=True)
    ws["A1"] = "Project Orion – Status Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Reporting Period: Q1 2025"
    ws.append([])
    ws.append(["Milestone", "Due Date", "Owner", "Status"])
    for cell in ws[4]:
        cell.font = bold
    milestones = [
        ("Requirements Freeze", "Jan 15", "PM Team", "Complete"),
        ("Architecture Review", "Feb 1", "Tech Lead", "Complete"),
        ("Alpha Release", "Feb 28", "Dev Team", "In Progress"),
        ("Beta Testing", "Mar 31", "QA Team", "Not Started"),
        ("Production Deploy", "Apr 15", "DevOps", "Not Started"),
    ]
    for m in milestones:
        ws.append(list(m))
    ws.append([])
    team_img = _make_jpeg_bytes(160, 100, (200, 220, 200))
    _add_image(ws, team_img, "F4", width_px=160, height_px=100)
    save(wb, "classic90_project_status_with_milestones.xlsx")


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    ensure_output_dir()
    print(f"Generating 90 classic .xlsx files in: {OUTPUT_DIR}\n")

    generators = [
        classic01_basic_table_with_headers,
        classic02_multiple_worksheets,
        classic03_empty_workbook,
        classic04_single_cell,
        classic05_wide_table,
        classic06_tall_table,
        classic07_numbers_only,
        classic08_mixed_text_and_numbers,
        classic09_long_text,
        classic10_special_xml_characters,
        classic11_sparse_rows,
        classic12_sparse_columns,
        classic13_date_strings,
        classic14_decimal_numbers,
        classic15_negative_numbers,
        classic16_percentage_strings,
        classic17_currency_strings,
        classic18_large_dataset,
        classic19_single_column_list,
        classic20_all_empty_cells,
        classic21_header_only,
        classic22_long_sheet_name,
        classic23_unicode_text,
        classic24_red_text,
        classic25_multiple_colors,
        classic26_inline_strings,
        classic27_single_row,
        classic28_duplicate_values,
        classic29_formula_results,
        classic30_mixed_empty_and_filled_sheets,
        classic31_bold_header_row,
        classic32_right_aligned_numbers,
        classic33_centered_text,
        classic34_explicit_column_widths,
        classic35_explicit_row_heights,
        classic36_merged_cells,
        classic37_freeze_panes,
        classic38_hyperlink_cell,
        classic39_financial_table,
        classic40_scientific_notation,
        classic41_integer_vs_float,
        classic42_boolean_values,
        classic43_inventory_report,
        classic44_employee_roster,
        classic45_sales_by_region,
        classic46_grade_book,
        classic47_time_series,
        classic48_survey_results,
        classic49_contact_list,
        classic50_budget_vs_actuals,
        classic51_product_catalog,
        classic52_pivot_summary,
        classic53_invoice,
        classic54_multi_level_header,
        classic55_error_values,
        classic56_alternating_row_colors,
        classic57_cjk_only,
        classic58_mixed_numeric_formats,
        classic59_multi_sheet_summary,
        classic60_large_wide_table,
        # 61-90: image cases
        classic61_product_card_with_image,
        classic62_company_logo_header,
        classic63_two_products_side_by_side,
        classic64_employee_directory_with_photo,
        classic65_inventory_with_product_photos,
        classic66_invoice_with_logo,
        classic67_real_estate_listing,
        classic68_restaurant_menu,
        classic69_image_only_sheet,
        classic70_product_catalog_with_images,
        classic71_multi_sheet_with_images,
        classic72_bar_chart_image_with_data,
        classic73_event_flyer_with_banner,
        classic74_dashboard_with_kpi_image,
        classic75_certificate_with_seal,
        classic76_product_image_grid,
        classic77_news_article_with_hero_image,
        classic78_small_icon_per_row,
        classic79_wide_panoramic_banner,
        classic80_portrait_tall_image,
        classic81_step_by_step_with_images,
        classic82_before_after_images,
        classic83_color_swatch_palette,
        classic84_travel_destination_cards,
        classic85_lab_results_with_image,
        classic86_software_screenshot_features,
        classic87_sports_results_with_logos,
        classic88_image_after_data,
        classic89_nutrition_label_with_image,
        classic90_project_status_with_milestones,
    ]

    for gen in generators:
        gen()

    print(f"\nDone! {len(generators)} files generated.")


if __name__ == "__main__":
    main()
