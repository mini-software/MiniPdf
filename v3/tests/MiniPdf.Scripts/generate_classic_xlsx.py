"""
Generate 180 classic .xlsx files for testing Excel-to-PDF conversion.
Each file corresponds to a test case in ClassicExcelToPdfTests.cs.

Cases 61-90 include embedded images to exercise MiniPdf image rendering.
Cases 91-120 include openpyxl chart objects (bar, line, pie, area, etc.).
Cases 151-180 include multilingual text, special marks, and emoji.

Usage:
    pip install openpyxl pillow
    python generate_classic_xlsx.py

Output directory: ./output/
"""

import io
import os
import string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart, BarChart3D, LineChart, LineChart3D, PieChart, PieChart3D,
    AreaChart, AreaChart3D, ScatterChart, DoughnutChart, RadarChart,
    BubbleChart, Reference,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False


def _make_jpeg_bytes(width: int, height: int, color: tuple) -> bytes:
    """Return raw JPEG bytes for a solid-color image (requires Pillow)."""
    img = PILImage.new("RGB", (width, height), color)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


def _add_image(ws, jpeg_bytes: bytes, anchor: str, width_px: int = 120, height_px: int = 90):
    """Embed a JPEG image into the worksheet at the given cell anchor."""
    if not HAS_PIL:
        return
    from openpyxl.drawing.image import Image as XLImage
    buf = io.BytesIO(jpeg_bytes)
    xl_img = XLImage(buf)
    xl_img.width = width_px
    xl_img.height = height_px
    xl_img.anchor = anchor
    ws.add_image(xl_img)

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def save(wb: Workbook, filename: str):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  [OK] {filename}")


# ── 01. Basic table with headers ────────────────────────────────────────
def classic01_basic_table_with_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 30, "New York"])
    ws.append(["Bob", 25, "London"])
    ws.append(["Charlie", 35, "Tokyo"])
    ws.append(["Diana", 28, "Paris"])
    save(wb, "classic01_basic_table_with_headers.xlsx")


# ── 02. Multiple worksheets ─────────────────────────────────────────────
def classic02_multiple_worksheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Quarter", "Revenue"])
    ws1.append(["Q1", 100])
    ws1.append(["Q2", 200])
    ws1.append(["Q3", 350])
    ws1.append(["Q4", 480])

    ws2 = wb.create_sheet("Costs")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 500])
    ws2.append(["Salary", 3000])
    ws2.append(["Utilities", 200])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Revenue", 1130])
    ws3.append(["Total Costs", 3700])
    ws3.append(["Net", -2570])
    save(wb, "classic02_multiple_worksheets.xlsx")


# ── 03. Empty workbook (no data rows) ───────────────────────────────────
def classic03_empty_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # No data at all
    save(wb, "classic03_empty_workbook.xlsx")


# ── 04. Single cell ─────────────────────────────────────────────────────
def classic04_single_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    save(wb, "classic04_single_cell.xlsx")


# ── 05. Wide table (26 columns A–Z) ─────────────────────────────────────
def classic05_wide_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = list(string.ascii_uppercase)  # A-Z
    ws.append(headers)
    for row_idx in range(1, 6):
        ws.append([f"{ch}{row_idx}" for ch in headers])
    save(wb, "classic05_wide_table.xlsx")


# ── 06. Tall table (200 rows → multi-page) ──────────────────────────────
def classic06_tall_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Row#", "Value", "Description"])
    for i in range(1, 201):
        ws.append([f"Row{i}", f"Val{i}", f"This is the description for row number {i}"])
    save(wb, "classic06_tall_table.xlsx")


# ── 07. Numbers only ────────────────────────────────────────────────────
def classic07_numbers_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([1.0, 2.0, 3.0])
    ws.append([4.0, 5.0, 6.0])
    ws.append([7.0, 8.0, 9.0])
    ws.append([10.0, 100.0, 1000.0])
    save(wb, "classic07_numbers_only.xlsx")


# ── 08. Mixed text and numbers ──────────────────────────────────────────
def classic08_mixed_text_and_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Amount"])
    ws.append(["Item", 10.5])
    ws.append(["Tax", 0.08])
    ws.append(["Total", 10.58])
    ws.append(["Discount", -1.5])
    ws.append(["Final", 9.08])
    save(wb, "classic08_mixed_text_and_numbers.xlsx")


# ── 09. Long text content ───────────────────────────────────────────────
def classic09_long_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Long Text Column"])
    ws.append(["X" * 500])
    ws.append(["A" * 300 + " " + "B" * 200])
    ws.append(["Short"])
    ws.append(["Y" * 1000])
    save(wb, "classic09_long_text.xlsx")


# ── 10. Special XML characters ──────────────────────────────────────────
def classic10_special_xml_characters():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Special Characters"])
    ws.append(["A&B"])
    ws.append(["<tag>"])
    ws.append(['"quoted"'])
    ws.append(["it's"])
    ws.append(["Tom & Jerry < Batman > Superman"])
    ws.append(['He said "hello" & she replied \'hi\''])
    save(wb, "classic10_special_xml_characters.xlsx")


# ── 11. Sparse rows (gaps between data rows) ────────────────────────────
def classic11_sparse_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="First")
    ws.cell(row=5, column=1, value="Fifth")
    ws.cell(row=10, column=1, value="Tenth")
    ws.cell(row=20, column=1, value="Twentieth")
    ws.cell(row=50, column=1, value="Fiftieth")
    save(wb, "classic11_sparse_rows.xlsx")


# ── 12. Sparse columns (A, D filled; B, C empty) ───────────────────────
def classic12_sparse_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Left"
    ws["D1"] = "Right"
    ws["A2"] = "Data1"
    ws["F2"] = "FarRight"
    ws["A3"] = "Row3"
    ws["J3"] = "VeryFar"
    save(wb, "classic12_sparse_columns.xlsx")


# ── 13. Date-like strings ───────────────────────────────────────────────
def classic13_date_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Date", "Event"])
    ws.append(["2025-01-15", "Launch"])
    ws.append(["2025-06-30", "Release"])
    ws.append(["2025-12-25", "Holiday"])
    ws.append(["2026-01-01", "New Year"])
    ws.append(["2026-02-23", "Today"])
    save(wb, "classic13_date_strings.xlsx")


# ── 14. Decimal numbers ─────────────────────────────────────────────────
def classic14_decimal_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Constant", "Value"])
    ws.append(["Pi", 3.14159])
    ws.append(["e", 2.71828])
    ws.append(["Sqrt(2)", 1.41421])
    ws.append(["Phi", 1.61803])
    ws.append(["Ln(2)", 0.69315])
    save(wb, "classic14_decimal_numbers.xlsx")


# ── 15. Negative numbers ────────────────────────────────────────────────
def classic15_negative_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value"])
    ws.append(["Loss", -100.0])
    ws.append(["Small Loss", -0.5])
    ws.append(["Zero", 0.0])
    ws.append(["Gain", 50.0])
    ws.append(["Big Loss", -99999.99])
    ws.append(["Tiny", -0.001])
    save(wb, "classic15_negative_numbers.xlsx")


# ── 16. Percentage-like strings ──────────────────────────────────────────
def classic16_percentage_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Metric", "Rate"])
    ws.append(["Conversion", "12.5%"])
    ws.append(["Bounce", "45.3%"])
    ws.append(["Retention", "88.7%"])
    ws.append(["Churn", "3.2%"])
    ws.append(["Growth", "156.0%"])
    save(wb, "classic16_percentage_strings.xlsx")


# ── 17. Currency-like strings ───────────────────────────────────────────
def classic17_currency_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Price"])
    ws.append(["Widget", "$19.99"])
    ws.append(["Gadget", "$149.00"])
    ws.append(["Premium", "$1,299.99"])
    ws.append(["Budget", "$4.50"])
    ws.append(["Euro Item", "€49.99"])
    ws.append(["Yen Item", "¥5000"])
    save(wb, "classic17_currency_strings.xlsx")


# ── 18. Large dataset (1000 rows × 10 cols) ─────────────────────────────
def classic18_large_dataset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [f"Col{c}" for c in range(10)]
    ws.append(headers)
    for r in range(1000):
        ws.append([f"R{r}C{c}" for c in range(10)])
    save(wb, "classic18_large_dataset.xlsx")


# ── 19. Single column list ──────────────────────────────────────────────
def classic19_single_column_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Items"])
    for i in range(1, 21):
        ws.append([f"Item {i}"])
    save(wb, "classic19_single_column_list.xlsx")


# ── 20. All empty cells ─────────────────────────────────────────────────
def classic20_all_empty_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    ws.append(["", "", ""])
    save(wb, "classic20_all_empty_cells.xlsx")


# ── 21. Header only (no data rows) ──────────────────────────────────────
def classic21_header_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Col1", "Col2", "Col3", "Col4", "Col5"])
    save(wb, "classic21_header_only.xlsx")


# ── 22. Very long sheet name ────────────────────────────────────────────
def classic22_long_sheet_name():
    wb = Workbook()
    # Excel sheet name max is 31 characters
    ws = wb.active
    ws.title = "VeryLongSheetNameThatIsMaxLen"
    ws.append(["Data", "Value"])
    ws.append(["Row1", 100])
    ws.append(["Row2", 200])
    save(wb, "classic22_long_sheet_name.xlsx")


# ── 23. Unicode / CJK text ──────────────────────────────────────────────
def classic23_unicode_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Language", "Greeting", "Extra"])
    ws.append(["English", "Hello", "World"])
    ws.append(["Chinese", "你好", "世界"])
    ws.append(["Japanese", "こんにちは", "世界"])
    ws.append(["Korean", "안녕하세요", "세계"])
    ws.append(["Arabic", "مرحبا", "العالم"])
    ws.append(["Emoji", "😀🎉", "✅❌"])
    save(wb, "classic23_unicode_text.xlsx")


# ── 24. Red text (colored) ──────────────────────────────────────────────
def classic24_red_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    red_font = Font(color="FF0000", size=11)
    normal_font = Font(size=11)

    ws.append(["Status", "Message"])
    cell_a = ws.cell(row=2, column=1, value="Error")
    cell_a.font = red_font
    cell_b = ws.cell(row=2, column=2, value="Something went wrong")
    cell_b.font = red_font

    cell_a2 = ws.cell(row=3, column=1, value="OK")
    cell_a2.font = normal_font
    cell_b2 = ws.cell(row=3, column=2, value="All systems operational")
    cell_b2.font = normal_font

    cell_a3 = ws.cell(row=4, column=1, value="Warning")
    cell_a3.font = Font(color="FFA500", size=11)
    cell_b3 = ws.cell(row=4, column=2, value="Check disk space")
    cell_b3.font = Font(color="FFA500", size=11)
    save(wb, "classic24_red_text.xlsx")


# ── 25. Multiple colors ─────────────────────────────────────────────────
def classic25_multiple_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    colors = [
        ("Red", "FF0000"),
        ("Green", "00FF00"),
        ("Blue", "0000FF"),
        ("Yellow", "FFFF00"),
        ("Magenta", "FF00FF"),
        ("Cyan", "00FFFF"),
        ("Orange", "FFA500"),
        ("Purple", "800080"),
    ]
    ws.append(["Color Name", "Sample Text"])
    for name, color_hex in colors:
        row = ws.max_row + 1
        cell_a = ws.cell(row=row, column=1, value=name)
        cell_a.font = Font(color=color_hex, size=11)
        cell_b = ws.cell(row=row, column=2, value=f"This is {name.lower()} text")
        cell_b.font = Font(color=color_hex, size=11)
    save(wb, "classic25_multiple_colors.xlsx")


# ── 26. Inline strings ──────────────────────────────────────────────────
def classic26_inline_strings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Inline1", "Inline2", "Inline3"])
    ws.append(["ValueA", "ValueB", "ValueC"])
    ws.append(["Test1", "Test2", "Test3"])
    save(wb, "classic26_inline_strings.xlsx")


# ── 27. Single row (horizontal data) ────────────────────────────────────
def classic27_single_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
    save(wb, "classic27_single_row.xlsx")


# ── 28. Duplicate values ────────────────────────────────────────────────
def classic28_duplicate_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Yes", "No", "Yes", "No"])
    ws.append(["No", "Yes", "No", "Yes"])
    ws.append(["Yes", "Yes", "Yes", "Yes"])
    ws.append(["No", "No", "No", "No"])
    ws.append(["Yes", "No", "Yes", "No"])
    save(wb, "classic28_duplicate_values.xlsx")


# ── 29. Formula-result values (pre-computed, no formula strings) ─────────
def classic29_formula_results():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B", "Sum", "Product"])
    ws.append([10, 20, 30, 200])
    ws.append([5, 15, 20, 75])
    ws.append([100, 200, 300, 20000])
    ws.append(["", "", 350, 20275])
    save(wb, "classic29_formula_results.xlsx")


# ── 30. Mixed empty and filled sheets ───────────────────────────────────
def classic30_mixed_empty_and_filled_sheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Empty"
    # No data in first sheet

    ws2 = wb.create_sheet("Data")
    ws2.append(["Hello", "World"])
    ws2.append(["Foo", "Bar"])
    ws2.append(["Baz", "Qux"])

    ws3 = wb.create_sheet("AlsoEmpty")
    # No data in third sheet

    ws4 = wb.create_sheet("MoreData")
    ws4.append(["Column1", "Column2", "Column3"])
    ws4.append([1, 2, 3])
    save(wb, "classic30_mixed_empty_and_filled_sheets.xlsx")


# ── 31. Bold header row ──────────────────────────────────────────────────
def classic31_bold_header_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bold = Font(bold=True, size=11)
    headers = ["Product", "Category", "Price", "Stock"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
    ws.append(["Laptop", "Electronics", 999.99, 50])
    ws.append(["Desk", "Furniture", 349.00, 20])
    ws.append(["Pen", "Stationery", 1.99, 500])
    ws.append(["Chair", "Furniture", 199.00, 30])
    save(wb, "classic31_bold_header_row.xlsx")


# ── 32. Right-aligned numbers ─────────────────────────────────────────────
def classic32_right_aligned_numbers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    right = Alignment(horizontal="right")
    ws.append(["Label", "Amount"])
    data = [("Revenue", 125000), ("Expenses", 87500), ("Profit", 37500)]
    for row_idx, (label, amount) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        c = ws.cell(row=row_idx, column=2, value=amount)
        c.alignment = right
    save(wb, "classic32_right_aligned_numbers.xlsx")


# ── 33. Centered text ────────────────────────────────────────────────────
def classic33_centered_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    center = Alignment(horizontal="center")
    ws.append(["Mon", "Tue", "Wed", "Thu", "Fri"])
    data_rows = [
        [9, 10, 11, 9, 8],
        [12, 11, 10, 13, 12],
    ]
    for dr in data_rows:
        row_num = ws.max_row + 1
        for col, val in enumerate(dr, start=1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.alignment = center
    save(wb, "classic33_centered_text.xlsx")


# ── 34. Column width set explicitly ──────────────────────────────────────
def classic34_explicit_column_widths():
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.append(["ID", "Description", "Value"])
    ws.append([1, "Short", 10])
    ws.append([2, "A much longer description text here", 200])
    ws.append([3, "Medium length description", 55])
    save(wb, "classic34_explicit_column_widths.xlsx")


# ── 35. Row height set explicitly ────────────────────────────────────────
def classic35_explicit_row_heights():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 50
    ws.append(["Tall Header", "Value"])
    ws.append(["Extra Tall Row", 42])
    ws.append(["Normal Row", 10])
    save(wb, "classic35_explicit_row_heights.xlsx")


# ── 36. Merged cells ──────────────────────────────────────────────────────
def classic36_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.merge_cells("A1:C1")
    ws["A1"] = "Merged Header Spanning Three Columns"
    ws.append(["Col1", "Col2", "Col3"])
    ws.append(["Row2A", "Row2B", "Row2C"])
    ws.append(["Row3A", "Row3B", "Row3C"])
    save(wb, "classic36_merged_cells.xlsx")


# ── 37. Freeze panes ──────────────────────────────────────────────────────
def classic37_freeze_panes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.freeze_panes = "A2"  # freeze first row
    ws.append(["ID", "Name", "Score", "Grade"])
    for i in range(1, 21):
        grade = "A" if i > 17 else "B" if i > 13 else "C" if i > 9 else "D"
        ws.append([i, f"Student{i:02d}", i * 5, grade])
    save(wb, "classic37_freeze_panes.xlsx")


# ── 38. Hyperlink cell ────────────────────────────────────────────────────
def classic38_hyperlink_cell():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Resource", "URL"])
    cell = ws.cell(row=2, column=1, value="GitHub")
    cell.hyperlink = "https://github.com"
    cell.style = "Hyperlink"
    ws.cell(row=2, column=2, value="https://github.com")
    ws.cell(row=3, column=1, value="Docs")
    ws.cell(row=3, column=2, value="https://docs.microsoft.com")
    save(wb, "classic38_hyperlink_cell.xlsx")


# ── 39. Conditional-style financial table ────────────────────────────────
def classic39_financial_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    red = Font(color="FF0000")
    green = Font(color="008000")
    ws.append(["Month", "Budget", "Actual", "Variance"])
    data = [
        ("Jan", 10000, 9500, -500),
        ("Feb", 10000, 10800, 800),
        ("Mar", 10000, 9900, -100),
        ("Apr", 10000, 11200, 1200),
        ("May", 10000, 9700, -300),
        ("Jun", 10000, 10050, 50),
    ]
    for month, budget, actual, variance in data:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=month)
        ws.cell(row=row_idx, column=2, value=budget)
        ws.cell(row=row_idx, column=3, value=actual)
        var_cell = ws.cell(row=row_idx, column=4, value=variance)
        var_cell.font = green if variance >= 0 else red
    save(wb, "classic39_financial_table.xlsx")


# ── 40. Scientific notation numbers ──────────────────────────────────────
def classic40_scientific_notation():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Label", "Value"])
    ws.append(["Avogadro", 6.022e23])
    ws.append(["Planck", 6.626e-34])
    ws.append(["Speed of Light", 2.998e8])
    ws.append(["Electron mass", 9.109e-31])
    ws.append(["Pi approx", 3.14159265358979])
    save(wb, "classic40_scientific_notation.xlsx")


# ── 41. Integer vs float mixed ────────────────────────────────────────────
def classic41_integer_vs_float():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Type", "Value"])
    ws.append(["Integer", 42])
    ws.append(["Float", 42.0])
    ws.append(["NegInt", -7])
    ws.append(["NegFloat", -7.5])
    ws.append(["Zero", 0])
    ws.append(["ZeroFloat", 0.0])
    ws.append(["Large", 1000000])
    ws.append(["Small", 0.000001])
    save(wb, "classic41_integer_vs_float.xlsx")


# ── 42. Boolean values ────────────────────────────────────────────────────
def classic42_boolean_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Feature", "Enabled"])
    ws.append(["Dark Mode", True])
    ws.append(["Notifications", False])
    ws.append(["Auto-save", True])
    ws.append(["Analytics", False])
    ws.append(["Beta Features", True])
    save(wb, "classic42_boolean_values.xlsx")


# ── 43. Inventory report ─────────────────────────────────────────────────
def classic43_inventory_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    bold = Font(bold=True)
    headers = ["SKU", "Name", "Category", "Qty", "Unit Price", "Total Value"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
    inventory = [
        ("SKU001", "Widget A", "Widgets", 100, 5.99, 599.00),
        ("SKU002", "Widget B", "Widgets", 250, 3.49, 872.50),
        ("SKU003", "Gadget X", "Gadgets", 50, 29.99, 1499.50),
        ("SKU004", "Gadget Y", "Gadgets", 75, 19.99, 1499.25),
        ("SKU005", "Tool Z", "Tools", 30, 49.99, 1499.70),
        ("SKU006", "Part P", "Parts", 500, 0.99, 495.00),
        ("SKU007", "Part Q", "Parts", 1000, 0.49, 490.00),
    ]
    for row in inventory:
        ws.append(list(row))
    save(wb, "classic43_inventory_report.xlsx")


# ── 44. Employee roster ───────────────────────────────────────────────────
def classic44_employee_roster():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(["EmpID", "First", "Last", "Dept", "Title", "Email"])
    employees = [
        (1001, "Alice", "Smith", "Engineering", "Senior Engineer", "alice@example.com"),
        (1002, "Bob", "Jones", "Marketing", "Marketing Manager", "bob@example.com"),
        (1003, "Carol", "Williams", "HR", "HR Specialist", "carol@example.com"),
        (1004, "David", "Brown", "Engineering", "Junior Engineer", "david@example.com"),
        (1005, "Eve", "Davis", "Finance", "Financial Analyst", "eve@example.com"),
        (1006, "Frank", "Miller", "Sales", "Sales Representative", "frank@example.com"),
        (1007, "Grace", "Wilson", "Engineering", "Tech Lead", "grace@example.com"),
        (1008, "Henry", "Moore", "Support", "Support Specialist", "henry@example.com"),
    ]
    for emp in employees:
        ws.append(list(emp))
    save(wb, "classic44_employee_roster.xlsx")


# ── 45. Sales by region (multi-sheet) ────────────────────────────────────
def classic45_sales_by_region():
    wb = Workbook()
    regions = [
        ("North", [("Q1", 45000), ("Q2", 52000), ("Q3", 61000), ("Q4", 71000)]),
        ("South", [("Q1", 38000), ("Q2", 41000), ("Q3", 39000), ("Q4", 44000)]),
        ("East",  [("Q1", 55000), ("Q2", 59000), ("Q3", 63000), ("Q4", 68000)]),
        ("West",  [("Q1", 47000), ("Q2", 51000), ("Q3", 55000), ("Q4", 60000)]),
    ]
    first = True
    for region, quarters in regions:
        ws = wb.active if first else wb.create_sheet(region)
        if first:
            ws.title = region
            first = False
        ws.append(["Quarter", "Sales"])
        for q, s in quarters:
            ws.append([q, s])
    save(wb, "classic45_sales_by_region.xlsx")


# ── 46. Grade book ───────────────────────────────────────────────────────
def classic46_grade_book():
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"
    ws.append(["Student", "HW1", "HW2", "Midterm", "Final", "Total", "Grade"])
    students = [
        ("Alice",   95, 88, 92, 90, 365, "A"),
        ("Bob",     70, 75, 68, 72, 285, "C"),
        ("Carol",   85, 90, 88, 91, 354, "A-"),
        ("David",   60, 55, 62, 58, 235, "D"),
        ("Eve",     92, 95, 97, 98, 382, "A+"),
        ("Frank",   78, 80, 75, 82, 315, "B"),
        ("Grace",   88, 85, 89, 87, 349, "B+"),
    ]
    for s in students:
        ws.append(list(s))
    save(wb, "classic46_grade_book.xlsx")


# ── 47. Time series data ──────────────────────────────────────────────────
def classic47_time_series():
    wb = Workbook()
    ws = wb.active
    ws.title = "Temperatures"
    ws.append(["Day", "High", "Low", "Avg"])
    import random
    random.seed(42)
    for day in range(1, 32):
        # Use "Day-NN" label to avoid LibreOffice auto-converting ISO dates
        day_label = f"Day-{day:02d}"
        high = round(random.uniform(15, 30), 1)
        low = round(random.uniform(5, 15), 1)
        avg = round((high + low) / 2, 1)
        ws.append([day_label, high, low, avg])
    save(wb, "classic47_time_series.xlsx")


# ── 48. Survey results ───────────────────────────────────────────────────
def classic48_survey_results():
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"
    # Deliberately keep question text short (≤16 chars) so all 6 columns fit
    # on one page: Q-col natural width ≈ (16+2)*5 = 90pt, total ≈ 90+80+35+45+50+95
    # = 395pt + 5*20 padding = 495pt ≤ 512pt usable → single column group.
    ws.append(["Question", "StrongAgree", "Agree", "Neutral", "Disagree", "StrongDisagree"])
    questions = [
        ("Easy to use",   30, 45, 15, 7, 3),
        ("Recommend",     25, 40, 20, 10, 5),
        ("Fair price",    20, 35, 25, 15, 5),
        ("Good support",  35, 40, 15, 7, 3),
        ("Satisfied",     28, 42, 18, 8, 4),
    ]
    for q in questions:
        ws.append(list(q))
    save(wb, "classic48_survey_results.xlsx")


# ── 49. Contact list ──────────────────────────────────────────────────────
def classic49_contact_list():
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Name", "Phone", "Email", "City", "Country"])
    contacts = [
        ("Alice Smith",   "+1-555-0101", "alice@example.com",   "New York",  "USA"),
        ("Bob Jones",     "+44-20-7946-0958", "bob@example.co.uk", "London", "UK"),
        ("Carol Wang",    "+86-10-1234-5678", "carol@example.cn", "Beijing",  "China"),
        ("David Muller",  "+49-30-1234567",   "david@example.de", "Berlin",   "Germany"),
        ("Eve Martin",    "+33-1-23-45-67-89","eve@example.fr",   "Paris",    "France"),
        ("Frank Tanaka",  "+81-3-1234-5678",  "frank@example.jp","Tokyo",    "Japan"),
        ("Grace Kim",     "+82-2-1234-5678",  "grace@example.kr","Seoul",    "Korea"),
    ]
    for c in contacts:
        ws.append(list(c))
    save(wb, "classic49_contact_list.xlsx")


# ── 50. Budget vs actuals (three-sheet) ──────────────────────────────────
def classic50_budget_vs_actuals():
    wb = Workbook()

    # Sheet 1: Budget
    ws_budget = wb.active
    ws_budget.title = "Budget"
    ws_budget.append(["Department", "Q1", "Q2", "Q3", "Q4", "Annual"])
    budget_data = [
        ("Engineering", 200000, 200000, 210000, 220000, 830000),
        ("Marketing",   80000,  90000,  85000,  95000, 350000),
        ("Sales",       120000, 130000, 140000, 150000, 540000),
        ("HR",          40000,  40000,  42000,  43000, 165000),
        ("Finance",     35000,  35000,  37000,  38000, 145000),
    ]
    for row in budget_data:
        ws_budget.append(list(row))

    # Sheet 2: Actuals
    ws_actual = wb.create_sheet("Actuals")
    ws_actual.append(["Department", "Q1", "Q2", "Q3", "Q4", "Annual"])
    actual_data = [
        ("Engineering", 195000, 205000, 215000, 225000, 840000),
        ("Marketing",   82000,  88000,  91000,  97000, 358000),
        ("Sales",       118000, 135000, 142000, 148000, 543000),
        ("HR",          39000,  41000,  41500,  44000, 165500),
        ("Finance",     34000,  36000,  37500,  39000, 146500),
    ]
    for row in actual_data:
        ws_actual.append(list(row))

    # Sheet 3: Variance
    ws_var = wb.create_sheet("Variance")
    ws_var.append(["Department", "Q1", "Q2", "Q3", "Q4", "Annual"])
    for b, a in zip(budget_data, actual_data):
        ws_var.append([b[0]] + [a[i] - b[i] for i in range(1, 6)])

    save(wb, "classic50_budget_vs_actuals.xlsx")


# ── 51. Product catalog ───────────────────────────────────────────────────
def classic51_product_catalog():
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws.append(["Part#", "Name", "Description", "Weight(g)", "Price"])
    catalog = [
        ("P-001", "Basic Widget", "Standard widget for everyday use", 150, 4.99),
        ("P-002", "Pro Widget", "Enhanced widget with premium features", 180, 12.99),
        ("P-003", "Mini Gadget", "Compact gadget for mobile use", 90, 19.99),
        ("P-004", "Max Gadget", "Full-size gadget, industrial grade", 450, 89.99),
        ("P-005", "Connector A", "Type-A connector cable, 1m", 80, 7.49),
        ("P-006", "Connector B", "Type-B connector cable, 2m", 110, 9.99),
        ("P-007", "Adapter X", "Universal power adapter", 200, 15.99),
        ("P-008", "Adapter Y", "Travel power adapter", 120, 11.99),
        ("P-009", "Mount Bracket", "Wall mount bracket, steel", 600, 24.99),
        ("P-010", "Carry Case", "Padded carry case, waterproof", 350, 34.99),
    ]
    for row in catalog:
        ws.append(list(row))
    save(wb, "classic51_product_catalog.xlsx")


# ── 52. Pivot-style summary ───────────────────────────────────────────────
def classic52_pivot_summary():
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Region", "Electronics", "Furniture", "Clothing", "Food", "Total"])
    data = [
        ("North", 45000, 12000, 8000, 22000, 87000),
        ("South", 38000, 15000, 11000, 25000, 89000),
        ("East",  52000, 9000,  14000, 18000, 93000),
        ("West",  41000, 18000, 10000, 21000, 90000),
        ("Total", 176000, 54000, 43000, 86000, 359000),
    ]
    bold = Font(bold=True)
    for row_idx, row in enumerate(data, start=2):
        ws.append(list(row))
        if row_idx == len(data) + 1:  # Total row
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).font = bold
    save(wb, "classic52_pivot_summary.xlsx")


# ── 53. Invoice layout ────────────────────────────────────────────────────
def classic53_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    bold = Font(bold=True)
    # Header section
    ws["A1"] = "INVOICE"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Invoice #:"
    ws["B3"] = "INV-2025-0042"
    ws["A4"] = "Date:"
    ws["B4"] = "2025-03-01"
    ws["A5"] = "Due Date:"
    ws["B5"] = "2025-03-31"
    ws["A7"] = "Bill To:"
    ws["A7"].font = bold
    ws["A8"] = "ACME Corporation"
    ws["A9"] = "123 Business Rd, Suite 400"
    ws["A10"] = "New York, NY 10001"
    # Line items
    ws["A12"] = "Item"
    ws["B12"] = "Qty"
    ws["C12"] = "Unit Price"
    ws["D12"] = "Total"
    for col in range(1, 5):
        ws.cell(row=12, column=col).font = bold
    items = [
        ("Consulting Services", 10, 150.00, 1500.00),
        ("Software License", 5, 99.00, 495.00),
        ("Hardware", 2, 249.99, 499.98),
        ("Support Plan (annual)", 1, 1200.00, 1200.00),
    ]
    for r_idx, item in enumerate(items, start=13):
        ws.cell(row=r_idx, column=1, value=item[0])
        ws.cell(row=r_idx, column=2, value=item[1])
        ws.cell(row=r_idx, column=3, value=item[2])
        ws.cell(row=r_idx, column=4, value=item[3])
    ws.cell(row=18, column=3, value="Subtotal")
    ws.cell(row=18, column=4, value=3694.98)
    ws.cell(row=19, column=3, value="Tax (8%)")
    ws.cell(row=19, column=4, value=295.60)
    ws.cell(row=20, column=3, value="Total Due")
    ws.cell(row=20, column=3).font = bold
    ws.cell(row=20, column=4, value=3990.58)
    ws.cell(row=20, column=4).font = bold
    save(wb, "classic53_invoice.xlsx")


# ── 54. Multi-level header simulation ────────────────────────────────────
def classic54_multi_level_header():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bold = Font(bold=True)
    # Row 1: group headers
    ws["A1"] = ""
    ws["B1"] = "Q1"
    ws["D1"] = "Q2"
    ws["F1"] = "Q3"
    for cell_ref in ["B1", "D1", "F1"]:
        ws[cell_ref].font = bold
    # Row 2: sub-headers
    sub = ["ID", "Revenue", "Cost", "Revenue", "Cost", "Revenue", "Cost"]
    for col, h in enumerate(sub, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = bold
    # Data rows
    data = [
        (1, 50000, 30000, 55000, 32000, 60000, 35000),
        (2, 45000, 28000, 48000, 29000, 52000, 31000),
        (3, 60000, 35000, 65000, 37000, 70000, 40000),
    ]
    for row in data:
        ws.append(list(row))
    save(wb, "classic54_multi_level_header.xlsx")


# ── 55. Error / N/A values (as strings) ──────────────────────────────────
def classic55_error_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Metric", "Value", "Status"])
    ws.append(["Sales", 12345, "OK"])
    ws.append(["Revenue", "#N/A", "Missing"])
    ws.append(["Cost", "#REF!", "Broken ref"])
    ws.append(["Profit", "#DIV/0!", "Div by zero"])
    ws.append(["Units", "#VALUE!", "Wrong type"])
    ws.append(["Target", 15000, "OK"])
    save(wb, "classic55_error_values.xlsx")


# ── 56. Alternating row fill colors ──────────────────────────────────────
def classic56_alternating_row_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    fill_light = PatternFill(fill_type="solid", fgColor="DCE6F1")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    ws.append(["#", "Product", "Price"])
    for i in range(1, 11):
        row_num = ws.max_row + 1
        ws.append([i, f"Product {i}", i * 10.0])
        fill = fill_light if i % 2 == 0 else fill_white
        for col in range(1, 4):
            ws.cell(row=row_num, column=col).fill = fill
    save(wb, "classic56_alternating_row_colors.xlsx")


# ── 57. UTF-8 Asian characters (CJK only sheet) ──────────────────────────
def classic57_cjk_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "中文"
    ws.append(["序号", "产品名称", "价格", "库存"])
    ws.append([1, "笔记本电脑", 5999, 100])
    ws.append([2, "智能手机", 2999, 250])
    ws.append([3, "平板电脑", 1999, 150])
    ws.append([4, "蓝牙耳机", 299, 500])
    ws.append([5, "充电器", 99, 1000])
    save(wb, "classic57_cjk_only.xlsx")


# ── 58. Mixed numeric formats ─────────────────────────────────────────────
def classic58_mixed_numeric_formats():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Type", "Value"])
    ws.append(["Integer",           1000000])
    ws.append(["Float 2dp",         3.14])
    ws.append(["Float 5dp",         3.14159])
    ws.append(["Negative int",      -42])
    ws.append(["Negative float",    -3.14])
    ws.append(["Very small",        0.0001])
    ws.append(["Very large",        9999999.99])
    ws.append(["Zero",              0])
    ws.append(["Scientific approx", 1.23e10])
    save(wb, "classic58_mixed_numeric_formats.xlsx")


# ── 59. Multi-sheet with data on each plus a summary ─────────────────────
def classic59_multi_sheet_summary():
    wb = Workbook()

    months = ["Jan", "Feb", "Mar"]
    totals = []
    first = True
    for month in months:
        ws = wb.active if first else wb.create_sheet(month)
        if first:
            ws.title = month
            first = False
        ws.append(["Product", "Units", "Revenue"])
        revenue_total = 0
        for p in range(1, 6):
            units = (p * 7 + len(month)) % 50 + 10
            price = p * 5.0 + 9.99
            revenue = round(units * price, 2)
            revenue_total += revenue
            ws.append([f"Prod{p}", units, revenue])
        totals.append((month, revenue_total))

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Month", "Total Revenue"])
    for month, total in totals:
        ws_summary.append([month, round(total, 2)])

    save(wb, "classic59_multi_sheet_summary.xlsx")


# ── 60. Large wide table (20 columns × 50 rows) ───────────────────────────
def classic60_large_wide_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    num_cols = 20
    num_rows = 50
    headers = [f"Col{c+1:02d}" for c in range(num_cols)]
    ws.append(headers)
    for r in range(1, num_rows + 1):
        ws.append([f"R{r:02d}C{c+1:02d}" for c in range(num_cols)])
    save(wb, "classic60_large_wide_table.xlsx")


# ── 61. Product card with image ──────────────────────────────────────────
def classic61_product_card_with_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Product"
    bold = Font(bold=True)
    img = _make_jpeg_bytes(120, 120, (220, 80, 80))
    _add_image(ws, img, "A1", width_px=120, height_px=90)
    ws.row_dimensions[1].height = 68
    ws["D1"] = "Product Name"
    ws["D1"].font = bold
    ws["D2"] = "Widget Pro 3000"
    ws["D3"] = "Price"
    ws["D3"].font = bold
    ws["D4"] = "$29.99"
    ws["D5"] = "In Stock"
    ws["D6"] = 150
    save(wb, "classic61_product_card_with_image.xlsx")


# ── 62. Company logo header ───────────────────────────────────────────────
def classic62_company_logo_header():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True, size=14)
    logo = _make_jpeg_bytes(160, 60, (0, 80, 160))
    _add_image(ws, logo, "A1", width_px=160, height_px=60)
    ws.row_dimensions[1].height = 45
    ws["C1"] = "ACME Corporation"
    ws["C1"].font = bold
    ws["C2"] = "Annual Report 2025"
    ws.append([])
    ws.append(["Department", "Q1", "Q2", "Q3", "Q4"])
    for dept, q1, q2, q3, q4 in [
        ("Sales", 120, 135, 142, 160),
        ("Engineering", 85, 90, 95, 100),
        ("Marketing", 60, 65, 70, 75),
    ]:
        ws.append([dept, q1, q2, q3, q4])
    save(wb, "classic62_company_logo_header.xlsx")


# ── 63. Two products side by side ─────────────────────────────────────────
def classic63_two_products_side_by_side():
    wb = Workbook()
    ws = wb.active
    ws.title = "Compare"
    img_a = _make_jpeg_bytes(100, 100, (255, 140, 0))
    img_b = _make_jpeg_bytes(100, 100, (0, 160, 100))
    _add_image(ws, img_a, "A1", width_px=90, height_px=90)
    _add_image(ws, img_b, "D1", width_px=90, height_px=90)
    height_rows = 68
    for r in range(1, 7):
        ws.row_dimensions[r].height = height_rows // 6
    ws["A7"] = "Product A"
    ws["D7"] = "Product B"
    ws["A8"] = "Price: $19.99"
    ws["D8"] = "Price: $24.99"
    ws["A9"] = "Rating: 4.2"
    ws["D9"] = "Rating: 4.7"
    save(wb, "classic63_two_products_side_by_side.xlsx")


# ── 64. Employee directory with photo ────────────────────────────────────
def classic64_employee_directory_with_photo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Directory"
    bold = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="003366")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(["Photo", "Name", "Title", "Department", "Email"])
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    employees = [
        ("Alice Chen", "Engineer", "R&D", "alice@example.com", (100, 149, 237)),
        ("Bob Smith", "Manager", "Sales", "bob@example.com", (60, 179, 113)),
        ("Carol Wang", "Designer", "UX", "carol@example.com", (218, 112, 214)),
    ]
    for i, (name, title, dept, email, color) in enumerate(employees, start=2):
        photo = _make_jpeg_bytes(60, 60, color)
        _add_image(ws, photo, f"A{i}", width_px=50, height_px=45)
        ws.row_dimensions[i].height = 36
        ws[f"B{i}"] = name
        ws[f"C{i}"] = title
        ws[f"D{i}"] = dept
        ws[f"E{i}"] = email
    save(wb, "classic64_employee_directory_with_photo.xlsx")


# ── 65. Inventory with product photos ─────────────────────────────────────
def classic65_inventory_with_product_photos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    bold = Font(bold=True)
    ws.append(["Image", "SKU", "Name", "Qty", "Price"])
    ws[1][0].font = bold
    items = [
        ("SKU-001", "Red Widget", 50, 9.99, (200, 50, 50)),
        ("SKU-002", "Blue Gadget", 30, 14.99, (50, 80, 200)),
        ("SKU-003", "Green Tool", 100, 4.49, (50, 160, 50)),
        ("SKU-004", "Yellow Device", 25, 29.99, (220, 190, 0)),
        ("SKU-005", "Purple Gear", 75, 7.99, (140, 50, 200)),
    ]
    for i, (sku, name, qty, price, color) in enumerate(items, start=2):
        img = _make_jpeg_bytes(60, 60, color)
        _add_image(ws, img, f"A{i}", width_px=50, height_px=45)
        ws.row_dimensions[i].height = 36
        ws[f"B{i}"] = sku
        ws[f"C{i}"] = name
        ws[f"D{i}"] = qty
        ws[f"E{i}"] = price
    save(wb, "classic65_inventory_with_product_photos.xlsx")


# ── 66. Invoice with company logo ─────────────────────────────────────────
def classic66_invoice_with_logo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    logo = _make_jpeg_bytes(140, 55, (20, 60, 140))
    _add_image(ws, logo, "A1", width_px=140, height_px=55)
    ws.row_dimensions[1].height = 42
    bold = Font(bold=True)
    ws["D1"] = "INVOICE"
    ws["D1"].font = Font(bold=True, size=18)
    ws["D2"] = "Invoice #: INV-20250301"
    ws["D3"] = "Date: 2025-03-01"
    ws.append([])
    ws.append(["Description", "Qty", "Unit Price", "Total"])
    for cell in ws[5]:
        cell.font = bold
    ws.append(["Consulting Services", 8, 150.0, 1200.0])
    ws.append(["Software License", 1, 299.0, 299.0])
    ws.append(["Support Package", 1, 99.0, 99.0])
    ws.append([])
    ws.append(["", "", "Total", 1598.0])
    save(wb, "classic66_invoice_with_logo.xlsx")


# ── 67. Real-estate listing with photo ────────────────────────────────────
def classic67_real_estate_listing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Listing"
    house_img = _make_jpeg_bytes(200, 130, (180, 160, 130))
    _add_image(ws, house_img, "A1", width_px=200, height_px=130)
    for r in range(1, 11):
        ws.row_dimensions[r].height = 13
    bold = Font(bold=True)
    ws["D1"] = "123 Maple Street"
    ws["D1"].font = Font(bold=True, size=14)
    ws["D2"] = "Springfield, ST 12345"
    ws["D3"] = "List Price: $485,000"
    ws["D3"].font = bold
    ws.append([])
    ws.append(["Feature", "Detail"])
    for feat, detail in [
        ("Bedrooms", 4), ("Bathrooms", 2.5), ("Sq Ft", 2100),
        ("Lot Size", "0.25 acres"), ("Year Built", 1998),
    ]:
        ws.append([feat, detail])
    save(wb, "classic67_real_estate_listing.xlsx")


# ── 68. Restaurant menu with food photos ──────────────────────────────────
def classic68_restaurant_menu():
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    ws["A1"] = "Today's Menu"
    ws["A1"].font = Font(bold=True, size=16)
    ws.row_dimensions[1].height = 20
    menu_items = [
        ("Grilled Salmon", "$18.99", "Fresh Atlantic salmon with herbs", (240, 180, 100)),
        ("Caesar Salad", "$12.99", "Romaine lettuce, croutons, parmesan", (180, 220, 130)),
        ("Beef Burger", "$14.99", "8oz Angus beef, brioche bun", (190, 130, 80)),
        ("Pasta Primavera", "$13.99", "Seasonal vegetables, olive oil", (255, 220, 150)),
    ]
    row = 2
    for name, price, description, color in menu_items:
        # Text in columns A-C (left); image in column E (right)
        ws[f"A{row}"] = name
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = price
        ws[f"A{row + 1}"] = description
        ws.row_dimensions[row].height = 57
        img = _make_jpeg_bytes(80, 80, color)
        _add_image(ws, img, f"E{row}", width_px=80, height_px=75)
        row += 3
    save(wb, "classic68_restaurant_menu.xlsx")


# ── 69. Image-only sheet ──────────────────────────────────────────────────
def classic69_image_only_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    cover_img = _make_jpeg_bytes(300, 200, (40, 100, 180))
    _add_image(ws, cover_img, "A1", width_px=300, height_px=200)
    for r in range(1, 16):
        ws.row_dimensions[r].height = 14
    save(wb, "classic69_image_only_sheet.xlsx")


# ── 70. Product catalog with 3 images ─────────────────────────────────────
def classic70_product_catalog_with_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    bold = Font(bold=True)
    ws["A1"] = "Product Catalog - Spring 2025"
    ws["A1"].font = Font(bold=True, size=14)
    ws.row_dimensions[1].height = 20
    products = [
        ("Classic Pen", "$3.99", "A reliable ballpoint pen", (60, 100, 180)),
        ("Leather Notebook", "$12.99", "Premium A5 notebook", (140, 100, 60)),
        ("Desk Organizer", "$24.99", "Bamboo desk tidy set", (100, 160, 100)),
    ]
    row = 3
    for name, price, desc, color in products:
        img = _make_jpeg_bytes(100, 100, color)
        _add_image(ws, img, f"A{row}", width_px=90, height_px=90)
        ws.row_dimensions[row].height = 68
        ws[f"C{row}"] = name
        ws[f"C{row}"].font = bold
        ws[f"D{row}"] = price
        ws[f"C{row + 1}"] = desc
        row += 4
    save(wb, "classic70_product_catalog_with_images.xlsx")


# ── 71. Multi-sheet workbook, each sheet has an image ─────────────────────
def classic71_multi_sheet_with_images():
    wb = Workbook()
    sheets_data = [
        ("Overview", (30, 100, 180), [("Metric", "Value"), ("Revenue", 1_200_000), ("Cost", 800_000)]),
        ("Marketing", (180, 60, 60), [("Channel", "Budget"), ("Digital", 50_000), ("Print", 20_000)]),
        ("HR", (60, 160, 80), [("Department", "Headcount"), ("Engineering", 45), ("Sales", 30)]),
    ]
    for i, (title, color, data) in enumerate(sheets_data):
        ws = wb.active if i == 0 else wb.create_sheet(title)
        ws.title = title
        img = _make_jpeg_bytes(100, 70, color)
        _add_image(ws, img, "A1", width_px=100, height_px=70)
        ws.row_dimensions[1].height = 53
        for j, row_data in enumerate(data, start=2):
            ws.append(list(row_data))
    save(wb, "classic71_multi_sheet_with_images.xlsx")


# ── 72. Bar chart screenshot + data ───────────────────────────────────────
def classic72_bar_chart_image_with_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    bold = Font(bold=True)
    ws["A1"] = "Monthly Sales Data"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Month", "Revenue", "Target"])
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    revenues = [42000, 48000, 51000, 45000, 56000, 62000]
    targets = [45000, 47000, 50000, 50000, 54000, 60000]
    for m, r, t in zip(months, revenues, targets):
        ws.append([m, r, t])
    # Embed a simulated chart image
    chart_img = _make_jpeg_bytes(220, 140, (245, 245, 245))
    _add_image(ws, chart_img, "E3", width_px=220, height_px=140)
    save(wb, "classic72_bar_chart_image_with_data.xlsx")


# ── 73. Event flyer with banner ───────────────────────────────────────────
def classic73_event_flyer_with_banner():
    wb = Workbook()
    ws = wb.active
    ws.title = "Event"
    banner = _make_jpeg_bytes(300, 100, (80, 30, 140))
    _add_image(ws, banner, "A1", width_px=300, height_px=100)
    ws.row_dimensions[1].height = 76
    bold = Font(bold=True)
    ws["A8"] = "Tech Summit 2025"
    ws["A8"].font = Font(bold=True, size=16)
    ws["A9"] = "Date: April 15, 2025"
    ws["A10"] = "Venue: Convention Center Hall A"
    ws["A11"] = "Speakers: 20+ Industry Leaders"
    ws.append([])
    ws.append(["Time", "Session", "Speaker"])
    for row in [
        ("09:00", "Opening Keynote", "Dr. Jane Kim"),
        ("10:30", "AI in Practice", "Prof. Mark Liu"),
        ("13:00", "Cloud Architecture", "Eng. Sara Patel"),
        ("15:00", "Panel Discussion", "All Speakers"),
    ]:
        ws.append(list(row))
    save(wb, "classic73_event_flyer_with_banner.xlsx")


# ── 74. Dashboard with KPIs and screenshot ────────────────────────────────
def classic74_dashboard_with_kpi_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    bold = Font(bold=True)
    ws["A1"] = "Executive Dashboard Q1 2025"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["KPI", "Target", "Actual", "Status"])
    for cell in ws[3]:
        cell.font = bold
    kpis = [
        ("Revenue", 500000, 523000, "✓ Above"),
        ("New Customers", 200, 187, "✗ Below"),
        ("NPS Score", 70, 74, "✓ Above"),
        ("Churn Rate", "< 3%", "2.8%", "✓ Above"),
    ]
    for kpi in kpis:
        ws.append(list(kpi))
    dash_img = _make_jpeg_bytes(180, 120, (240, 248, 255))
    _add_image(ws, dash_img, "F2", width_px=180, height_px=120)
    save(wb, "classic74_dashboard_with_kpi_image.xlsx")


# ── 75. Certificate with seal image ───────────────────────────────────────
def classic75_certificate_with_seal():
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate"
    bg_fill = PatternFill(fill_type="solid", fgColor="FEFEFE")
    bold_large = Font(bold=True, size=20)
    ws["B2"] = "Certificate of Achievement"
    ws["B2"].font = bold_large
    ws["B4"] = "This certifies that"
    ws["B5"] = "Alice Johnson"
    ws["B5"].font = Font(bold=True, size=16)
    ws["B6"] = "has successfully completed the Advanced Python Training"
    ws["B7"] = "Issued: March 1, 2025"
    seal = _make_jpeg_bytes(90, 90, (200, 160, 0))
    _add_image(ws, seal, "F3", width_px=90, height_px=90)
    save(wb, "classic75_certificate_with_seal.xlsx")


# ── 76. Four product images in a grid ─────────────────────────────────────
def classic76_product_image_grid():
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid"
    ws["A1"] = "Best Sellers"
    ws["A1"].font = Font(bold=True, size=14)
    # 2x2 grid compact layout: products in cols A-B and D-E to stay within page width
    products = [
        ("A3", "B3", (220, 80, 80),  "Red Phone Case",  "$9.99"),
        ("D3", "E3", (80, 120, 220), "Blue Speakers",   "$49.99"),
        ("A9", "B9", (80, 180, 80),  "Green Backpack",  "$34.99"),
        ("D9", "E9", (220, 160, 0),  "Yellow Headset",  "$29.99"),
    ]
    for img_anchor, text_anchor, color, name, price in products:
        img = _make_jpeg_bytes(80, 80, color)
        _add_image(ws, img, img_anchor, width_px=80, height_px=80)
        row_num = int("".join(filter(str.isdigit, text_anchor)))
        ws[text_anchor] = name
        ws[text_anchor].font = Font(bold=True)
        next_cell = text_anchor[0] + str(row_num + 1)
        ws[next_cell] = price
    save(wb, "classic76_product_image_grid.xlsx")


# ── 77. News article layout with hero image ───────────────────────────────
def classic77_news_article_with_hero_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Article"
    hero = _make_jpeg_bytes(280, 140, (100, 130, 160))
    _add_image(ws, hero, "A1", width_px=280, height_px=140)
    ws.row_dimensions[1].height = 107
    bold = Font(bold=True)
    ws["A11"] = "AI Transforms Modern Workplaces"
    ws["A11"].font = Font(bold=True, size=14)
    ws["A12"] = "By Jane Reporter | March 1, 2025"
    ws["A12"].font = Font(italic=True, color="808080")
    paragraphs = [
        "Artificial intelligence is reshaping how businesses operate globally.",
        "From automated workflows to intelligent decision support, the impact is clear.",
        "Companies adopting AI early report 30% productivity gains on average.",
        "Experts predict continued acceleration through 2030 and beyond.",
    ]
    for i, para in enumerate(paragraphs, start=14):
        ws[f"A{i}"] = para
    save(wb, "classic77_news_article_with_hero_image.xlsx")


# ── 78. Small icon in every row ───────────────────────────────────────────
def classic78_small_icon_per_row():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    bold = Font(bold=True)
    ws.append(["Icon", "Task", "Assignee", "Status"])
    for cell in ws[1]:
        cell.font = bold
    tasks = [
        ("Fix login bug", "Alice", "Done", (80, 200, 80)),
        ("Write unit tests", "Bob", "In Progress", (255, 165, 0)),
        ("Deploy to staging", "Carol", "Pending", (200, 80, 80)),
        ("Code review PR #42", "Alice", "Done", (80, 200, 80)),
        ("Update docs", "Dave", "In Progress", (255, 165, 0)),
    ]
    for i, (task, assignee, status, color) in enumerate(tasks, start=2):
        icon = _make_jpeg_bytes(30, 30, color)
        _add_image(ws, icon, f"A{i}", width_px=25, height_px=25)
        ws.row_dimensions[i].height = 20
        ws[f"B{i}"] = task
        ws[f"C{i}"] = assignee
        ws[f"D{i}"] = status
    save(wb, "classic78_small_icon_per_row.xlsx")


# ── 79. Wide panoramic banner ─────────────────────────────────────────────
def classic79_wide_panoramic_banner():
    wb = Workbook()
    ws = wb.active
    ws.title = "Banner"
    banner = _make_jpeg_bytes(480, 80, (10, 60, 120))
    _add_image(ws, banner, "A1", width_px=480, height_px=80)
    ws.row_dimensions[1].height = 61
    ws["A9"] = "Product Launch 2025"
    ws["A9"].font = Font(bold=True, size=18)
    ws["A10"] = "Introducing the next generation of innovation."
    ws["A11"] = "Available starting April 1, 2025"
    ws.append([])
    ws.append(["Model", "Storage", "RAM", "Price"])
    for cfg in [("Pro", "256GB", "16GB", "$999"), ("Max", "512GB", "32GB", "$1499")]:
        ws.append(list(cfg))
    save(wb, "classic79_wide_panoramic_banner.xlsx")


# ── 80. Portrait tall image ───────────────────────────────────────────────
def classic80_portrait_tall_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Portrait"
    portrait = _make_jpeg_bytes(100, 200, (160, 200, 240))
    _add_image(ws, portrait, "A1", width_px=90, height_px=180)
    ws["D1"] = "Profile"
    ws["D1"].font = Font(bold=True, size=14)
    ws["D2"] = "Name: Dr. Emily Zhao"
    ws["D3"] = "Title: Chief Scientist"
    ws["D4"] = "Dept: Research & Innovation"
    ws["D5"] = "Location: Singapore Office"
    ws["D6"] = "Email: emily.zhao@example.com"
    ws["D7"] = "LinkedIn: linkedin.com/in/ezhao"
    save(wb, "classic80_portrait_tall_image.xlsx")


# ── 81. Step-by-step guide with images ────────────────────────────────────
def classic81_step_by_step_with_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "Guide"
    ws["A1"] = "Quick Start Guide"
    ws["A1"].font = Font(bold=True, size=14)
    steps = [
        ("Step 1: Unbox", "Remove the product from packaging.", (200, 230, 255)),
        ("Step 2: Charge", "Connect USB-C cable, charge 2 hours.", (200, 255, 210)),
        ("Step 3: Power On", "Hold power button 3 seconds.", (255, 235, 200)),
        ("Step 4: Configure", "Follow on-screen setup wizard.", (240, 200, 255)),
    ]
    row = 3
    for title, desc, color in steps:
        step_img = _make_jpeg_bytes(80, 80, color)
        _add_image(ws, step_img, f"A{row}", width_px=75, height_px=75)
        ws.row_dimensions[row].height = 57
        ws[f"C{row}"] = title
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"C{row + 1}"] = desc
        row += 4
    save(wb, "classic81_step_by_step_with_images.xlsx")


# ── 82. Before/After comparison with images ───────────────────────────────
def classic82_before_after_images():
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    bold = Font(bold=True)
    ws["A1"] = "Before"
    ws["A1"].font = bold
    ws["D1"] = "After"
    ws["D1"].font = bold
    before_img = _make_jpeg_bytes(130, 130, (180, 140, 130))
    after_img = _make_jpeg_bytes(130, 130, (100, 180, 160))
    _add_image(ws, before_img, "A2", width_px=130, height_px=130)
    _add_image(ws, after_img, "D2", width_px=130, height_px=130)
    for r in range(2, 12):
        ws.row_dimensions[r].height = 13
    ws["A12"] = "Old design – legacy UI"
    ws["D12"] = "New design – modern UI"
    ws.append([])
    ws.append(["Metric", "Before", "After", "Delta"])
    for m, b, a, d in [("Load time", "4.2s", "1.1s", "-74%"), ("Conversion", "2.1%", "4.8%", "+129%")]:
        ws.append([m, b, a, d])
    save(wb, "classic82_before_after_images.xlsx")


# ── 83. Color swatch palette ──────────────────────────────────────────────
def classic83_color_swatch_palette():
    wb = Workbook()
    ws = wb.active
    ws.title = "Palette"
    ws["A1"] = "Brand Color Palette"
    ws["A1"].font = Font(bold=True, size=13)
    colors = [
        ("Primary Blue", (0, 82, 165)),
        ("Primary Red", (197, 27, 50)),
        ("Accent Green", (0, 163, 108)),
        ("Neutral Grey", (128, 128, 128)),
        ("Warm Yellow", (255, 193, 7)),
        ("Dark Navy", (10, 30, 70)),
    ]
    row = 3
    for name, rgb in colors:
        swatch = _make_jpeg_bytes(60, 40, rgb)
        _add_image(ws, swatch, f"A{row}", width_px=55, height_px=38)
        ws.row_dimensions[row].height = 29
        ws[f"C{row}"] = name
        ws[f"D{row}"] = f"RGB({rgb[0]}, {rgb[1]}, {rgb[2]})"
        row += 2
    save(wb, "classic83_color_swatch_palette.xlsx")


# ── 84. Travel destination cards ──────────────────────────────────────────
def classic84_travel_destination_cards():
    wb = Workbook()
    ws = wb.active
    ws.title = "Destinations"
    ws["A1"] = "Top Travel Destinations 2025"
    ws["A1"].font = Font(bold=True, size=14)
    destinations = [
        ("Kyoto, Japan", "Cherry blossoms and ancient temples", (180, 120, 150)),
        ("Prague, Czech Republic", "Medieval architecture and vibrant culture", (130, 160, 200)),
        ("Cape Town, South Africa", "Mountains, ocean, and wildlife", (160, 200, 140)),
    ]
    row = 3
    for place, desc, color in destinations:
        dest_img = _make_jpeg_bytes(140, 90, color)
        _add_image(ws, dest_img, f"A{row}", width_px=130, height_px=90)
        ws.row_dimensions[row].height = 68
        ws[f"D{row}"] = place
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"D{row + 1}"] = desc
        row += 5
    save(wb, "classic84_travel_destination_cards.xlsx")


# ── 85. Science lab results with specimen image ───────────────────────────
def classic85_lab_results_with_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lab Results"
    bold = Font(bold=True)
    ws["A1"] = "Sample Analysis Report"
    ws["A1"].font = Font(bold=True, size=13)
    specimen_img = _make_jpeg_bytes(120, 100, (210, 220, 230))
    _add_image(ws, specimen_img, "E2", width_px=120, height_px=100)
    ws.append([])
    ws.append(["Parameter", "Value", "Unit", "Reference Range", "Flag"])
    for cell in ws[3]:
        cell.font = bold
    results = [
        ("pH", 7.35, "", "7.35 – 7.45", "Normal"),
        ("Glucose", 5.2, "mmol/L", "3.9 – 5.5", "Normal"),
        ("Sodium", 142, "mEq/L", "136 – 145", "Normal"),
        ("Potassium", 5.0, "mEq/L", "3.5 – 5.0", "Normal"),
        ("Creatinine", 1.4, "mg/dL", "0.6 – 1.2", "High"),
    ]
    for r in results:
        ws.append(list(r))
    save(wb, "classic85_lab_results_with_image.xlsx")


# ── 86. Software screenshot + feature list ────────────────────────────────
def classic86_software_screenshot_features():
    wb = Workbook()
    ws = wb.active
    ws.title = "Features"
    screenshot = _make_jpeg_bytes(200, 150, (230, 240, 250))
    _add_image(ws, screenshot, "A1", width_px=200, height_px=150)
    ws["D1"] = "MiniApp v2.0"
    ws["D1"].font = Font(bold=True, size=15)
    ws["D2"] = "The fastest lightweight app"
    for r in range(1, 13):
        ws.row_dimensions[r].height = 12
    ws.append([])
    ws.append(["Feature", "Available"])
    bold = Font(bold=True)
    ws[ws.max_row][0].font = bold
    features = [
        ("Dark Mode", "Yes"), ("Auto Save", "Yes"), ("Cloud Sync", "Yes"),
        ("Offline Mode", "Yes"), ("API Access", "Pro only"), ("Export to PDF", "Yes"),
    ]
    for f in features:
        ws.append(list(f))
    save(wb, "classic86_software_screenshot_features.xlsx")


# ── 87. Sports results with team logos ────────────────────────────────────
def classic87_sports_results_with_logos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    bold = Font(bold=True)
    ws["A1"] = "League Standings - Season 2025"
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(["Logo", "Team", "W", "L", "D", "Pts"])
    for cell in ws[3]:
        cell.font = bold
    teams = [
        ("Eagles", 18, 4, 2, 56, (0, 80, 180)),
        ("Tigers", 15, 7, 2, 47, (220, 120, 0)),
        ("Sharks", 13, 8, 3, 42, (0, 160, 200)),
        ("Wolves", 10, 10, 4, 34, (120, 30, 30)),
    ]
    row = 4
    for team, w, l, d, pts, color in teams:
        logo = _make_jpeg_bytes(40, 40, color)
        _add_image(ws, logo, f"A{row}", width_px=35, height_px=32)
        ws.row_dimensions[row].height = 26
        ws[f"B{row}"] = team
        ws[f"C{row}"] = w
        ws[f"D{row}"] = l
        ws[f"E{row}"] = d
        ws[f"F{row}"] = pts
        row += 1
    save(wb, "classic87_sports_results_with_logos.xlsx")


# ── 88. Image after data rows ─────────────────────────────────────────────
def classic88_image_after_data():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    bold = Font(bold=True)
    ws.append(["Quarter", "Revenue", "Expenses", "Profit"])
    for cell in ws[1]:
        cell.font = bold
    for q, r, e in [("Q1", 120000, 80000), ("Q2", 135000, 88000),
                    ("Q3", 142000, 91000), ("Q4", 160000, 95000)]:
        ws.append([q, r, e, r - e])
    ws.append([])
    footer_img = _make_jpeg_bytes(200, 80, (220, 235, 255))
    _add_image(ws, footer_img, "A8", width_px=200, height_px=80)
    ws["D8"] = "Prepared by Finance Team"
    ws["D9"] = "Confidential - Q4 2025"
    save(wb, "classic88_image_after_data.xlsx")


# ── 89. Nutrition label with product image ────────────────────────────────
def classic89_nutrition_label_with_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Nutrition"
    bold = Font(bold=True)
    product_img = _make_jpeg_bytes(100, 130, (230, 200, 160))
    _add_image(ws, product_img, "A1", width_px=95, height_px=125)
    ws["C1"] = "Nutrition Facts"
    ws["C1"].font = Font(bold=True, size=14)
    ws["C2"] = "Serving Size: 30g (approx. 1 cup)"
    ws.append([])
    ws.append(["Nutrient", "Amount per serving", "% Daily Value"])
    for cell in ws[4]:
        cell.font = bold
    nutrients = [
        ("Calories", "120 kcal", ""),
        ("Total Fat", "3g", "4%"),
        ("Saturated Fat", "0.5g", "3%"),
        ("Sodium", "160mg", "7%"),
        ("Total Carbohydrate", "22g", "8%"),
        ("Dietary Fiber", "3g", "11%"),
        ("Sugars", "4g", ""),
        ("Protein", "3g", ""),
    ]
    for n in nutrients:
        ws.append(list(n))
    save(wb, "classic89_nutrition_label_with_image.xlsx")


# ── 90. Project status with milestone images ──────────────────────────────
def classic90_project_status_with_milestones():
    wb = Workbook()
    ws = wb.active
    ws.title = "Project"
    bold = Font(bold=True)
    ws["A1"] = "Project Orion – Status Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Reporting Period: Q1 2025"
    ws.append([])
    ws.append(["Milestone", "Due Date", "Owner", "Status"])
    for cell in ws[4]:
        cell.font = bold
    milestones = [
        ("Requirements Freeze", "Jan 15", "PM Team", "Complete"),
        ("Architecture Review", "Feb 1", "Tech Lead", "Complete"),
        ("Alpha Release", "Feb 28", "Dev Team", "In Progress"),
        ("Beta Testing", "Mar 31", "QA Team", "Not Started"),
        ("Production Deploy", "Apr 15", "DevOps", "Not Started"),
    ]
    for m in milestones:
        ws.append(list(m))
    ws.append([])
    team_img = _make_jpeg_bytes(160, 100, (200, 220, 200))
    _add_image(ws, team_img, "F4", width_px=160, height_px=100)
    save(wb, "classic90_project_status_with_milestones.xlsx")


# ── 91. Simple vertical bar chart ─────────────────────────────────────────
def classic91_simple_bar_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Product", "Revenue"])
    for prod, rev in [("Widget A", 12000), ("Widget B", 18500),
                      ("Widget C", 9200), ("Widget D", 22000), ("Widget E", 15600)]:
        ws.append([prod, rev])
    chart = BarChart()
    chart.type = "col"
    chart.title = "Product Revenue"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Product"
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic91_simple_bar_chart.xlsx")


# ── 92. Horizontal bar chart ─────────────────────────────────────────────
def classic92_horizontal_bar_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Departments"
    ws.append(["Department", "Headcount"])
    for dept, hc in [("Engineering", 45), ("Sales", 30), ("Marketing", 18),
                     ("HR", 12), ("Finance", 15), ("Operations", 25)]:
        ws.append([dept, hc])
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Headcount by Department"
    data = Reference(ws, min_col=2, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic92_horizontal_bar_chart.xlsx")


# ── 93. Line chart ──────────────────────────────────────────────────────
def classic93_line_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Temperature"
    ws.append(["Month", "Avg Temp (C)"])
    temps = [3, 5, 10, 15, 20, 25, 28, 27, 22, 15, 8, 4]
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for m, t in zip(months, temps):
        ws.append([m, t])
    chart = LineChart()
    chart.title = "Monthly Average Temperature"
    chart.y_axis.title = "Temperature (C)"
    data = Reference(ws, min_col=2, min_row=1, max_row=13)
    cats = Reference(ws, min_col=1, min_row=2, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic93_line_chart.xlsx")


# ── 94. Pie chart ────────────────────────────────────────────────────────
def classic94_pie_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Market"
    ws.append(["Segment", "Share (%)"])
    for seg, sh in [("Enterprise", 35), ("SMB", 28), ("Consumer", 22),
                    ("Government", 10), ("Education", 5)]:
        ws.append([seg, sh])
    chart = PieChart()
    chart.title = "Market Share by Segment"
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic94_pie_chart.xlsx")


# ── 95. Area chart ──────────────────────────────────────────────────────
def classic95_area_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Traffic"
    ws.append(["Hour", "Users"])
    for h in range(0, 24):
        users = int(200 + 800 * (1.0 / (1 + abs(h - 14) ** 1.5)))
        ws.append([f"{h:02d}:00", users])
    chart = AreaChart()
    chart.title = "Website Traffic by Hour"
    chart.y_axis.title = "Users"
    data = Reference(ws, min_col=2, min_row=1, max_row=25)
    cats = Reference(ws, min_col=1, min_row=2, max_row=25)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic95_area_chart.xlsx")


# ── 96. Scatter (XY) chart ──────────────────────────────────────────────
def classic96_scatter_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Correlation"
    ws.append(["Ad Spend ($K)", "Sales ($K)"])
    import random
    random.seed(42)
    for _ in range(20):
        spend = random.randint(5, 50)
        sales = int(spend * 2.3 + random.randint(-10, 10))
        ws.append([spend, sales])
    chart = ScatterChart()
    chart.title = "Ad Spend vs Sales"
    chart.x_axis.title = "Ad Spend ($K)"
    chart.y_axis.title = "Sales ($K)"
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=22)
    yvalues = Reference(ws, min_col=2, min_row=2, max_row=22)
    series = chart.series
    from openpyxl.chart import Series
    s = Series(yvalues, xvalues, title="Data Points")
    chart.series.append(s)
    chart.width = 14
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic96_scatter_chart.xlsx")


# ── 97. Doughnut chart ──────────────────────────────────────────────────
def classic97_doughnut_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws.append(["Category", "Amount"])
    for cat, amt in [("Salaries", 50000), ("Rent", 12000), ("Marketing", 8000),
                     ("R&D", 15000), ("Other", 5000)]:
        ws.append([cat, amt])
    chart = DoughnutChart()
    chart.title = "Budget Allocation"
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic97_doughnut_chart.xlsx")


# ── 98. Radar chart ─────────────────────────────────────────────────────
def classic98_radar_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Skills"
    ws.append(["Skill", "Score"])
    for skill, score in [("Python", 9), ("SQL", 8), ("Communication", 7),
                         ("Leadership", 6), ("Design", 5), ("DevOps", 7)]:
        ws.append([skill, score])
    chart = RadarChart()
    chart.title = "Developer Skill Radar"
    data = Reference(ws, min_col=2, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 12
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic98_radar_chart.xlsx")


# ── 99. Bubble chart ────────────────────────────────────────────────────
def classic99_bubble_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["Price ($)", "Rating", "Units Sold"])
    for price, rating, units in [(10, 4.2, 500), (25, 4.5, 300),
                                  (50, 3.8, 150), (15, 4.0, 420),
                                  (35, 4.7, 200), (8, 3.5, 600)]:
        ws.append([price, rating, units])
    chart = BubbleChart()
    chart.title = "Product Comparison"
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
    yvalues = Reference(ws, min_col=2, min_row=2, max_row=7)
    bubbles = Reference(ws, min_col=3, min_row=2, max_row=7)
    from openpyxl.chart import Series
    s = Series(yvalues, xvalues, zvalues=bubbles, title="Products")
    chart.series.append(s)
    chart.x_axis.title = "Price ($)"
    chart.y_axis.title = "Rating"
    chart.width = 14
    chart.height = 10
    ws.add_chart(chart, "E2")
    save(wb, "classic99_bubble_chart.xlsx")


# ── 100. Stacked bar chart ──────────────────────────────────────────────
def classic100_stacked_bar_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Quarters"
    ws.append(["Region", "Q1", "Q2", "Q3", "Q4"])
    for region, q1, q2, q3, q4 in [("North", 30, 40, 35, 50),
                                     ("South", 25, 30, 45, 40),
                                     ("East", 40, 35, 30, 45),
                                     ("West", 20, 25, 40, 35)]:
        ws.append([region, q1, q2, q3, q4])
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Quarterly Revenue by Region"
    data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "A8")
    save(wb, "classic100_stacked_bar_chart.xlsx")


# ── 101. 100% stacked bar chart ─────────────────────────────────────────
def classic101_percent_stacked_bar():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sources"
    ws.append(["Year", "Organic", "Paid", "Referral", "Direct"])
    for yr, org, paid, ref, direct in [(2021, 40, 25, 20, 15),
                                        (2022, 38, 30, 18, 14),
                                        (2023, 35, 32, 20, 13),
                                        (2024, 33, 35, 18, 14),
                                        (2025, 30, 38, 17, 15)]:
        ws.append([yr, org, paid, ref, direct])
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "percentStacked"
    chart.title = "Traffic Source Mix by Year"
    data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "A9")
    save(wb, "classic101_percent_stacked_bar.xlsx")


# ── 102. Line chart with markers ────────────────────────────────────────
def classic102_line_chart_with_markers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Growth"
    ws.append(["Year", "Users (K)", "Revenue (K)"])
    for yr, users, rev in [(2020, 10, 50), (2021, 25, 120), (2022, 55, 280),
                            (2023, 90, 500), (2024, 140, 780), (2025, 200, 1100)]:
        ws.append([yr, users, rev])
    chart = LineChart()
    chart.title = "Company Growth"
    chart.y_axis.title = "Value (K)"
    chart.style = 10
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for s in chart.series:
        s.graphicalProperties.line.width = 25000
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "E2")
    save(wb, "classic102_line_chart_with_markers.xlsx")


# ── 103. Pie chart with data labels ─────────────────────────────────────
def classic103_pie_chart_with_labels():
    wb = Workbook()
    ws = wb.active
    ws.title = "OS"
    ws.append(["OS", "Share (%)"])
    for os_name, share in [("Windows", 42), ("macOS", 28), ("Linux", 15),
                            ("ChromeOS", 10), ("Other", 5)]:
        ws.append([os_name, share])
    chart = PieChart()
    chart.title = "Desktop OS Market Share"
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.width = 13
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic103_pie_chart_with_labels.xlsx")


# ── 104. Combo chart (bar + line) ───────────────────────────────────────
def classic104_combo_bar_line_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Combo"
    ws.append(["Month", "Sales", "Target"])
    for m, s, t in [("Jan", 42, 45), ("Feb", 48, 47), ("Mar", 51, 50),
                     ("Apr", 45, 50), ("May", 56, 54), ("Jun", 62, 60)]:
        ws.append([m, s, t])
    bar = BarChart()
    bar.type = "col"
    bar.title = "Sales vs Target"
    bar_data = Reference(ws, min_col=2, min_row=1, max_row=7)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=2, max_row=7))
    line = LineChart()
    line_data = Reference(ws, min_col=3, min_row=1, max_row=7)
    line.add_data(line_data, titles_from_data=True)
    line.y_axis.axId = 200
    bar += line
    bar.width = 16
    bar.height = 10
    ws.add_chart(bar, "E2")
    save(wb, "classic104_combo_bar_line_chart.xlsx")


# ── 105. 3D bar chart ───────────────────────────────────────────────────
def classic105_3d_bar_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Regions"
    ws.append(["Region", "2024", "2025"])
    for rg, v24, v25 in [("APAC", 120, 145), ("EMEA", 95, 110),
                          ("Americas", 150, 175), ("LATAM", 40, 55)]:
        ws.append([rg, v24, v25])
    chart = BarChart3D()
    chart.title = "Revenue by Region (3D)"
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "E2")
    save(wb, "classic105_3d_bar_chart.xlsx")


# ── 106. 3D pie chart ───────────────────────────────────────────────────
def classic106_3d_pie_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(["Category", "Amount"])
    for cat, amt in [("Food", 800), ("Housing", 1500), ("Transport", 400),
                     ("Entertainment", 300), ("Savings", 700), ("Other", 200)]:
        ws.append([cat, amt])
    chart = PieChart3D()
    chart.title = "Monthly Expense Breakdown (3D)"
    data = Reference(ws, min_col=2, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 13
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic106_3d_pie_chart.xlsx")


# ── 107. Multi-series line chart ────────────────────────────────────────
def classic107_multi_series_line():
    wb = Workbook()
    ws = wb.active
    ws.title = "Stocks"
    ws.append(["Day", "AAPL", "GOOG", "MSFT"])
    import random
    random.seed(107)
    prices = {"AAPL": 180, "GOOG": 140, "MSFT": 400}
    for d in range(1, 21):
        row = [f"Day {d}"]
        for ticker in ["AAPL", "GOOG", "MSFT"]:
            prices[ticker] += random.uniform(-3, 3)
            row.append(round(prices[ticker], 2))
        ws.append(row)
    chart = LineChart()
    chart.title = "Stock Price Trend (20 Days)"
    chart.y_axis.title = "Price ($)"
    data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=21)
    cats = Reference(ws, min_col=1, min_row=2, max_row=21)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "F2")
    save(wb, "classic107_multi_series_line.xlsx")


# ── 108. Stacked area chart ─────────────────────────────────────────────
def classic108_stacked_area_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Channels"
    ws.append(["Month", "Email", "Social", "Search", "Direct"])
    data_rows = [
        ("Jan", 120, 80, 200, 100), ("Feb", 130, 90, 210, 105),
        ("Mar", 125, 110, 230, 115), ("Apr", 140, 120, 250, 120),
        ("May", 150, 130, 240, 125), ("Jun", 160, 140, 260, 130),
    ]
    for row in data_rows:
        ws.append(list(row))
    chart = AreaChart()
    chart.grouping = "stacked"
    chart.title = "Traffic by Channel (Stacked)"
    data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "A10")
    save(wb, "classic108_stacked_area_chart.xlsx")


# ── 109. Scatter with trend line ─────────────────────────────────────────
def classic109_scatter_with_trendline():
    wb = Workbook()
    ws = wb.active
    ws.title = "Study"
    ws.append(["Study Hours", "Exam Score"])
    import random
    random.seed(109)
    for _ in range(15):
        hrs = random.randint(1, 10)
        score = min(100, max(30, int(hrs * 8 + 20 + random.randint(-8, 8))))
        ws.append([hrs, score])
    chart = ScatterChart()
    chart.title = "Study Hours vs Exam Score"
    chart.x_axis.title = "Hours"
    chart.y_axis.title = "Score"
    xv = Reference(ws, min_col=1, min_row=2, max_row=16)
    yv = Reference(ws, min_col=2, min_row=2, max_row=16)
    from openpyxl.chart import Series
    s = Series(yv, xv, title="Students")
    from openpyxl.chart.trendline import Trendline
    s.trendline = Trendline(trendlineType="linear")
    chart.series.append(s)
    chart.width = 14
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic109_scatter_with_trendline.xlsx")


# ── 110. Chart with title and legend ─────────────────────────────────────
def classic110_chart_with_legend():
    wb = Workbook()
    ws = wb.active
    ws.title = "Browser"
    ws.append(["Browser", "2024 (%)", "2025 (%)"])
    for br, v24, v25 in [("Chrome", 65, 62), ("Safari", 18, 20),
                          ("Firefox", 8, 7), ("Edge", 6, 8), ("Other", 3, 3)]:
        ws.append([br, v24, v25])
    chart = BarChart()
    chart.type = "col"
    chart.title = "Browser Market Share Comparison"
    chart.y_axis.title = "Market Share (%)"
    chart.legend.position = "b"
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "E2")
    save(wb, "classic110_chart_with_legend.xlsx")


# ── 111. Chart with axis labels ──────────────────────────────────────────
def classic111_chart_with_axis_labels():
    wb = Workbook()
    ws = wb.active
    ws.title = "Emissions"
    ws.append(["Country", "CO2 (Mt)"])
    for country, co2 in [("China", 10500), ("USA", 5000), ("India", 2700),
                          ("Russia", 1700), ("Japan", 1100), ("Germany", 700)]:
        ws.append([country, co2])
    chart = BarChart()
    chart.type = "bar"
    chart.title = "CO2 Emissions by Country"
    chart.x_axis.title = "Country"
    chart.y_axis.title = "CO2 Emissions (Megatons)"
    chart.y_axis.numFmt = "#,##0"
    data = Reference(ws, min_col=2, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic111_chart_with_axis_labels.xlsx")


# ── 112. Multiple charts on one sheet ────────────────────────────────────
def classic112_multiple_charts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.append(["Month", "Revenue", "Costs", "Profit"])
    for m, r, c in [("Jan", 50, 30), ("Feb", 55, 32), ("Mar", 60, 35),
                     ("Apr", 52, 28), ("May", 70, 40), ("Jun", 75, 42)]:
        ws.append([m, r, c, r - c])
    # Bar chart for revenue
    bar = BarChart()
    bar.type = "col"
    bar.title = "Revenue & Costs"
    bar_data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=7)
    bar_cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.width = 14
    bar.height = 9
    ws.add_chart(bar, "F2")
    # Line chart for profit
    line = LineChart()
    line.title = "Profit Trend"
    line_data = Reference(ws, min_col=4, min_row=1, max_row=7)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(bar_cats)
    line.width = 14
    line.height = 9
    ws.add_chart(line, "F18")
    save(wb, "classic112_multiple_charts.xlsx")


# ── 113. Chart on separate chart sheet ───────────────────────────────────
def classic113_chart_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Quarter", "Revenue"])
    for q, r in [("Q1", 250), ("Q2", 310), ("Q3", 285), ("Q4", 400)]:
        ws.append([q, r])
    chart = BarChart()
    chart.type = "col"
    chart.title = "Quarterly Revenue"
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 12
    # Place chart on data sheet (openpyxl chartsheet API is limited)
    ws.add_chart(chart, "D2")
    save(wb, "classic113_chart_sheet.xlsx")


# ── 114. Chart with large dataset ────────────────────────────────────────
def classic114_chart_large_dataset():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Day", "Value"])
    import random
    random.seed(114)
    val = 100
    for d in range(1, 101):
        val += random.uniform(-5, 6)
        ws.append([d, round(val, 1)])
    chart = LineChart()
    chart.title = "100-Day Value Trend"
    data = Reference(ws, min_col=2, min_row=1, max_row=101)
    cats = Reference(ws, min_col=1, min_row=2, max_row=101)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic114_chart_large_dataset.xlsx")


# ── 115. Chart with negative values ─────────────────────────────────────
def classic115_chart_negative_values():
    wb = Workbook()
    ws = wb.active
    ws.title = "PnL"
    ws.append(["Month", "Profit/Loss"])
    for m, pl in [("Jan", 15), ("Feb", -8), ("Mar", 22), ("Apr", -3),
                   ("May", 30), ("Jun", -12), ("Jul", 18), ("Aug", 5)]:
        ws.append([m, pl])
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Profit & Loss"
    chart.y_axis.title = "Amount ($K)"
    data = Reference(ws, min_col=2, min_row=1, max_row=9)
    cats = Reference(ws, min_col=1, min_row=2, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic115_chart_negative_values.xlsx")


# ── 116. 100% stacked area chart ────────────────────────────────────────
def classic116_percent_stacked_area():
    wb = Workbook()
    ws = wb.active
    ws.title = "Energy"
    ws.append(["Year", "Coal", "Gas", "Nuclear", "Renewable"])
    for yr, coal, gas, nuc, ren in [(2015, 40, 25, 20, 15),
                                     (2017, 35, 27, 20, 18),
                                     (2019, 30, 28, 19, 23),
                                     (2021, 25, 28, 18, 29),
                                     (2023, 20, 26, 17, 37),
                                     (2025, 15, 24, 16, 45)]:
        ws.append([yr, coal, gas, nuc, ren])
    chart = AreaChart()
    chart.grouping = "percentStacked"
    chart.title = "Energy Mix Transition"
    data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=7)
    cats = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws.add_chart(chart, "A10")
    save(wb, "classic116_percent_stacked_area.xlsx")


# ── 117. Stock-like OHLC bar chart ──────────────────────────────────────
def classic117_stock_ohlc_chart():
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(["Day", "Open", "High", "Low", "Close"])
    import random
    random.seed(117)
    price = 150.0
    for d in range(1, 11):
        o = round(price + random.uniform(-2, 2), 2)
        h = round(o + random.uniform(0, 5), 2)
        l = round(o - random.uniform(0, 5), 2)
        c = round(random.uniform(l, h), 2)
        price = c
        ws.append([f"Day {d}", o, h, l, c])
    # Use a bar chart to visualize Open/Close range
    chart = BarChart()
    chart.type = "col"
    chart.title = "Stock OHLC (10 Days)"
    chart.y_axis.title = "Price ($)"
    data = Reference(ws, min_col=2, max_col=5, min_row=1, max_row=11)
    cats = Reference(ws, min_col=1, min_row=2, max_row=11)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "G2")
    save(wb, "classic117_stock_ohlc_chart.xlsx")


# ── 118. Bar chart with custom colors ───────────────────────────────────
def classic118_bar_chart_custom_colors():
    wb = Workbook()
    ws = wb.active
    ws.title = "Survey"
    ws.append(["Rating", "Count"])
    for rating, count in [("Excellent", 45), ("Good", 30), ("Average", 15),
                           ("Poor", 7), ("Very Poor", 3)]:
        ws.append([rating, count])
    chart = BarChart()
    chart.type = "col"
    chart.title = "Customer Satisfaction Survey"
    data = Reference(ws, min_col=2, min_row=1, max_row=6)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    colors = ["228B22", "32CD32", "FFD700", "FF8C00", "DC143C"]
    series = chart.series[0]
    for i, color in enumerate(colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)
    chart.width = 15
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic118_bar_chart_custom_colors.xlsx")


# ── 119. Dashboard with multiple chart types ────────────────────────────
def classic119_dashboard_multi_charts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    bold = Font(bold=True)
    # KPI row
    ws["A1"] = "KPI Dashboard - Q4 2025"
    ws["A1"].font = Font(bold=True, size=14)
    # Revenue data (rows 3-8)
    ws["A3"] = "Month"
    ws["B3"] = "Revenue"
    ws["C3"] = "Expenses"
    for cell in ws[3]:
        if cell.value:
            cell.font = bold
    for m, r, e in [("Oct", 85, 60), ("Nov", 92, 65), ("Dec", 110, 70)]:
        ws.append([m, r, e])
    # Segment data (rows 10-14)
    ws["A10"] = "Segment"
    ws["B10"] = "Share"
    ws["A10"].font = bold
    ws["B10"].font = bold
    for seg, sh in [("Enterprise", 45), ("SMB", 30), ("Consumer", 25)]:
        ws.append([seg, sh])
    # Bar chart
    bar = BarChart()
    bar.type = "col"
    bar.title = "Revenue vs Expenses"
    bar_data = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=6)
    bar_cats = Reference(ws, min_col=1, min_row=4, max_row=6)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    bar.width = 12
    bar.height = 8
    ws.add_chart(bar, "E2")
    # Pie chart
    pie = PieChart()
    pie.title = "Revenue by Segment"
    pie_data = Reference(ws, min_col=2, min_row=10, max_row=13)
    pie_cats = Reference(ws, min_col=1, min_row=11, max_row=13)
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)
    pie.width = 10
    pie.height = 8
    ws.add_chart(pie, "E16")
    save(wb, "classic119_dashboard_multi_charts.xlsx")


# ── 120. Chart with date axis ───────────────────────────────────────────
def classic120_chart_with_date_axis():
    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline"
    ws.append(["Date", "Downloads"])
    from datetime import date, timedelta
    start = date(2025, 1, 1)
    import random
    random.seed(120)
    downloads = 500
    for i in range(12):
        d = start + timedelta(days=i * 30)
        downloads += random.randint(-50, 100)
        ws.append([d.strftime("%Y-%m-%d"), downloads])
    chart = LineChart()
    chart.title = "Monthly Downloads (2025)"
    chart.y_axis.title = "Downloads"
    chart.x_axis.title = "Date"
    data = Reference(ws, min_col=2, min_row=1, max_row=13)
    cats = Reference(ws, min_col=1, min_row=2, max_row=13)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 18
    chart.height = 10
    ws.add_chart(chart, "D2")
    save(wb, "classic120_chart_with_date_axis.xlsx")


# ── 121-150: Style, Border & Background cases ───────────────────────────

# ── 121. Thin borders on all cells ───────────────────────────────────────
def classic121_thin_borders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Borders"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Item", "Qty", "Price", "Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.border = border
    data = [
        ("Widget A", 10, 5.00, 50.00),
        ("Widget B", 25, 3.50, 87.50),
        ("Widget C", 8, 12.00, 96.00),
        ("Widget D", 15, 7.25, 108.75),
        ("Widget E", 30, 2.00, 60.00),
    ]
    for r, row_data in enumerate(data, start=2):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
    save(wb, "classic121_thin_borders.xlsx")


# ── 122. Thick outer border with thin inner grid ────────────────────────
def classic122_thick_outer_thin_inner():
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed Borders"
    thick = Side(style="thick", color="000000")
    thin = Side(style="thin", color="000000")
    rows, cols = 6, 4
    headers = ["Region", "Q1", "Q2", "Q3"]
    regions = ["North", "South", "East", "West", "Central"]
    import random
    random.seed(122)
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            if r == 1:
                cell = ws.cell(row=r, column=c, value=headers[c - 1])
                cell.font = Font(bold=True)
            else:
                if c == 1:
                    cell = ws.cell(row=r, column=c, value=regions[r - 2])
                else:
                    cell = ws.cell(row=r, column=c, value=random.randint(1000, 9999))
            left = thick if c == 1 else thin
            right = thick if c == cols else thin
            top = thick if r == 1 else thin
            bottom = thick if r == rows else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
    save(wb, "classic122_thick_outer_thin_inner.xlsx")


# ── 123. Dashed borders ─────────────────────────────────────────────────
def classic123_dashed_borders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashed"
    styles = ["dashed", "dotted", "dashDot", "dashDotDot", "mediumDashed"]
    ws.cell(row=1, column=1, value="Border Style").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Sample").font = Font(bold=True)
    for i, style in enumerate(styles, start=2):
        side = Side(style=style, color="333333")
        b = Border(left=side, right=side, top=side, bottom=side)
        ws.cell(row=i, column=1, value=style)
        c = ws.cell(row=i, column=2, value="Bordered cell")
        c.border = b
    save(wb, "classic123_dashed_borders.xlsx")


# ── 124. Colored borders ────────────────────────────────────────────────
def classic124_colored_borders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Colored Borders"
    colors = [
        ("Red", "FF0000"), ("Green", "00AA00"), ("Blue", "0000FF"),
        ("Orange", "FF8800"), ("Purple", "880088"), ("Teal", "008888"),
    ]
    ws.cell(row=1, column=1, value="Color").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Cell").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Description").font = Font(bold=True)
    for i, (name, hex_color) in enumerate(colors, start=2):
        side = Side(style="medium", color=hex_color)
        b = Border(left=side, right=side, top=side, bottom=side)
        ws.cell(row=i, column=1, value=name)
        c = ws.cell(row=i, column=2, value="Sample")
        c.border = b
        ws.cell(row=i, column=3, value=f"Medium {name.lower()} border")
    save(wb, "classic124_colored_borders.xlsx")


# ── 125. Solid background fills ─────────────────────────────────────────
def classic125_solid_fills():
    wb = Workbook()
    ws = wb.active
    ws.title = "Solid Fills"
    fills = [
        ("Light Blue", "DAEEF3"), ("Light Green", "EBF1DE"),
        ("Light Yellow", "FFFFCC"), ("Light Red", "F2DCDB"),
        ("Light Purple", "E4DFEC"), ("Light Orange", "FDE9D9"),
        ("Gray 25%", "D9D9D9"), ("Sky Blue", "B7DEE8"),
    ]
    ws.cell(row=1, column=1, value="Fill Name").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Filled Cell").font = Font(bold=True)
    for i, (name, color) in enumerate(fills, start=2):
        ws.cell(row=i, column=1, value=name)
        c = ws.cell(row=i, column=2, value="Background")
        c.fill = PatternFill(fill_type="solid", fgColor=color)
    save(wb, "classic125_solid_fills.xlsx")


# ── 126. Header with dark background and white text ─────────────────────
def classic126_dark_header():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dark Header"
    dark_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    white_bold = Font(bold=True, color="FFFFFF", size=11)
    headers = ["Employee", "Department", "Salary", "Start Date"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = dark_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="center")
    data = [
        ("Alice Smith", "Engineering", 95000, "2020-03-15"),
        ("Bob Jones", "Marketing", 72000, "2019-07-01"),
        ("Carol Lee", "Finance", 88000, "2021-01-10"),
        ("David Kim", "Engineering", 102000, "2018-11-20"),
        ("Eva Chen", "HR", 68000, "2022-05-03"),
    ]
    for row_data in data:
        ws.append(list(row_data))
    save(wb, "classic126_dark_header.xlsx")


# ── 127. Multiple font styles (bold, italic, underline, strikethrough) ──
def classic127_font_styles():
    wb = Workbook()
    ws = wb.active
    ws.title = "Font Styles"
    ws.cell(row=1, column=1, value="Style").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Example").font = Font(bold=True)
    styles = [
        ("Bold", Font(bold=True, size=11)),
        ("Italic", Font(italic=True, size=11)),
        ("Underline", Font(underline="single", size=11)),
        ("Strikethrough", Font(strike=True, size=11)),
        ("Bold Italic", Font(bold=True, italic=True, size=11)),
        ("Bold Underline", Font(bold=True, underline="single", size=11)),
        ("Double Underline", Font(underline="double", size=11)),
        ("Bold + Red", Font(bold=True, color="FF0000", size=11)),
    ]
    for i, (name, font) in enumerate(styles, start=2):
        ws.cell(row=i, column=1, value=name)
        c = ws.cell(row=i, column=2, value=f"Sample {name} text")
        c.font = font
    save(wb, "classic127_font_styles.xlsx")


# ── 128. Font sizes ─────────────────────────────────────────────────────
def classic128_font_sizes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Font Sizes"
    sizes = [8, 9, 10, 11, 12, 14, 16, 18, 20, 24]
    ws.cell(row=1, column=1, value="Size").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Sample").font = Font(bold=True)
    for i, sz in enumerate(sizes, start=2):
        ws.cell(row=i, column=1, value=sz)
        c = ws.cell(row=i, column=2, value=f"Font size {sz}")
        c.font = Font(size=sz)
    ws.row_dimensions[11].height = 30  # space for size 24
    save(wb, "classic128_font_sizes.xlsx")


# ── 129. Cell alignment combinations ────────────────────────────────────
def classic129_alignment_combos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Alignment"
    h_aligns = ["left", "center", "right"]
    v_aligns = ["top", "center", "bottom"]
    ws.cell(row=1, column=1, value="").font = Font(bold=True)
    for ci, ha in enumerate(h_aligns, start=2):
        ws.cell(row=1, column=ci, value=ha).font = Font(bold=True)
    for ri, va in enumerate(v_aligns, start=2):
        ws.cell(row=ri, column=1, value=va).font = Font(bold=True)
        ws.row_dimensions[ri].height = 40
        for ci, ha in enumerate(h_aligns, start=2):
            c = ws.cell(row=ri, column=ci, value=f"{ha}/{va}")
            c.alignment = Alignment(horizontal=ha, vertical=va)
    for col in range(2, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18
    save(wb, "classic129_alignment_combos.xlsx")


# ── 130. Text wrapping and indent ────────────────────────────────────────
def classic130_wrap_and_indent():
    wb = Workbook()
    ws = wb.active
    ws.title = "Wrap & Indent"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    ws.cell(row=1, column=1, value="Wrapped Text").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Indented Text").font = Font(bold=True)
    long_text = "This is a long text that should wrap within the cell when text wrapping is enabled."
    c1 = ws.cell(row=2, column=1, value=long_text)
    c1.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 50
    for indent_level in range(5):
        r = indent_level + 3
        c = ws.cell(row=r, column=2, value=f"Indent level {indent_level}")
        c.alignment = Alignment(indent=indent_level)
    save(wb, "classic130_wrap_and_indent.xlsx")


# ── 131. Number format patterns ──────────────────────────────────────────
def classic131_number_formats():
    wb = Workbook()
    ws = wb.active
    ws.title = "Number Formats"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.cell(row=1, column=1, value="Format").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Display").font = Font(bold=True)
    formats = [
        ("#,##0", 1234567),
        ("#,##0.00", 1234567.891),
        ("$#,##0.00", 9876.5),
        ("0.00%", 0.8523),
        ("0.00E+00", 123456789),
        ("0000", 42),
        ("#,##0;(#,##0)", -5000),
        ("yyyy-mm-dd", 45658),
        ("dd/mm/yyyy", 45658),
        ("hh:mm:ss", 0.75),
    ]
    for i, (fmt, val) in enumerate(formats, start=2):
        ws.cell(row=i, column=1, value=fmt)
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt
        ws.cell(row=i, column=3, value=val)
    save(wb, "classic131_number_formats.xlsx")


# ── 132. Striped table (banded rows with borders) ───────────────────────
def classic132_striped_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Striped"
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_light = PatternFill(fill_type="solid", fgColor="F2F2F2")
    fill_white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Product", "Category", "Price", "Rating"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    import random
    random.seed(132)
    categories = ["Electronics", "Books", "Clothing", "Food", "Sports"]
    for i in range(1, 11):
        r = i + 1
        ws.cell(row=r, column=1, value=f"Product {i}").border = border
        ws.cell(row=r, column=2, value=random.choice(categories)).border = border
        ws.cell(row=r, column=3, value=round(random.uniform(5, 500), 2)).border = border
        ws.cell(row=r, column=4, value=round(random.uniform(1, 5), 1)).border = border
        fill = fill_light if i % 2 == 0 else fill_white
        for col in range(1, 5):
            ws.cell(row=r, column=col).fill = fill
    save(wb, "classic132_striped_table.xlsx")


# ── 133. Color gradient rows (simulated with progressive fills) ──────────
def classic133_gradient_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Gradient"
    ws.cell(row=1, column=1, value="Step").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Color Intensity").font = Font(bold=True)
    for i in range(10):
        r = i + 2
        intensity = int(255 - (i * 25))
        hex_color = f"00{intensity:02X}00"  # green gradient
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=(i + 1) * 10)
        for col in range(1, 4):
            c = ws.cell(row=r, column=col) if col > 1 else ws.cell(row=r, column=col)
            c.fill = PatternFill(fill_type="solid", fgColor=hex_color)
            if intensity < 128:
                c.font = Font(color="FFFFFF")
        ws.cell(row=r, column=3, value=f"Green {hex_color}")
    save(wb, "classic133_gradient_rows.xlsx")


# ── 134. Conditional-style heatmap (manual fills based on value) ─────────
def classic134_heatmap():
    wb = Workbook()
    ws = wb.active
    ws.title = "Heatmap"
    import random
    random.seed(134)
    headers = [""] + [f"Col{c}" for c in range(1, 7)]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    for r in range(2, 9):
        ws.cell(row=r, column=1, value=f"Row{r - 1}").font = Font(bold=True)
        for c in range(2, 8):
            val = random.randint(0, 100)
            cell = ws.cell(row=r, column=c, value=val)
            # Map value to red intensity
            red = int(255 * val / 100)
            green = int(255 * (100 - val) / 100)
            hex_color = f"{red:02X}{green:02X}30"
            cell.fill = PatternFill(fill_type="solid", fgColor=hex_color)
            if val > 60:
                cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
    save(wb, "classic134_heatmap.xlsx")


# ── 135. Bottom border only (underline style) ───────────────────────────
def classic135_bottom_border_only():
    wb = Workbook()
    ws = wb.active
    ws.title = "Bottom Borders"
    bottom_thin = Border(bottom=Side(style="thin", color="000000"))
    bottom_medium = Border(bottom=Side(style="medium", color="000000"))
    bottom_thick = Border(bottom=Side(style="thick", color="000000"))
    bottom_double = Border(bottom=Side(style="double", color="000000"))
    ws.cell(row=1, column=1, value="Section Header").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).border = bottom_thick
    ws.cell(row=1, column=2).border = bottom_thick
    ws.cell(row=1, column=3).border = bottom_thick
    ws.cell(row=3, column=1, value="Thin bottom")
    ws.cell(row=3, column=1).border = bottom_thin
    ws.cell(row=4, column=1, value="Medium bottom")
    ws.cell(row=4, column=1).border = bottom_medium
    ws.cell(row=5, column=1, value="Thick bottom")
    ws.cell(row=5, column=1).border = bottom_thick
    ws.cell(row=6, column=1, value="Double bottom")
    ws.cell(row=6, column=1).border = bottom_double
    ws.column_dimensions["A"].width = 20
    save(wb, "classic135_bottom_border_only.xlsx")


# ── 136. Financial report with styled cells ──────────────────────────────
def classic136_financial_report_styled():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    bold = Font(bold=True)
    money_fmt = "$#,##0.00"
    header_fill = PatternFill(fill_type="solid", fgColor="002060")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_border = Border(top=Side(style="double", color="000000"),
                          bottom=Side(style="double", color="000000"))
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    # Header
    for col, h in enumerate(["Category", "2024", "2025"], start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    rows = [
        ("Revenue", 450000, 520000),
        ("Cost of Goods", -180000, -195000),
        ("Gross Profit", 270000, 325000),
        ("", None, None),
        ("Operating Expenses", -120000, -135000),
        ("R&D", -45000, -55000),
        ("Marketing", -30000, -38000),
        ("", None, None),
        ("Net Income", 75000, 97000),
    ]
    for i, (label, v1, v2) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).border = border
        if label in ("Gross Profit", "Net Income"):
            ws.cell(row=i, column=1).font = bold
        if v1 is not None:
            c1 = ws.cell(row=i, column=2, value=v1)
            c1.number_format = money_fmt
            c1.border = border
            c2 = ws.cell(row=i, column=3, value=v2)
            c2.number_format = money_fmt
            c2.border = border
        if label == "Net Income":
            for col in range(1, 4):
                ws.cell(row=i, column=col).border = total_border
                ws.cell(row=i, column=col).font = Font(bold=True, size=12)
    save(wb, "classic136_financial_report_styled.xlsx")


# ── 137. Checkerboard pattern ────────────────────────────────────────────
def classic137_checkerboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "Checkerboard"
    fill_dark = PatternFill(fill_type="solid", fgColor="333333")
    fill_light = PatternFill(fill_type="solid", fgColor="EEEEEE")
    for r in range(1, 9):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c, value=" ")
            if (r + c) % 2 == 0:
                cell.fill = fill_dark
                cell.font = Font(color="FFFFFF")
            else:
                cell.fill = fill_light
            ws.column_dimensions[get_column_letter(c)].width = 5
        ws.row_dimensions[r].height = 25
    save(wb, "classic137_checkerboard.xlsx")


# ── 138. Multi-color background grid ────────────────────────────────────
def classic138_color_grid():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Grid"
    colors = [
        "FF6B6B", "FFD93D", "6BCB77", "4D96FF",
        "FF8E71", "C780FF", "FFB4B4", "B5DEFF",
        "E8FFC1", "FFC0D0", "D5AAFF", "A0E7E5",
    ]
    for r in range(1, 5):
        for c in range(1, 4):
            idx = (r - 1) * 3 + (c - 1)
            color = colors[idx]
            cell = ws.cell(row=r, column=c, value=f"#{color}")
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.row_dimensions[r].height = 30
    save(wb, "classic138_color_grid.xlsx")


# ── 139. Pattern fills (hatches) ─────────────────────────────────────────
def classic139_pattern_fills():
    wb = Workbook()
    ws = wb.active
    ws.title = "Patterns"
    patterns = [
        ("solid", "4472C4"), ("darkGray", "808080"),
        ("mediumGray", "808080"), ("lightGray", "808080"),
        ("gray125", "808080"), ("gray0625", "808080"),
        ("darkHorizontal", "0000FF"), ("darkVertical", "0000FF"),
        ("darkDown", "FF0000"), ("darkUp", "FF0000"),
        ("lightHorizontal", "00AA00"), ("lightVertical", "00AA00"),
    ]
    ws.cell(row=1, column=1, value="Pattern Type").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Sample").font = Font(bold=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    for i, (pat, color) in enumerate(patterns, start=2):
        ws.cell(row=i, column=1, value=pat)
        c = ws.cell(row=i, column=2, value="Pattern")
        c.fill = PatternFill(fill_type=pat, fgColor=color, bgColor="FFFFFF")
    save(wb, "classic139_pattern_fills.xlsx")


# ── 140. Rotated text ───────────────────────────────────────────────────
def classic140_rotated_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "Rotated"
    # openpyxl: 0-90 = counter-clockwise, 91-180 maps to -(180-val) clockwise
    angles = [0, 15, 30, 45, 60, 90, 105, 120, 135, 150, 180]
    ws.cell(row=1, column=1, value="Rotation").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Text").font = Font(bold=True)
    for i, angle in enumerate(angles, start=2):
        ws.cell(row=i, column=1, value=f"{angle}")
        c = ws.cell(row=i, column=2, value="Rotated")
        c.alignment = Alignment(text_rotation=angle)
        ws.row_dimensions[i].height = 40
    ws.column_dimensions["B"].width = 18
    save(wb, "classic140_rotated_text.xlsx")


# ── 141. Mixed border styles per edge ───────────────────────────────────
def classic141_mixed_edge_borders():
    wb = Workbook()
    ws = wb.active
    ws.title = "Edge Borders"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25
    combos = [
        ("thick left, thin others",
         Border(left=Side("thick", "FF0000"), right=Side("thin"),
                top=Side("thin"), bottom=Side("thin"))),
        ("double bottom",
         Border(bottom=Side("double", "0000FF"))),
        ("medium top + bottom",
         Border(top=Side("medium", "00AA00"), bottom=Side("medium", "00AA00"))),
        ("dashed all sides",
         Border(left=Side("dashed"), right=Side("dashed"),
                top=Side("dashed"), bottom=Side("dashed"))),
        ("thick all, red",
         Border(left=Side("thick", "FF0000"), right=Side("thick", "FF0000"),
                top=Side("thick", "FF0000"), bottom=Side("thick", "FF0000"))),
    ]
    ws.cell(row=1, column=1, value="Description").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Cell").font = Font(bold=True)
    for i, (desc, b) in enumerate(combos, start=2):
        ws.cell(row=i, column=1, value=desc)
        c = ws.cell(row=i, column=2, value="Styled")
        c.border = b
        ws.row_dimensions[i].height = 30
    save(wb, "classic141_mixed_edge_borders.xlsx")


# ── 142. Invoice with full styling ──────────────────────────────────────
def classic142_styled_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "INVOICE"
    title_cell.font = Font(bold=True, size=20, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="D6E4F0")
    # Info section
    ws["A3"] = "Invoice #:"
    ws["B3"] = "INV-2025-0099"
    ws["A4"] = "Date:"
    ws["B4"] = "2025-06-15"
    ws["A3"].font = Font(bold=True)
    ws["A4"].font = Font(bold=True)
    # Table header
    thin = Side(style="thin", color="1F4E79")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    h_row = 6
    for col, h in enumerate(["Item", "Description", "Qty", "Price", "Total"], start=1):
        c = ws.cell(row=h_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    items = [
        ("SVC-001", "Web Development", 40, 125.00, 5000.00),
        ("SVC-002", "UI/UX Design", 20, 100.00, 2000.00),
        ("SVC-003", "Testing & QA", 15, 90.00, 1350.00),
        ("LIC-001", "Annual License", 1, 2400.00, 2400.00),
    ]
    alt_fill = PatternFill(fill_type="solid", fgColor="EDF2F9")
    for i, item_data in enumerate(items):
        r = h_row + 1 + i
        for col, val in enumerate(item_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if i % 2 == 1:
                c.fill = alt_fill
            if col >= 4:
                c.number_format = "$#,##0.00"
    # Totals
    total_r = h_row + len(items) + 2
    ws.cell(row=total_r, column=4, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=total_r, column=5, value=10750.00).number_format = "$#,##0.00"
    ws.cell(row=total_r + 1, column=4, value="Tax (10%):").font = Font(bold=True)
    ws.cell(row=total_r + 1, column=5, value=1075.00).number_format = "$#,##0.00"
    total_border = Border(top=Side("double", "1F4E79"), bottom=Side("double", "1F4E79"))
    ws.cell(row=total_r + 2, column=4, value="Total:").font = Font(bold=True, size=13)
    tc = ws.cell(row=total_r + 2, column=5, value=11825.00)
    tc.number_format = "$#,##0.00"
    tc.font = Font(bold=True, size=13)
    tc.border = total_border
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    save(wb, "classic142_styled_invoice.xlsx")


# ── 143. Colored tabs (multiple sheets with tab colors) ──────────────────
def classic143_colored_tabs():
    wb = Workbook()
    tab_data = [
        ("Sales", "FF0000", [("Product A", 100), ("Product B", 200)]),
        ("Inventory", "00AA00", [("Widget", 500), ("Gadget", 300)]),
        ("Finance", "0000FF", [("Revenue", 50000), ("Cost", 30000)]),
        ("HR", "FF8800", [("Employees", 50), ("Open Roles", 5)]),
    ]
    first = True
    for name, color, rows in tab_data:
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = color
        ws.append(["Label", "Value"])
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        for label, val in rows:
            ws.append([label, val])
    save(wb, "classic143_colored_tabs.xlsx")


# ── 144. Cell comments / notes style (simulated with styled cells) ───────
def classic144_note_style_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    note_fill = PatternFill(fill_type="solid", fgColor="FFFFD5")
    note_border = Border(
        left=Side("thin", "CCCC00"), right=Side("thin", "CCCC00"),
        top=Side("thin", "CCCC00"), bottom=Side("thin", "CCCC00"))
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 40
    ws.cell(row=1, column=1, value="Field").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Note").font = Font(bold=True)
    notes = [
        ("Name", "Must be full legal name"),
        ("Email", "Use company email only"),
        ("Phone", "Include country code"),
        ("Address", "PO boxes not accepted"),
        ("DOB", "Format: YYYY-MM-DD"),
    ]
    for i, (field, note) in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=field)
        c = ws.cell(row=i, column=2, value=note)
        c.fill = note_fill
        c.border = note_border
        c.font = Font(italic=True, color="666600")
    save(wb, "classic144_note_style_cells.xlsx")


# ── 145. Status badge style (colored fills for status labels) ────────────
def classic145_status_badges():
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Status"
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_styles = {
        "Completed": (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100", bold=True)),
        "In Progress": (PatternFill("solid", fgColor="FFEB9C"), Font(color="9C5700", bold=True)),
        "Blocked": (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006", bold=True)),
        "Not Started": (PatternFill("solid", fgColor="D9D9D9"), Font(color="333333")),
        "In Review": (PatternFill("solid", fgColor="B4C6E7"), Font(color="1F3864", bold=True)),
    }
    headers = ["Task", "Owner", "Status", "Due Date"]
    header_fill = PatternFill("solid", fgColor="404040")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.border = border
    tasks = [
        ("Backend API", "Alice", "Completed", "2025-02-01"),
        ("Frontend UI", "Bob", "In Progress", "2025-03-15"),
        ("Database Migration", "Carol", "Blocked", "2025-02-20"),
        ("Documentation", "David", "Not Started", "2025-04-01"),
        ("Code Review", "Eva", "In Review", "2025-03-10"),
        ("Deployment", "Frank", "Not Started", "2025-04-15"),
        ("Testing", "Grace", "In Progress", "2025-03-20"),
    ]
    for i, (task, owner, status, due) in enumerate(tasks, start=2):
        ws.cell(row=i, column=1, value=task).border = border
        ws.cell(row=i, column=2, value=owner).border = border
        sc = ws.cell(row=i, column=3, value=status)
        sc.border = border
        sc.alignment = Alignment(horizontal="center")
        if status in status_styles:
            sc.fill, sc.font = status_styles[status]
        ws.cell(row=i, column=4, value=due).border = border
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    save(wb, "classic145_status_badges.xlsx")


# ── 146. Double border table ────────────────────────────────────────────
def classic146_double_border_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Double Border"
    double = Side(style="double", color="000000")
    thin = Side(style="thin", color="000000")
    headers = ["Name", "Role", "Years", "Rating"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.border = Border(left=double, right=double, top=double, bottom=double)
        c.fill = PatternFill("solid", fgColor="E2EFDA")
    data = [
        ("Alice", "Engineer", 5, "Excellent"),
        ("Bob", "Designer", 3, "Good"),
        ("Carol", "Manager", 8, "Excellent"),
        ("David", "Analyst", 2, "Satisfactory"),
    ]
    for r, row_data in enumerate(data, start=2):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Bottom edge double
    last_r = len(data) + 1
    for col in range(1, 5):
        ws.cell(row=last_r, column=col).border = Border(
            left=thin, right=thin, top=thin, bottom=double)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    save(wb, "classic146_double_border_table.xlsx")


# ── 147. Multi-sheet styled report ──────────────────────────────────────
def classic147_multi_sheet_styled():
    wb = Workbook()
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheets_data = {
        "Summary": {
            "header_color": "2F5496",
            "headers": ["Metric", "Value"],
            "data": [("Total Revenue", "$1,200,000"), ("Total Costs", "$780,000"),
                     ("Net Profit", "$420,000"), ("Margin", "35%")],
        },
        "Quarterly": {
            "header_color": "548235",
            "headers": ["Quarter", "Revenue", "Costs", "Profit"],
            "data": [("Q1", 280000, 190000, 90000), ("Q2", 310000, 200000, 110000),
                     ("Q3", 290000, 185000, 105000), ("Q4", 320000, 205000, 115000)],
        },
        "Employees": {
            "header_color": "BF8F00",
            "headers": ["Name", "Department", "Salary"],
            "data": [("Alice", "Eng", 110000), ("Bob", "Sales", 85000),
                     ("Carol", "Eng", 105000), ("David", "HR", 75000)],
        },
    }
    first = True
    for sheet_name, info in sheets_data.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        h_fill = PatternFill("solid", fgColor=info["header_color"])
        for col, h in enumerate(info["headers"], start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = h_fill
            c.border = border
        for r, row_data in enumerate(info["data"], start=2):
            for col, val in enumerate(row_data, start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = border
        for col in range(1, len(info["headers"]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
    save(wb, "classic147_multi_sheet_styled.xlsx")


# ── 148. Frozen header with styled grid ──────────────────────────────────
def classic148_frozen_styled_grid():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Grid"
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["ID", "Name", "Category", "Value", "Status", "Date"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")
    import random
    random.seed(148)
    categories = ["Alpha", "Beta", "Gamma"]
    statuses = ["Active", "Inactive", "Pending"]
    fill_even = PatternFill("solid", fgColor="D9E2F3")
    for i in range(1, 21):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=f"Item-{i:03d}").border = border
        ws.cell(row=r, column=3, value=random.choice(categories)).border = border
        ws.cell(row=r, column=4, value=round(random.uniform(10, 999), 2)).border = border
        ws.cell(row=r, column=5, value=random.choice(statuses)).border = border
        ws.cell(row=r, column=6, value=f"2025-{random.randint(1,12):02d}-{random.randint(1,28):02d}").border = border
        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = fill_even
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14
    save(wb, "classic148_frozen_styled_grid.xlsx")


# ── 149. Merged header with styled sub-sections ─────────────────────────
def classic149_merged_styled_sections():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    # Title
    ws.merge_cells("A1:D1")
    tc = ws["A1"]
    tc.value = "Quarterly Performance Report"
    tc.font = Font(bold=True, size=16, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="1F4E79")
    tc.alignment = Alignment(horizontal="center")
    # Section 1
    ws.merge_cells("A3:D3")
    s1 = ws["A3"]
    s1.value = "Revenue Breakdown"
    s1.font = Font(bold=True, size=12, color="1F4E79")
    s1.fill = PatternFill("solid", fgColor="D6E4F0")
    for col, h in enumerate(["Source", "Q1", "Q2", "Total"], start=1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True)
        c.border = border
    rev_data = [("Online", 120000, 140000, 260000),
                ("Retail", 90000, 85000, 175000),
                ("Wholesale", 60000, 70000, 130000)]
    for r, row_data in enumerate(rev_data, start=5):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if col >= 2:
                c.number_format = "#,##0"
    # Section 2
    ws.merge_cells("A9:D9")
    s2 = ws["A9"]
    s2.value = "Expense Summary"
    s2.font = Font(bold=True, size=12, color="FFFFFF")
    s2.fill = PatternFill("solid", fgColor="C00000")
    for col, h in enumerate(["Category", "Q1", "Q2", "Total"], start=1):
        c = ws.cell(row=10, column=col, value=h)
        c.font = Font(bold=True)
        c.border = border
    exp_data = [("Salaries", 200000, 210000, 410000),
                ("Marketing", 30000, 35000, 65000),
                ("Operations", 50000, 48000, 98000)]
    for r, row_data in enumerate(exp_data, start=11):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            if col >= 2:
                c.number_format = "#,##0"
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 16
    save(wb, "classic149_merged_styled_sections.xlsx")


# ── 150. Kitchen sink — all style features combined ─────────────────────
def classic150_kitchen_sink_styles():
    wb = Workbook()
    ws = wb.active
    ws.title = "All Styles"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    # Row 1: merged bold title
    ws.merge_cells("A1:C1")
    tc = ws["A1"]
    tc.value = "Style Showcase"
    tc.font = Font(bold=True, size=18, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="2F5496")
    tc.alignment = Alignment(horizontal="center")
    # Row 2: sub-header with bottom border
    for col, h in enumerate(["Feature", "Example", "Notes"], start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, size=12)
        c.border = Border(bottom=Side("medium", "2F5496"))
    # Rows 3+: various style demos
    r = 3
    # Bold + italic + red
    ws.cell(row=r, column=1, value="Bold Italic Red")
    c = ws.cell(row=r, column=2, value="Styled Text")
    c.font = Font(bold=True, italic=True, color="FF0000", size=12)
    ws.cell(row=r, column=3, value="Font combo")
    r += 1
    # Underline double
    ws.cell(row=r, column=1, value="Double Underline")
    c = ws.cell(row=r, column=2, value="Important Value")
    c.font = Font(underline="double", size=11)
    ws.cell(row=r, column=3, value="Emphasis")
    r += 1
    # Strikethrough
    ws.cell(row=r, column=1, value="Strikethrough")
    c = ws.cell(row=r, column=2, value="Deprecated Item")
    c.font = Font(strike=True, color="999999")
    ws.cell(row=r, column=3, value="Removed")
    r += 1
    # Solid fill + white text
    ws.cell(row=r, column=1, value="Dark Fill")
    c = ws.cell(row=r, column=2, value="White on Dark")
    c.fill = PatternFill("solid", fgColor="333333")
    c.font = Font(color="FFFFFF", bold=True)
    ws.cell(row=r, column=3, value="Inverted")
    r += 1
    # Colored border
    ws.cell(row=r, column=1, value="Red Border")
    c = ws.cell(row=r, column=2, value="Attention!")
    red_side = Side("thick", "FF0000")
    c.border = Border(left=red_side, right=red_side, top=red_side, bottom=red_side)
    ws.cell(row=r, column=3, value="Alert style")
    r += 1
    # Number format
    ws.cell(row=r, column=1, value="Currency Format")
    c = ws.cell(row=r, column=2, value=12345.678)
    c.number_format = "$#,##0.00"
    ws.cell(row=r, column=3, value="Formatted number")
    r += 1
    # Percentage
    ws.cell(row=r, column=1, value="Percentage")
    c = ws.cell(row=r, column=2, value=0.8756)
    c.number_format = "0.00%"
    ws.cell(row=r, column=3, value="Percent display")
    r += 1
    # Wrapped + centered
    ws.cell(row=r, column=1, value="Wrap + Center")
    c = ws.cell(row=r, column=2, value="This text wraps in the cell nicely")
    c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 40
    ws.cell(row=r, column=3, value="Multi-line")
    r += 1
    # Pattern fill
    ws.cell(row=r, column=1, value="Pattern Fill")
    c = ws.cell(row=r, column=2, value="Gray pattern")
    c.fill = PatternFill(fill_type="lightGray", fgColor="4472C4", bgColor="FFFFFF")
    ws.cell(row=r, column=3, value="Hatched")
    r += 1
    # Large font
    ws.cell(row=r, column=1, value="Large Font")
    c = ws.cell(row=r, column=2, value="BIG")
    c.font = Font(size=24, bold=True, color="2F5496")
    ws.row_dimensions[r].height = 35
    ws.cell(row=r, column=3, value="Size 24")
    save(wb, "classic150_kitchen_sink_styles.xlsx")


# ── 151. Multilingual greetings — broad script coverage ─────────────────
def classic151_multilingual_greetings():
    wb = Workbook()
    ws = wb.active
    ws.title = "Greetings"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.append(["Language", "Hello", "Thank you"])
    ws.append(["English", "Hello", "Thank you"])
    ws.append(["Chinese", "你好", "谢谢"])
    ws.append(["Japanese", "こんにちは", "ありがとう"])
    ws.append(["Korean", "안녕하세요", "감사합니다"])
    ws.append(["Thai", "สวัสดี", "ขอบคุณ"])
    ws.append(["Hindi", "नमस्ते", "धन्यवाद"])
    ws.append(["Arabic", "مرحبا", "شكرا"])
    ws.append(["Hebrew", "שלום", "תודה"])
    ws.append(["Greek", "Γεια σου", "Ευχαριστώ"])
    ws.append(["Russian", "Привет", "Спасибо"])
    ws.append(["Vietnamese", "Xin chào", "Cảm ơn"])
    ws.append(["Turkish", "Merhaba", "Teşekkürler"])
    save(wb, "classic151_multilingual_greetings.xlsx")


# ── 152. Emoji sampler — common categories ──────────────────────────────
def classic152_emoji_sampler():
    wb = Workbook()
    ws = wb.active
    ws.title = "Emoji"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.append(["Category", "Emoji"])
    ws.append(["Faces", "😀😃😄😁😆"])
    ws.append(["Hearts", "❤️🧡💛💚💙"])
    ws.append(["Animals", "🐶🐱🐭🐹🐰"])
    ws.append(["Food", "🍎🍐🍊🍋🍌"])
    ws.append(["Travel", "✈️🚗🚌🚂🚀"])
    ws.append(["Sports", "⚽🏀🏈⚾🎾"])
    ws.append(["Symbols", "✅❌⚠️🔴🟢"])
    ws.append(["Hands", "👍👎👏🤝✌️"])
    save(wb, "classic152_emoji_sampler.xlsx")


# ── 153. Currency symbols around the world ──────────────────────────────
def classic153_currency_symbols():
    wb = Workbook()
    ws = wb.active
    ws.title = "Currencies"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.append(["Currency", "Symbol", "Example"])
    data = [
        ("US Dollar", "$", "$1,234.56"),
        ("Euro", "€", "€1.234,56"),
        ("British Pound", "£", "£1,234.56"),
        ("Japanese Yen", "¥", "¥123,456"),
        ("Chinese Yuan", "¥", "¥1,234.56"),
        ("Korean Won", "₩", "₩1,234,560"),
        ("Indian Rupee", "₹", "₹1,23,456"),
        ("Thai Baht", "฿", "฿1,234.56"),
        ("Russian Ruble", "₽", "₽1 234,56"),
        ("Turkish Lira", "₺", "₺1.234,56"),
        ("Bitcoin", "₿", "₿0.05"),
        ("Swiss Franc", "CHF", "CHF 1'234.56"),
    ]
    for row in data:
        ws.append(list(row))
    save(wb, "classic153_currency_symbols.xlsx")


# ── 154. Mathematical symbols and formulas ──────────────────────────────
def classic154_math_symbols():
    wb = Workbook()
    ws = wb.active
    ws.title = "Math"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.append(["Category", "Symbols"])
    ws.append(["Operators", "± × ÷ ≠ ≤ ≥ ≈ ∝ ∓"])
    ws.append(["Greek Letters", "α β γ δ ε ζ η θ ι κ λ μ"])
    ws.append(["Set Theory", "∈ ∉ ⊂ ⊃ ∪ ∩ ∅ ℵ"])
    ws.append(["Logic", "∀ ∃ ¬ ∧ ∨ ⊕ ⟹ ⟺"])
    ws.append(["Calculus", "∫ ∬ ∮ ∂ ∇ ∑ ∏ √"])
    ws.append(["Arrows", "→ ← ↑ ↓ ↔ ⇒ ⇐ ⇔"])
    ws.append(["Misc", "∞ ℏ ℝ ℤ ℚ ℕ ℂ"])
    ws.append(["Superscripts", "x² y³ aⁿ eⁱ"])
    ws.append(["Subscripts", "H₂O CO₂ xₙ aᵢ"])
    save(wb, "classic154_math_symbols.xlsx")


# ── 155. Diacritical marks — Latin script with accents ──────────────────
def classic155_diacritical_marks():
    wb = Workbook()
    ws = wb.active
    ws.title = "Diacritics"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    ws.append(["Type", "Examples"])
    ws.append(["Acute", "á é í ó ú ý ś ź ć ń"])
    ws.append(["Grave", "à è ì ò ù ỳ"])
    ws.append(["Circumflex", "â ê î ô û ŵ ŷ ĉ"])
    ws.append(["Umlaut", "ä ë ï ö ü ÿ"])
    ws.append(["Tilde", "ã ñ õ ũ ĩ"])
    ws.append(["Cedilla", "ç ş ţ ḑ ȩ"])
    ws.append(["Ring", "å ů"])
    ws.append(["Caron", "č š ž ř ň ě ď ť"])
    ws.append(["Stroke", "ø đ ħ ł ŧ"])
    ws.append(["Ligatures", "æ œ ß ĳ"])
    save(wb, "classic155_diacritical_marks.xlsx")


# ── 156. RTL and BiDi text — Arabic and Hebrew paragraphs ───────────────
def classic156_rtl_bidi_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "RTL"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.append(["Script", "Text"])
    ws.append(["Arabic", "مرحبا بالعالم"])
    ws.append(["Hebrew", "שלום עולם"])
    ws.append(["Persian", "سلام دنیا"])
    ws.append(["Urdu", "ہیلو دنیا"])
    save(wb, "classic156_rtl_bidi_text.xlsx")


# ── 157. CJK extended — Traditional Chinese, rare kanji ─────────────────
def classic157_cjk_extended():
    wb = Workbook()
    ws = wb.active
    ws.title = "CJK"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.append(["Variant", "Text", "Notes"])
    ws.append(["Simplified CN", "简体中文测试字符串", "Mainland China"])
    ws.append(["Traditional CN", "繁體中文測試字串", "Taiwan / HK"])
    ws.append(["Japanese mixed", "漢字とひらがなとカタカナ", "Kanji + Hiragana + Katakana"])
    ws.append(["Korean mixed", "한글과 漢字 혼용 텍스트", "Hangul + Hanja"])
    ws.append(["Rare CJK", "𠀀𠀁𠀂𠀃𠀄", "CJK Ext-B (SMP)"])
    ws.append(["Full-width", "ＡＢＣＤ１２３４", "Full-width alphanumeric"])
    ws.append(["Half-width kana", "ｱｲｳｴｵ ｶｷｸｹｺ", "Half-width katakana"])
    save(wb, "classic157_cjk_extended.xlsx")


# ── 158. Emoji skin tone variants ───────────────────────────────────────
def classic158_emoji_skin_tones():
    wb = Workbook()
    ws = wb.active
    ws.title = "SkinTones"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.append(["Gesture", "Skin Tones (light → dark)"])
    ws.append(["Thumbs up", "👍🏻👍🏼👍🏽👍🏾👍🏿"])
    ws.append(["Waving", "👋🏻👋🏼👋🏽👋🏾👋🏿"])
    ws.append(["Clapping", "👏🏻👏🏼👏🏽👏🏾👏🏿"])
    ws.append(["Raised fist", "✊🏻✊🏼✊🏽✊🏾✊🏿"])
    ws.append(["Person", "🧑🏻🧑🏼🧑🏽🧑🏾🧑🏿"])
    save(wb, "classic158_emoji_skin_tones.xlsx")


# ── 159. ZWJ emoji sequences — compound emoji ──────────────────────────
def classic159_zwj_emoji():
    wb = Workbook()
    ws = wb.active
    ws.title = "ZWJ"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.append(["Description", "Emoji"])
    ws.append(["Family", "👨‍👩‍👧‍👦"])
    ws.append(["Couple with heart", "👩‍❤️‍👨"])
    ws.append(["Woman technologist", "👩‍💻"])
    ws.append(["Man cook", "👨‍🍳"])
    ws.append(["Rainbow flag", "🏳️‍🌈"])
    ws.append(["Trans flag", "🏳️‍⚧️"])
    ws.append(["Firefighter", "🧑‍🚒"])
    ws.append(["Health worker", "🧑‍⚕️"])
    ws.append(["Service dog", "🐕‍🦺"])
    save(wb, "classic159_zwj_emoji.xlsx")


# ── 160. Punctuation marks from various scripts ─────────────────────────
def classic160_punctuation_marks():
    wb = Workbook()
    ws = wb.active
    ws.title = "Punctuation"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.append(["Type", "Characters"])
    ws.append(["Latin", ". , ; : ! ? … — – ' ' " " « »"])
    ws.append(["CJK", "。、；：！？「」『』【】（）"])
    ws.append(["Arabic", "، ؛ ؟ ٪ ﷽"])
    ws.append(["Devanagari", "। ॥ ꣸ ꣹ ꣺"])
    ws.append(["Thai", "ฯ ๆ ๏ ๚ ๛"])
    ws.append(["Misc brackets", "⟨⟩ ⟪⟫ ⌈⌉ ⌊⌋ ‖"])
    ws.append(["Typographic", "† ‡ § ¶ © ® ™ ℠"])
    save(wb, "classic160_punctuation_marks.xlsx")


# ── 161. Box drawing and block characters ───────────────────────────────
def classic161_box_drawing():
    wb = Workbook()
    ws = wb.active
    ws.title = "BoxDraw"
    mono = Font(name="Consolas", size=11)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.append(["Type", "Characters"])
    rows = [
        ("Light box", "┌──┬──┐│  │  │├──┼──┤└──┴──┘"),
        ("Heavy box", "┏━━┳━━┓┃  ┃  ┃┣━━╋━━┫┗━━┻━━┛"),
        ("Double box", "╔══╦══╗║  ║  ║╠══╬══╣╚══╩══╝"),
        ("Blocks", "▀▁▂▃▄▅▆▇█ ░▒▓"),
        ("Geometric", "■□▪▫▲△▼▽◆◇○●◎"),
        ("Braille", "⠁⠂⠃⠄⠅⠆⠇⠈⠉⠊"),
    ]
    for label, chars in rows:
        r = ws.append([label, chars])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.font = mono
    save(wb, "classic161_box_drawing.xlsx")


# ── 162. Mixed CJK + emoji styled table ─────────────────────────────────
def classic162_cjk_emoji_styled():
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D35400")
    headers = ["Icon", "Dish", "Price", "Rating"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    data = [
        ("🍣", "寿司", "¥1,200", "★★★★★"),
        ("🍜", "ラーメン", "¥850", "★★★★"),
        ("🍱", "弁当", "¥1,500", "★★★★★"),
        ("🍙", "おにぎり", "¥400", "★★★"),
    ]
    for r, row_data in enumerate(data, start=2):
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=col, value=val)
    save(wb, "classic162_cjk_emoji_styled.xlsx")


# ── 163. Cyrillic alphabets — Russian, Ukrainian, Serbian ───────────────
def classic163_cyrillic_alphabets():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cyrillic"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 45
    ws.append(["Language", "Sample Text"])
    ws.append(["Russian", "Съешь ещё этих мягких французских булок, да выпей чаю."])
    ws.append(["Ukrainian", "Жебракують філософи при ґанку церкви в Гадячі."])
    ws.append(["Serbian", "Ђурађ Бранковић је био владар Србије."])
    ws.append(["Bulgarian", "Щъркел яде бялата жаба."])
    ws.append(["Mongolian", "Би монгол хэл дээр бичиж байна."])
    save(wb, "classic163_cyrillic_alphabets.xlsx")


# ── 164. Indic scripts — Devanagari, Tamil, Bengali, Telugu ─────────────
def classic164_indic_scripts():
    wb = Workbook()
    ws = wb.active
    ws.title = "Indic"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.append(["Script", "Sample"])
    ws.append(["Devanagari", "नमस्ते"])
    ws.append(["Tamil", "வணக்கம்"])
    ws.append(["Bengali", "নমস্কার"])
    ws.append(["Telugu", "నమస్కారం"])
    ws.append(["Gujarati", "નમસ્તે"])
    save(wb, "classic164_indic_scripts.xlsx")


# ── 165. Southeast Asian scripts — Thai, Lao, Myanmar, Khmer ────────────
def classic165_southeast_asian():
    wb = Workbook()
    ws = wb.active
    ws.title = "SEAsia"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.append(["Script", "Sample"])
    ws.append(["Thai", "ภาษาไทยเป็นภาษาที่มีวรรณยุกต์"])
    ws.append(["Lao", "ພາສາລາວເປັນພາສາທີ່ສວຍງາມ"])
    ws.append(["Myanmar", "မြန်မာဘာသာစကားသည် လှပသည်"])
    ws.append(["Khmer", "ភាសាខ្មែរជាភាសាចំណាស់"])
    ws.append(["Tibetan", "བོད་ཀྱི་སྐད་ཡིག་ནི་གལ་ཆེན་པོ་ཡིན།"])
    save(wb, "classic165_southeast_asian.xlsx")


# ── 166. Emoji progress bars and status indicators ──────────────────────
def classic166_emoji_progress():
    wb = Workbook()
    ws = wb.active
    ws.title = "Progress"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.append(["Task", "Status", "Progress"])
    ws.append(["Design", "✅", "🟩🟩🟩🟩🟩🟩🟩🟩🟩🟩 100%"])
    ws.append(["Frontend", "🔄", "🟩🟩🟩🟩🟩🟩🟩⬜⬜⬜ 70%"])
    ws.append(["Backend", "🔄", "🟩🟩🟩🟩🟩⬜⬜⬜⬜⬜ 50%"])
    ws.append(["Testing", "⏳", "🟩🟩⬜⬜⬜⬜⬜⬜⬜⬜ 20%"])
    ws.append(["Deploy", "❌", "⬜⬜⬜⬜⬜⬜⬜⬜⬜⬜ 0%"])
    ws.append(["Docs", "🔄", "🟩🟩🟩🟩🟩🟩🟩🟩⬜⬜ 80%"])
    save(wb, "classic166_emoji_progress.xlsx")


# ── 167. Musical notation and special Unicode symbols ───────────────────
def classic167_musical_symbols():
    wb = Workbook()
    ws = wb.active
    ws.title = "Symbols"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.append(["Category", "Symbols"])
    ws.append(["Music", "♩ ♪ ♫ ♬ 🎵 🎶"])
    ws.append(["Chess", "♔♕♖♗♘♙"])
    ws.append(["Zodiac", "♈♉♊♋♌♍♎♏♐♑♒♓"])
    ws.append(["Dice", "⚀ ⚁ ⚂ ⚃ ⚄ ⚅"])
    ws.append(["Weather", "☀ ☁ ☂ ☃ ❄ ☔"])
    save(wb, "classic167_musical_symbols.xlsx")


# ── 168. Mixed LTR/RTL in styled table ─────────────────────────────────
def classic168_mixed_ltr_rtl_styled():
    wb = Workbook()
    ws = wb.active
    ws.title = "BiDi Table"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E86C1")
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(["Code", "Name", "Price"], start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
    data = [
        ("EN-001", "Programming Book", "$29.99"),
        ("FR-002", "Livre de code", "€25.00"),
        ("AR-003", "كتاب برمجة", "50 SAR"),
        ("HE-004", "ספר קוד", "₪120"),
    ]
    for r, row_data in enumerate(data, start=2):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
    save(wb, "classic168_mixed_ltr_rtl_styled.xlsx")


# ── 169. Korean full document — invoice style ──────────────────────────
def classic169_korean_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "거래명세서"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    header = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "거래명세서 (Transaction Statement)"
    c.font = header
    c.alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(["번호", "상품명", "수량", "단가", "금액"])
    for col in range(1, 6):
        ws.cell(row=3, column=col).font = Font(bold=True)
    data = [
        (1, "노트북 컴퓨터", 2, "₩1,200,000", "₩2,400,000"),
        (2, "무선 마우스", 5, "₩25,000", "₩125,000"),
        (3, "모니터 27인치", 2, "₩350,000", "₩700,000"),
        (4, "키보드 (기계식)", 3, "₩89,000", "₩267,000"),
        (5, "USB 허브", 10, "₩15,000", "₩150,000"),
    ]
    for row_data in data:
        ws.append(list(row_data))
    ws.append([])
    ws.append(["", "", "", "합계", "₩3,642,000"])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    save(wb, "classic169_korean_invoice.xlsx")


# ── 170. Emoji + colored backgrounds — dashboard KPIs ──────────────────
def classic170_emoji_dashboard():
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    green = PatternFill("solid", fgColor="27AE60")
    yellow = PatternFill("solid", fgColor="F1C40F")
    red = PatternFill("solid", fgColor="E74C3C")
    white_font = Font(bold=True, color="FFFFFF")
    ws.append(["", "Metric", "Value", "Target"])
    kpis = [
        ("🟢", "Revenue", "$1.2M", "$1.0M", green),
        ("🟡", "Pipeline", "$800K", "$900K", yellow),
        ("🔴", "Churn Rate", "5.2%", "3.0%", red),
        ("🟢", "NPS Score", "72", "65", green),
        ("🟡", "Response Time", "2.1s", "1.5s", yellow),
        ("🟢", "Uptime", "99.95%", "99.9%", green),
    ]
    for r, (icon, metric, val, target, fill) in enumerate(kpis, start=2):
        ws.cell(row=r, column=1, value=icon)
        c_metric = ws.cell(row=r, column=2, value=metric)
        c_val = ws.cell(row=r, column=3, value=val)
        c_val.fill = fill
        c_val.font = white_font
        ws.cell(row=r, column=4, value=target)
    save(wb, "classic170_emoji_dashboard.xlsx")


# ── 171. IPA phonetic alphabet ──────────────────────────────────────────
def classic171_ipa_phonetic():
    wb = Workbook()
    ws = wb.active
    ws.title = "IPA"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.append(["Category", "IPA Symbols"])
    ws.append(["Plosives", "p b t d ʈ ɖ c ɟ k ɡ q ɢ ʔ"])
    ws.append(["Nasals", "m ɱ n ɳ ɲ ŋ ɴ"])
    ws.append(["Fricatives", "ɸ β f v θ ð s z ʃ ʒ ʂ ʐ ç ʝ x ɣ"])
    ws.append(["Vowels", "i y ɨ ʉ ɯ u e ø ɘ ɵ ɤ o ɛ œ ɜ ɞ ʌ ɔ æ a ɶ ɑ ɒ"])
    ws.append(["Tones", "˥ ˦ ˧ ˨ ˩ ˥˩ ˩˥"])
    ws.append(["Diacritics", "ʰ ʷ ʲ ˠ ˤ ⁿ ˡ"])
    ws.append(["Example word", "/ˌɪntəˈnæʃənəl/ (international)"])
    save(wb, "classic171_ipa_phonetic.xlsx")


# ── 172. Emoji timeline — project milestones ────────────────────────────
def classic172_emoji_timeline():
    wb = Workbook()
    ws = wb.active
    ws.title = "Timeline"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 10
    ws.append(["Date", "Icon", "Milestone", "Status"])
    data = [
        ("2025-01-15", "💡", "Idea conceived", "✅"),
        ("2025-02-01", "📋", "Requirements gathered", "✅"),
        ("2025-03-10", "🎨", "Design completed", "✅"),
        ("2025-04-20", "🔨", "Development started", "✅"),
        ("2025-06-15", "🧪", "Testing phase", "🔄"),
        ("2025-07-01", "🐛", "Bug fixing", "🔄"),
        ("2025-08-01", "🚀", "Launch day", "⏳"),
        ("2025-09-01", "📊", "Post-launch review", "⏳"),
    ]
    for row in data:
        ws.append(list(row))
    save(wb, "classic172_emoji_timeline.xlsx")


# ── 173. African language sampler ───────────────────────────────────────
def classic173_african_languages():
    wb = Workbook()
    ws = wb.active
    ws.title = "African"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 22
    ws.append(["Language", "Greeting", "Region"])
    ws.append(["Swahili", "Habari! Karibu sana.", "East Africa"])
    ws.append(["Amharic", "ሰላም! እንኳን ደህና መጣህ.", "Ethiopia"])
    ws.append(["Yoruba", "Ẹ kú àárọ̀! Ẹ kú alẹ́!", "Nigeria"])
    ws.append(["Zulu", "Sawubona! Unjani?", "South Africa"])
    ws.append(["Hausa", "Sannu! Barka da zuwa.", "West Africa"])
    ws.append(["Igbo", "Nnọọ! Kedụ?", "Nigeria"])
    ws.append(["Tigrinya", "ሰላም! ከመይ ኣለኻ?", "Eritrea"])
    save(wb, "classic173_african_languages.xlsx")


# ── 174. Technical symbols and unit notations ───────────────────────────
def classic174_technical_symbols():
    wb = Workbook()
    ws = wb.active
    ws.title = "Technical"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.append(["Category", "Symbols / Examples"])
    ws.append(["SI Units", "kg · m · s · A · K · mol · cd"])
    ws.append(["Derived", "N · Pa · J · W · V · Ω · Hz"])
    ws.append(["Prefixes", "μ (micro) · m (milli) · k (kilo) · M (mega) · G (giga)"])
    ws.append(["Electrical", "Ω kΩ MΩ · μF nF pF · mH μH"])
    ws.append(["Temp", "100°C = 212°F = 373.15 K"])
    ws.append(["Copyright", "© 2025 Company™ — All Rights Reserved®"])
    ws.append(["Fractions", "½ ⅓ ¼ ⅕ ⅙ ⅛ ⅔ ¾ ⅘"])
    ws.append(["Roman nums", "Ⅰ Ⅱ Ⅲ Ⅳ Ⅴ Ⅵ Ⅶ Ⅷ Ⅸ Ⅹ Ⅺ Ⅻ"])
    save(wb, "classic174_technical_symbols.xlsx")


# ── 175. Multi-script product catalog with emoji ────────────────────────
def classic175_multiscript_catalog():
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    header_fill = PatternFill("solid", fgColor="1ABC9C")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(["#", "Product (EN)", "Product (Local)", "Price", "Icon"], start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    data = [
        (1, "Green Tea", "緑茶", "¥500", "🍵"),
        (2, "Kimchi", "김치", "₩3,000", "🥬"),
        (3, "Samosa", "समोसा", "₹50", "🥟"),
        (4, "Croissant", "Croissant", "€2.50", "🥐"),
        (5, "Taco", "Taco", "$3.99", "🌮"),
        (6, "Borscht", "Борщ", "₽250", "🍲"),
        (7, "Falafel", "فلافل", "₪15", "🧆"),
        (8, "Pad Thai", "ผัดไทย", "฿80", "🍜"),
    ]
    for row_data in data:
        ws.append(list(row_data))
    save(wb, "classic175_multiscript_catalog.xlsx")


# ── 176. Combining characters and stacked diacritics ────────────────────
def classic176_combining_characters():
    wb = Workbook()
    ws = wb.active
    ws.title = "Combining"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.append(["Type", "Examples"])
    ws.append(["Single combining", "é = e + \u0301   ñ = n + \u0303"])
    ws.append(["Double combining", "ệ = e + \u0323 + \u0302"])
    ws.append(["Vietnamese", "ắ ằ ẵ ẳ ặ ố ồ ỗ ổ ộ ứ ừ ữ ử ự"])
    ws.append(["Zalgo-like", "H\u0335\u0316\u0318e\u0337\u031d\u0323l\u0336\u0324l\u0334\u0325o\u0338\u032e"])
    ws.append(["Precomposed vs decomposed", "ü (precomposed) vs u\u0308 (decomposed)"])
    ws.append(["Hangul Jamo", "ㅎ ㅏ ㄴ ㄱ ㅡ ㄹ → 한글"])
    save(wb, "classic176_combining_characters.xlsx")


# ── 177. Emoji calendar — month-by-month events ────────────────────────
def classic177_emoji_calendar():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendar"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 28
    ws.append(["Month", "Emoji", "Event"])
    months = [
        ("January", "🎆", "New Year"),
        ("February", "💝", "Valentine's Day"),
        ("March", "🌸", "Spring Equinox"),
        ("April", "🐣", "Easter"),
        ("May", "👩", "Mother's Day"),
        ("June", "☀️", "Summer Solstice"),
        ("July", "🎆", "Independence Day"),
        ("August", "🏖️", "Vacation Season"),
        ("September", "📚", "Back to School"),
        ("October", "🎃", "Halloween"),
        ("November", "🦃", "Thanksgiving"),
        ("December", "🎄", "Christmas"),
    ]
    for row in months:
        ws.append(list(row))
    save(wb, "classic177_emoji_calendar.xlsx")


# ── 178. Georgian, Armenian, Ethiopic scripts ───────────────────────────
def classic178_caucasus_ethiopic():
    wb = Workbook()
    ws = wb.active
    ws.title = "Scripts"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.append(["Script", "Sample Text"])
    ws.append(["Georgian", "საქართველო არის ძველი ცივილიზაცია."])
    ws.append(["Armenian", "Հայաստանը հին քաղաքակրթություն ունի."])
    ws.append(["Ethiopic", "ኢትዮጵያ የጥንታዊ ሥልጣኔ ምድር ናት።"])
    ws.append(["Georgian mkhedruli", "ა ბ გ დ ე ვ ზ თ ი კ ლ მ ნ ო პ"])
    ws.append(["Armenian alphabet", "Ա Բ Գ Դ Ե Զ Է Ը Թ Ժ Ի Լ Խ Ծ Կ"])
    save(wb, "classic178_caucasus_ethiopic.xlsx")


# ── 179. All-emoji inventory with quantities ────────────────────────────
def classic179_emoji_inventory():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    header_fill = PatternFill("solid", fgColor="8E44AD")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(["Icon", "Item", "Stock", "Min", "Status"], start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    items = [
        ("📱", "Smartphone", 150, 50, "🟢 OK"),
        ("💻", "Laptop", 42, 30, "🟡 Low"),
        ("🖨️", "Printer", 8, 10, "🔴 Reorder"),
        ("🎧", "Headphones", 200, 40, "🟢 OK"),
        ("⌨️", "Keyboard", 75, 25, "🟢 OK"),
        ("🖱️", "Mouse", 18, 20, "🔴 Reorder"),
        ("📷", "Camera", 12, 10, "🟡 Low"),
        ("🔌", "Charger", 300, 100, "🟢 OK"),
        ("💾", "USB Drive", 5, 15, "🔴 Reorder"),
        ("🖥️", "Monitor", 35, 20, "🟢 OK"),
    ]
    for r, (icon, item, stock, min_qty, status) in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=icon)
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=stock)
        ws.cell(row=r, column=4, value=min_qty)
        ws.cell(row=r, column=5, value=status)
    save(wb, "classic179_emoji_inventory.xlsx")


# ── 180. Polyglot paragraph — long mixed-script text ────────────────────
def classic180_polyglot_paragraph():
    wb = Workbook()
    ws = wb.active
    ws.title = "Polyglot"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.append(["Language", "Greeting"])
    ws.append(["English", "The quick brown fox."])
    ws.append(["Japanese", "速い茶色の狐。"])
    ws.append(["Korean", "빠른 갈색 여우."])
    ws.append(["Russian", "Быстрая бурая лиса."])
    ws.append(["Greek", "Η γρήγορη αλεπού."])
    ws.append(["Thai", "สุนัขจิ้งจอกสีน้ำตาล"])
    ws.append(["Hindi", "तेज़ भूरी लोमड़ी"])
    ws.append(["Emoji", "🦊 ➡️ 🐕"])
    save(wb, "classic180_polyglot_paragraph.xlsx")


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    ensure_output_dir()
    print(f"Generating 180 classic .xlsx files in: {OUTPUT_DIR}\n")

    generators = [
        classic01_basic_table_with_headers,
        classic02_multiple_worksheets,
        classic03_empty_workbook,
        classic04_single_cell,
        classic05_wide_table,
        classic06_tall_table,
        classic07_numbers_only,
        classic08_mixed_text_and_numbers,
        classic09_long_text,
        classic10_special_xml_characters,
        classic11_sparse_rows,
        classic12_sparse_columns,
        classic13_date_strings,
        classic14_decimal_numbers,
        classic15_negative_numbers,
        classic16_percentage_strings,
        classic17_currency_strings,
        classic18_large_dataset,
        classic19_single_column_list,
        classic20_all_empty_cells,
        classic21_header_only,
        classic22_long_sheet_name,
        classic23_unicode_text,
        classic24_red_text,
        classic25_multiple_colors,
        classic26_inline_strings,
        classic27_single_row,
        classic28_duplicate_values,
        classic29_formula_results,
        classic30_mixed_empty_and_filled_sheets,
        classic31_bold_header_row,
        classic32_right_aligned_numbers,
        classic33_centered_text,
        classic34_explicit_column_widths,
        classic35_explicit_row_heights,
        classic36_merged_cells,
        classic37_freeze_panes,
        classic38_hyperlink_cell,
        classic39_financial_table,
        classic40_scientific_notation,
        classic41_integer_vs_float,
        classic42_boolean_values,
        classic43_inventory_report,
        classic44_employee_roster,
        classic45_sales_by_region,
        classic46_grade_book,
        classic47_time_series,
        classic48_survey_results,
        classic49_contact_list,
        classic50_budget_vs_actuals,
        classic51_product_catalog,
        classic52_pivot_summary,
        classic53_invoice,
        classic54_multi_level_header,
        classic55_error_values,
        classic56_alternating_row_colors,
        classic57_cjk_only,
        classic58_mixed_numeric_formats,
        classic59_multi_sheet_summary,
        classic60_large_wide_table,
        # 61-90: image cases
        classic61_product_card_with_image,
        classic62_company_logo_header,
        classic63_two_products_side_by_side,
        classic64_employee_directory_with_photo,
        classic65_inventory_with_product_photos,
        classic66_invoice_with_logo,
        classic67_real_estate_listing,
        classic68_restaurant_menu,
        classic69_image_only_sheet,
        classic70_product_catalog_with_images,
        classic71_multi_sheet_with_images,
        classic72_bar_chart_image_with_data,
        classic73_event_flyer_with_banner,
        classic74_dashboard_with_kpi_image,
        classic75_certificate_with_seal,
        classic76_product_image_grid,
        classic77_news_article_with_hero_image,
        classic78_small_icon_per_row,
        classic79_wide_panoramic_banner,
        classic80_portrait_tall_image,
        classic81_step_by_step_with_images,
        classic82_before_after_images,
        classic83_color_swatch_palette,
        classic84_travel_destination_cards,
        classic85_lab_results_with_image,
        classic86_software_screenshot_features,
        classic87_sports_results_with_logos,
        classic88_image_after_data,
        classic89_nutrition_label_with_image,
        classic90_project_status_with_milestones,
        # 91-120: chart cases
        classic91_simple_bar_chart,
        classic92_horizontal_bar_chart,
        classic93_line_chart,
        classic94_pie_chart,
        classic95_area_chart,
        classic96_scatter_chart,
        classic97_doughnut_chart,
        classic98_radar_chart,
        classic99_bubble_chart,
        classic100_stacked_bar_chart,
        classic101_percent_stacked_bar,
        classic102_line_chart_with_markers,
        classic103_pie_chart_with_labels,
        classic104_combo_bar_line_chart,
        classic105_3d_bar_chart,
        classic106_3d_pie_chart,
        classic107_multi_series_line,
        classic108_stacked_area_chart,
        classic109_scatter_with_trendline,
        classic110_chart_with_legend,
        classic111_chart_with_axis_labels,
        classic112_multiple_charts,
        classic113_chart_sheet,
        classic114_chart_large_dataset,
        classic115_chart_negative_values,
        classic116_percent_stacked_area,
        classic117_stock_ohlc_chart,
        classic118_bar_chart_custom_colors,
        classic119_dashboard_multi_charts,
        classic120_chart_with_date_axis,
        # 121-150: style, border & background cases
        classic121_thin_borders,
        classic122_thick_outer_thin_inner,
        classic123_dashed_borders,
        classic124_colored_borders,
        classic125_solid_fills,
        classic126_dark_header,
        classic127_font_styles,
        classic128_font_sizes,
        classic129_alignment_combos,
        classic130_wrap_and_indent,
        classic131_number_formats,
        classic132_striped_table,
        classic133_gradient_rows,
        classic134_heatmap,
        classic135_bottom_border_only,
        classic136_financial_report_styled,
        classic137_checkerboard,
        classic138_color_grid,
        classic139_pattern_fills,
        classic140_rotated_text,
        classic141_mixed_edge_borders,
        classic142_styled_invoice,
        classic143_colored_tabs,
        classic144_note_style_cells,
        classic145_status_badges,
        classic146_double_border_table,
        classic147_multi_sheet_styled,
        classic148_frozen_styled_grid,
        classic149_merged_styled_sections,
        classic150_kitchen_sink_styles,
        # 151-180: multilingual, marks & emoji cases
        classic151_multilingual_greetings,
        classic152_emoji_sampler,
        classic153_currency_symbols,
        classic154_math_symbols,
        classic155_diacritical_marks,
        classic156_rtl_bidi_text,
        classic157_cjk_extended,
        classic158_emoji_skin_tones,
        classic159_zwj_emoji,
        classic160_punctuation_marks,
        classic161_box_drawing,
        classic162_cjk_emoji_styled,
        classic163_cyrillic_alphabets,
        classic164_indic_scripts,
        classic165_southeast_asian,
        classic166_emoji_progress,
        classic167_musical_symbols,
        classic168_mixed_ltr_rtl_styled,
        classic169_korean_invoice,
        classic170_emoji_dashboard,
        classic171_ipa_phonetic,
        classic172_emoji_timeline,
        classic173_african_languages,
        classic174_technical_symbols,
        classic175_multiscript_catalog,
        classic176_combining_characters,
        classic177_emoji_calendar,
        classic178_caucasus_ethiopic,
        classic179_emoji_inventory,
        classic180_polyglot_paragraph,
    ]

    for gen in generators:
        gen()

    print(f"\nDone! {len(generators)} files generated.")


if __name__ == "__main__":
    main()
